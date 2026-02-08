from __future__ import annotations

import json
import os
import io
import re
from datetime import datetime, timedelta
from pathlib import Path
import secrets
from urllib.parse import quote

import openpyxl

from fastapi import FastAPI, Request, Depends, Form, UploadFile, File, BackgroundTasks, HTTPException, Query
from fastapi.responses import HTMLResponse, RedirectResponse, PlainTextResponse, JSONResponse, Response
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel
from sqlmodel import Session, select
from sqlalchemy import or_, and_, func, exists, delete

from db import create_db_and_tables, get_session, engine
from migrate import ensure_schema, start_multi_agent_backfill_background
from models import (
    User,
    Role,
    AnalysisBatch,
    ConversationAnalysis,
    Conversation,
    Message,
    TrainingTask,
    TrainingAttempt,
    TaskStatus,
    AgentBinding,
    UserThread,
    TrainingReflection,
    TrainingSimulation,
    TrainingSimulationMessage,
    AIAnalysisJob,
    DailyAISummaryReport,
    DailyAISummaryJob,
    DailyAISummaryShare,
    TagCategory,
    TagDefinition,
    ConversationTagHit,
    TagSuggestion,
    RejectedTagRule,
    ManualTagBinding,
    AssistantThread,
    AssistantMessage,
    AssistantJob,
)
from auth import (
    hash_password,
    verify_password,
    create_token,
    get_current_user,
    require_role,
    auth_settings,
    get_user_by_email,
    get_user_by_username,
)
from seed import seed_if_needed, ensure_default_admin, ensure_app_config, ensure_default_tag_categories
from prompts import (
    analysis_system_prompt,
    analysis_system_prompt_fixed_part,
    get_editable_qc_prompt,
)
from tag_reject import normalize_key

from ai_client import chat_completion, AIError, list_model_ids_sync, get_ai_settings
from notify import get_feishu_webhook_url, send_feishu_webhook
from daily_summary import DEFAULT_DAILY_SUMMARY_PROMPT, build_daily_input, generate_daily_summary, enqueue_daily_job, get_latest_daily_job
from tools.user_admin import validate_role_change
from tools.tag_management import update_tag_definition, normalize_standard_for_category
from marllen_assistant import (
    ASSISTANT_NAME as MARLLEN_ASSISTANT_NAME,
    activate_thread as activate_marllen_thread,
    archive_and_create_new_thread,
    get_or_create_active_thread,
    get_pending_job as get_marllen_pending_job,
    list_threads as list_marllen_threads,
    list_thread_messages as list_marllen_thread_messages,
    maybe_auto_title_thread as maybe_auto_title_marllen_thread,
)


app = FastAPI(title="å®¢æœè´¨æ£€ä¸åŸ¹è®­(MVP)")

BASE_DIR = Path(__file__).resolve().parent
templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))
app.mount("/static", StaticFiles(directory=str(BASE_DIR / "static")), name="static")


@app.get("/favicon.ico")
def favicon():
    # Avoid noisy 404s in browser console when no favicon is configured yet.
    return Response(status_code=204)


@app.on_event("startup")
def on_startup():
    create_db_and_tables()
    ensure_schema()
    # Backfill historical per-message agent identity without blocking startup.
    start_multi_agent_backfill_background()
    # Ensure bootstrap admin exists (idempotent)
    with Session(engine) as session:
        ensure_default_admin(session)
        ensure_app_config(session)
        ensure_default_tag_categories(session)
        # Seed demo data when requested
        seed_if_needed(session)


def _redirect(url: str) -> RedirectResponse:
    return RedirectResponse(url=url, status_code=302)


def _template(request: Request, name: str, **ctx):
    return templates.TemplateResponse(name, {"request": request, **ctx})


def _can_view_conversation(user: User, session: Session, conv: Conversation) -> bool:
    """Conversation permissions.

    å½“å‰ç‰ˆæœ¬æŒ‰ä½ çš„éœ€æ±‚ï¼š
    - åªè¦å·²ç™»å½•ï¼Œå°±å…è®¸æŸ¥çœ‹æ‰€æœ‰å¯¹è¯ï¼ˆåŒ…æ‹¬ CID é¢„è§ˆå¼¹çª—ï¼‰ã€‚
    """
    return bool(user)


def _sanitize_next(next_url: str | None) -> str | None:
    """Prevent open-redirect. Only allow same-site relative paths like "/a?b=c"."""
    s = (next_url or "").strip()
    if not s:
        return None
    if not s.startswith("/"):
        return None
    # Disallow scheme-relative URLs like "//evil.com"
    if s.startswith("//"):
        return None
    return s


@app.exception_handler(HTTPException)
async def _http_exception_handler(request: Request, exc: HTTPException):
    """Redirect unauthenticated HTML requests to login, preserving the original URL."""
    if exc.status_code == 401:
        path = str(request.url.path or "")
        # Avoid redirect loops
        if path not in {"/login", "/register"}:
            accept = (request.headers.get("accept") or "").lower()
            # Only redirect for browser HTML navigations (avoid breaking JSON/API calls)
            if "text/html" in accept and not path.startswith("/api/"):
                nxt = path
                if request.url.query:
                    nxt = f"{nxt}?{request.url.query}"
                return RedirectResponse(url=f"/login?next={quote(nxt)}", status_code=302)

    # Default behavior
    return JSONResponse(status_code=exc.status_code, content={"detail": exc.detail})


@app.get("/", response_class=HTMLResponse)
def home(request: Request):
    token = request.cookies.get(auth_settings.COOKIE_NAME)
    if token:
        return _redirect("/conversations")
    return _redirect("/login")


@app.get("/me")
def me_redirect(user: User = Depends(get_current_user)):
    # All roles share the same primary landing page: conversations list.
    return _redirect("/conversations")


@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request, error: str | None = None, next: str | None = None):
    return _template(
        request,
        "login.html",
        error=error,
        next=_sanitize_next(next),
        remember_days=int(getattr(auth_settings, "REMEMBER_EXPIRE_DAYS", 7) or 7),
    )


@app.post("/login")
def login_action(
    request: Request,
    background_tasks: BackgroundTasks,
    username: str = Form(...),
    password: str = Form(...),
    remember: str | None = Form(None),
    next: str | None = Form(None),
    session: Session = Depends(get_session),
):
    user = get_user_by_username(session, username.strip())
    if not user or not verify_password(password, user.password_hash):
        return _template(
            request,
            "login.html",
            error="ç”¨æˆ·åæˆ–å¯†ç ä¸æ­£ç¡®",
            next=_sanitize_next(next) or _sanitize_next(request.query_params.get("next")),
            remember_days=int(getattr(auth_settings, "REMEMBER_EXPIRE_DAYS", 7) or 7),
        )

    remember_val = ((remember or "on") or "").strip().lower()
    remember_on = remember_val in {"on", "1", "true", "yes"}
    expire_hours = None
    if remember_on:
        expire_hours = int(getattr(auth_settings, "REMEMBER_EXPIRE_DAYS", 7) or 7) * 24
        expire_hours = max(1, expire_hours)

    token = create_token(user, expire_hours=expire_hours)
    target = _sanitize_next(next) or _sanitize_next(request.query_params.get("next")) or "/conversations"
    resp = _redirect(target)
    cookie_kwargs = dict(
        httponly=True,
        samesite="lax",
        secure=bool(getattr(auth_settings, "COOKIE_SECURE", False)),
    )
    if remember_on and expire_hours:
        max_age = int(expire_hours * 3600)
        cookie_kwargs.update(max_age=max_age, expires=max_age)
    resp.set_cookie(auth_settings.COOKIE_NAME, token, **cookie_kwargs)


    # Feishu login success notification (best-effort; never blocks login)
    try:
        url = get_feishu_webhook_url(session)
        if url:
            ip = getattr(getattr(request, "client", None), "host", "") or ""
            background_tasks.add_task(
                send_feishu_webhook,
                url,
                title="ç™»å½•æˆåŠŸ",
                text=f"è´¦å·ï¼š{user.username or ''}ï¼ˆ{user.name}ï¼‰\nè§’è‰²ï¼š{user.role}\nIPï¼š{ip}".strip(),
            )
    except Exception:
        pass
    return resp


@app.get("/logout")
def logout():
    resp = _redirect("/login")
    resp.delete_cookie(auth_settings.COOKIE_NAME)
    return resp


@app.get("/register", response_class=HTMLResponse)
def register_page(
    request: Request,
    error: str | None = None,
    session: Session = Depends(get_session),
):
    from app_config import get_open_registration
    if get_open_registration(session):
        return _template(request, "register.html", error=error)
    return _template(request, "error.html", message="å½“å‰å·²å…³é—­è‡ªåŠ©æ³¨å†Œï¼Œè¯·è”ç³»ç®¡ç†å‘˜åˆ›å»ºè´¦å·")


@app.post("/register")
def register_action(
    request: Request,
    username: str = Form(...),
    name: str = Form(...),
    email: str = Form(...),
    password: str = Form(...),
    session: Session = Depends(get_session),
):
    from app_config import get_open_registration
    if not get_open_registration(session):
        return _template(request, "error.html", message="å½“å‰å·²å…³é—­è‡ªåŠ©æ³¨å†Œï¼Œè¯·è”ç³»ç®¡ç†å‘˜åˆ›å»ºè´¦å·")

    # open registration -> agent role
    username_norm = username.strip()
    if not username_norm:
        return _template(request, "register.html", error="ç”¨æˆ·åä¸èƒ½ä¸ºç©º")
    existed = session.exec(select(User).where(User.username == username_norm)).first()
    if existed:
        return _template(request, "register.html", error="è¯¥ç”¨æˆ·åå·²å­˜åœ¨")

    email_norm = email.strip().lower()
    if get_user_by_email(session, email_norm):
        return _template(request, "register.html", error="è¯¥é‚®ç®±å·²æ³¨å†Œ")
    if len(password) < 6:
        return _template(request, "register.html", error="å¯†ç è‡³å°‘6ä½")

    u = User(username=username_norm, email=email_norm, name=name.strip(), role=Role.agent, password_hash=hash_password(password))
    session.add(u)
    session.commit()

    expire_hours = int(getattr(auth_settings, "REMEMBER_EXPIRE_DAYS", 7) or 7) * 24
    expire_hours = max(1, expire_hours)
    token = create_token(u, expire_hours=expire_hours)
    resp = _redirect("/conversations")
    max_age = int(expire_hours * 3600)
    resp.set_cookie(
        auth_settings.COOKIE_NAME,
        token,
        httponly=True,
        samesite="lax",
        secure=bool(getattr(auth_settings, "COOKIE_SECURE", False)),
        max_age=max_age,
        expires=max_age,
    )
    return resp


@app.get("/dashboard", response_class=HTMLResponse)
def dashboard(
    request: Request,
    user: User = Depends(require_role("admin", "supervisor")),
    session: Session = Depends(get_session),
):
    from app_config import get_app_config
    from models import BucketObject, AIAnalysisJob

    cfg = get_app_config(session)

    # latest bucket sync rows
    bucket_rows = session.exec(
        select(BucketObject).order_by(BucketObject.id.desc()).limit(200)
    ).all()

    # job stats
    pending = session.exec(select(AIAnalysisJob).where(AIAnalysisJob.status == "pending")).all()
    running = session.exec(select(AIAnalysisJob).where(AIAnalysisJob.status == "running")).all()
    error = session.exec(select(AIAnalysisJob).where(AIAnalysisJob.status == "error")).all()

    from app_config import get_bucket_config_dict
    b = get_bucket_config_dict("TAOBAO", session)
    return _template(
        request,
        "dashboard.html",
        user=user,
        cfg=cfg,
        bucket_rows=bucket_rows,
        job_stats={
            "pending": len(pending),
            "running": len(running),
            "error": len(error),
        },
        bucket_name=b.get("bucket") or "",
        bucket_prefix=b.get("prefix") or "",
        cron_interval=int(getattr(cfg, "cron_interval_seconds", 600) or 600),
        worker_poll=int(getattr(cfg, "worker_poll_seconds", 5) or 5),
    )


@app.get("/settings/ai", response_class=HTMLResponse)
def ai_settings_page(
    request: Request,
    user: User = Depends(require_role("admin", "supervisor")),
    session: Session = Depends(get_session),
):
    from app_config import get_app_config

    cfg = get_app_config(session)

    ai = get_ai_settings(session)
    from app_config import get_bucket_config_dict
    b = get_bucket_config_dict("TAOBAO", session)
    return _template(
        request,
        "ai_settings.html",
        user=user,
        active_tab="ai",
        cfg=cfg,
        system_prompt=analysis_system_prompt(session),
        qc_editable_prompt=get_editable_qc_prompt(session),
        qc_fixed_prompt=analysis_system_prompt_fixed_part(session),
        base_url=ai.get("base_url") or "",
        model=ai.get("model") or "",
        timeout=str(int(ai.get("timeout_s") or 1800)),
        openai_api_key=(cfg.openai_api_key or "")[:8] + "..." if (cfg.openai_api_key or "").strip() else "",
        openai_fallback_model=getattr(cfg, "openai_fallback_model", None) or "",
        openai_retries=cfg.openai_retries if getattr(cfg, "openai_retries", None) is not None else 0,
        openai_reasoning_effort=(getattr(cfg, "openai_reasoning_effort", None) or "high").strip(),
        bucket_name=b.get("bucket") or "",
        bucket_prefix=b.get("prefix") or "",
    )


@app.post("/settings/ai")
def ai_settings_update(
    request: Request,
    auto_analysis_enabled: str = Form("off"),
    qc_system_prompt: str = Form(""),
    openai_base_url: str = Form(""),
    openai_api_key: str = Form(""),
    openai_model: str = Form(""),
    openai_fallback_model: str = Form(""),
    openai_timeout: str = Form(""),
    openai_retries: str = Form(""),
    openai_reasoning_effort: str = Form(""),
    user: User = Depends(require_role("admin", "supervisor")),
    session: Session = Depends(get_session),
):
    from app_config import get_app_config, set_auto_analysis, set_ai_settings

    enabled = auto_analysis_enabled == "on"
    set_auto_analysis(session, enabled)

    cfg = get_app_config(session)
    cfg.qc_system_prompt = (qc_system_prompt or "").strip()
    cfg.updated_at = datetime.utcnow()
    session.add(cfg)
    session.commit()

    try:
        t = int((openai_timeout or "1800").strip() or "1800")
        t = max(60, t)
    except Exception:
        t = 1800
    try:
        r = int((openai_retries or "0").strip() or "0")
        r = max(0, r)
    except Exception:
        r = 0
    api_key_val = (openai_api_key or "").strip()
    if api_key_val == "****" or not api_key_val:
        api_key_val = None  # leave unchanged
    set_ai_settings(
        session,
        openai_base_url=(openai_base_url or "").strip() or None,
        openai_api_key=api_key_val,
        openai_model=(openai_model or "").strip() or None,
        openai_fallback_model=(openai_fallback_model or "").strip() or None,
        openai_timeout=t,
        openai_retries=r,
        openai_reasoning_effort=(openai_reasoning_effort or "").strip() or None,
    )

    return _redirect("/settings/ai")


def _recent_daily_reports(session: Session, *, limit: int = 60) -> list[DailyAISummaryReport]:
    # Latest saved reports for quick navigation in UI.
    try:
        return session.exec(
            select(DailyAISummaryReport).order_by(DailyAISummaryReport.run_date.desc()).limit(int(limit))
        ).all()
    except Exception:
        return []



def _today_shanghai_str() -> str:
    # We treat the product's "daily" as Shanghai business date.
    return (datetime.utcnow() + timedelta(hours=8)).strftime("%Y-%m-%d")


@app.get("/settings/ai/daily", response_class=HTMLResponse)
def daily_ai_summary_page(
    request: Request,
    user: User = Depends(require_role("admin", "supervisor")),
    session: Session = Depends(get_session),
    date: str | None = None,
    saved: int | None = None,
    saved_defaults: int | None = None,
    queued: int | None = None,
):
    from app_config import get_app_config

    cfg = get_app_config(session)
    env_default_model = (get_ai_settings(session).get("model") or "gpt-5.1").strip() or "gpt-5.1"
    model_ids = list_model_ids_sync(session=session)
    run_date = (date or "").strip() or _today_shanghai_str()

    report = session.exec(select(DailyAISummaryReport).where(DailyAISummaryReport.run_date == run_date)).first()

    job = get_latest_daily_job(session, run_date=run_date)

    prompt = (cfg.daily_summary_prompt or "").strip() or DEFAULT_DAILY_SUMMARY_PROMPT
    threshold_messages = int(cfg.daily_summary_threshold or 8)

    return _template(
        request,
        "ai_daily_summary.html",
        user=user,
        active_tab="daily",
        cfg=cfg,
        model_ids=model_ids,
        env_default_model=env_default_model,
        recent_reports=_recent_daily_reports(session),
        run_date=run_date,
        threshold_messages=threshold_messages,
        prompt=prompt,
        estimate=None,
        report=report,
        job=job,
        saved=saved,
        saved_defaults=saved_defaults,
        queued=queued,
        error=None,
    )


@app.post("/settings/ai/daily", response_class=HTMLResponse)
async def daily_ai_summary_action(
    request: Request,
    user: User = Depends(require_role("admin", "supervisor")),
    session: Session = Depends(get_session),
    action: str = Form("estimate"),
    run_date: str = Form(...),
    threshold_messages: int = Form(8),
    prompt: str = Form(""),
    model: str = Form(""),
):
    from app_config import get_app_config

    cfg = get_app_config(session)
    run_date = (run_date or "").strip() or _today_shanghai_str()
    threshold_messages = max(1, int(threshold_messages or 8))
    prompt = (prompt or "").strip() or DEFAULT_DAILY_SUMMARY_PROMPT
    env_default_model = (get_ai_settings(session).get("model") or "gpt-5.1").strip() or "gpt-5.1"
    model = (model or "").strip() or (cfg.daily_summary_model or "").strip() or env_default_model

    # Save defaults (no estimate/generation)
    if action == "save_defaults":
        cfg.daily_summary_threshold = threshold_messages
        cfg.daily_summary_model = (model or "").strip() or env_default_model
        cfg.daily_summary_prompt = prompt
        cfg.updated_at = datetime.utcnow()
        session.add(cfg)
        session.commit()
        return _redirect(f"/settings/ai/daily?date={run_date}&saved_defaults=1")

    # Always compute estimate first (for UI + safety)
    estimate = build_daily_input(session, run_date=run_date, threshold_messages=threshold_messages)

    # If user only asked for estimate, just render.
    if action != "generate":
        report = session.exec(select(DailyAISummaryReport).where(DailyAISummaryReport.run_date == run_date)).first()
        return _template(
            request,
            "ai_daily_summary.html",
            user=user,
            active_tab="daily",
            cfg=cfg,
            recent_reports=_recent_daily_reports(session),
            model_ids=list_model_ids_sync(session=session),
            env_default_model=env_default_model,
            run_date=run_date,
            threshold_messages=threshold_messages,
            prompt=prompt,
            estimate={
                "input_chars": estimate.input_chars,
                "included_conversations": estimate.included_conversations,
                "included_messages": estimate.included_messages,
            },
            report=report,
            job=get_latest_daily_job(session, run_date=run_date),
            saved=None,
            saved_defaults=None,
            queued=None,
            error=None,
        )

    # Generate (overwrite)
    try:
        # Persist defaults opportunistically (so next time it feels consistent)
        cfg.daily_summary_threshold = threshold_messages
        cfg.daily_summary_prompt = prompt
        cfg.daily_summary_model = (model or "").strip() or (cfg.daily_summary_model or "").strip() or env_default_model
        cfg.updated_at = datetime.utcnow()
        session.add(cfg)
        session.commit()
        # åå°ä»»åŠ¡ï¼šç«‹å³å…¥é˜Ÿï¼Œé¿å…é¡µé¢ç­‰å¾…è¿‡ä¹…/åˆ·æ–°ä¸¢å¤±
        enqueue_daily_job(
            session,
            run_date=run_date,
            threshold_messages=threshold_messages,
            prompt=prompt,
            model=cfg.daily_summary_model,
            created_by_user_id=int(user.id or 0) if user.id else None,
            estimate={
                "input_chars": estimate.input_chars,
                "included_conversations": estimate.included_conversations,
                "included_messages": estimate.included_messages,
            },
        )
        return _redirect(f"/settings/ai/daily?date={run_date}&queued=1")
    except AIError as e:
        report = session.exec(select(DailyAISummaryReport).where(DailyAISummaryReport.run_date == run_date)).first()
        return _template(
            request,
            "ai_daily_summary.html",
            user=user,
            active_tab="daily",
            cfg=cfg,
            recent_reports=_recent_daily_reports(session),
            model_ids=list_model_ids_sync(session=session),
            env_default_model=env_default_model,
            run_date=run_date,
            threshold_messages=threshold_messages,
            prompt=prompt,
            estimate={
                "input_chars": estimate.input_chars,
                "included_conversations": estimate.included_conversations,
                "included_messages": estimate.included_messages,
            },
            report=report,
            job=get_latest_daily_job(session, run_date=run_date),
            saved=None,
            saved_defaults=None,
            queued=None,
            error=str(e),
        )



    except Exception as e:
        report = session.exec(select(DailyAISummaryReport).where(DailyAISummaryReport.run_date == run_date)).first()
        return _template(
            request,
            "ai_daily_summary.html",
            user=user,
            active_tab="daily",
            cfg=cfg,
            recent_reports=_recent_daily_reports(session),
            model_ids=list_model_ids_sync(session=session),
            env_default_model=env_default_model,
            run_date=run_date,
            threshold_messages=threshold_messages,
            prompt=prompt,
            estimate={
                "input_chars": estimate.input_chars,
                "included_conversations": estimate.included_conversations,
                "included_messages": estimate.included_messages,
            },
            report=report,
            saved=None,
            saved_defaults=None,
            error=f"ç”Ÿæˆå¤±è´¥ï¼š{type(e).__name__}: {str(e)[:300]}",
        )

@app.get("/settings/ai/daily/export")
def daily_ai_summary_export(
    request: Request,
    user: User = Depends(require_role("admin", "supervisor")),
    session: Session = Depends(get_session),
    date: str = "",
):
    run_date = (date or "").strip() or _today_shanghai_str()
    report = session.exec(select(DailyAISummaryReport).where(DailyAISummaryReport.run_date == run_date)).first()
    if not report:
        return PlainTextResponse("æœªæ‰¾åˆ°è¯¥æ—¥æœŸçš„æŠ¥å‘Šã€‚", status_code=404)

    txt = (report.report_text or "").strip() + "\n"
    headers = {
        "Content-Disposition": f'attachment; filename="daily_ai_summary_{run_date}.txt"'
    }
    return PlainTextResponse(txt, media_type="text/plain; charset=utf-8", headers=headers)




# ===== Daily AI summary reports (saved history, clickable by date) =====

@app.get("/reports/daily-ai", response_class=HTMLResponse)
def daily_ai_reports_list(
    request: Request,
    user: User = Depends(require_role("admin", "supervisor")),
    session: Session = Depends(get_session),
):
    rows = session.exec(select(DailyAISummaryReport).order_by(DailyAISummaryReport.run_date.desc())).all()
    return _template(
        request,
        "daily_ai_reports.html",
        user=user,
        reports=rows,
    )


@app.get("/reports/daily-ai/{run_date}", response_class=HTMLResponse)
def daily_ai_report_view(
    request: Request,
    run_date: str,
    user: User = Depends(require_role("admin", "supervisor")),
    session: Session = Depends(get_session),
):
    run_date = (run_date or "").strip()
    report = session.exec(select(DailyAISummaryReport).where(DailyAISummaryReport.run_date == run_date)).first()
    if not report:
        return PlainTextResponse("æœªæ‰¾åˆ°è¯¥æ—¥æœŸçš„æŠ¥å‘Šã€‚", status_code=404)

    from rendering import render_markdown_safe

    html_body = render_markdown_safe(report.report_text or "")
    return _template(
        request,
        "daily_ai_report_view.html",
        user=user,
        report=report,
        html_body=html_body,
    )


@app.get("/reports/daily-ai/{run_date}/export")
def daily_ai_report_export(
    request: Request,
    run_date: str,
    user: User = Depends(require_role("admin", "supervisor")),
    session: Session = Depends(get_session),
):
    run_date = (run_date or "").strip()
    report = session.exec(select(DailyAISummaryReport).where(DailyAISummaryReport.run_date == run_date)).first()
    if not report:
        return PlainTextResponse("æœªæ‰¾åˆ°è¯¥æ—¥æœŸçš„æŠ¥å‘Šã€‚", status_code=404)

    txt = (report.report_text or "").strip() + "\n"
    headers = {"Content-Disposition": f'attachment; filename="daily_ai_summary_{run_date}.txt"'}
    return PlainTextResponse(txt, media_type="text/plain; charset=utf-8", headers=headers)

# ===== Public share links for Daily AI reports =====

def _new_share_token() -> str:
    # URL-safe token; keep short but hard to guess
    return secrets.token_urlsafe(18)


@app.post("/api/reports/daily-ai/{run_date}/share")
def daily_ai_report_create_share(
    request: Request,
    run_date: str,
    user: User = Depends(require_role("admin", "supervisor")),
    session: Session = Depends(get_session),
):
    run_date = (run_date or "").strip()
    report = session.exec(select(DailyAISummaryReport).where(DailyAISummaryReport.run_date == run_date)).first()
    if not report:
        raise HTTPException(status_code=404, detail="æœªæ‰¾åˆ°è¯¥æ—¥æœŸçš„æŠ¥å‘Š")

    # Create a new share token each time (safer than reusing; supports rotating links)
    token = None
    for _ in range(8):
        t = _new_share_token()
        existed = session.exec(select(DailyAISummaryShare).where(DailyAISummaryShare.token == t)).first()
        if not existed:
            token = t
            break
    if not token:
        raise HTTPException(status_code=500, detail="ç”Ÿæˆåˆ†äº«é“¾æ¥å¤±è´¥ï¼Œè¯·é‡è¯•")

    row = DailyAISummaryShare(
        report_id=report.id,
        token=token,
        created_by_user_id=user.id,
    )
    session.add(row)
    session.commit()

    share_path = f"/share/daily-ai/{token}"
    absolute = str(request.base_url).rstrip("/") + share_path
    return {"url": absolute, "path": share_path}


@app.get("/share/daily-ai/{token}", response_class=HTMLResponse)
def daily_ai_report_shared_view(
    request: Request,
    token: str,
    session: Session = Depends(get_session),
):
    token = (token or "").strip()
    row = session.exec(
        select(DailyAISummaryShare).where(
            DailyAISummaryShare.token == token,
            DailyAISummaryShare.revoked_at.is_(None),
        )
    ).first()
    if not row:
        return _template(request, "error.html", user=None, message="è¯¥åˆ†äº«é“¾æ¥ä¸å­˜åœ¨æˆ–å·²å¤±æ•ˆ")

    report = session.get(DailyAISummaryReport, row.report_id)
    if not report:
        return _template(request, "error.html", user=None, message="è¯¥æŠ¥å‘Šå·²ä¸å­˜åœ¨")

    from rendering import render_markdown_safe
    html_body = render_markdown_safe(report.report_text or "")
    return _template(
        request,
        "daily_ai_report_shared_view.html",
        user=None,
        report=report,
        html_body=html_body,
        share_token=token,
    )


@app.get("/settings/general", response_class=HTMLResponse)
def settings_general_page(
    request: Request,
    user: User = Depends(require_role("admin", "supervisor")),
    session: Session = Depends(get_session),
    saved: int | None = None,
    export_ok: int | None = None,
    import_ok: int | None = None,
    import_error: str | None = None,
):
    from app_config import get_open_registration, get_app_config
    cfg = get_app_config(session)
    return _template(
        request,
        "settings_general.html",
        user=user,
        open_registration=get_open_registration(session),
        enable_enqueue_analysis=bool(getattr(cfg, "enable_enqueue_analysis", True)),
        cron_interval_seconds=int(getattr(cfg, "cron_interval_seconds", 600) or 600),
        worker_poll_seconds=int(getattr(cfg, "worker_poll_seconds", 5) or 5),
        saved=bool(saved),
        export_ok=bool(export_ok),
        import_ok=bool(import_ok),
        import_error=import_error or "",
    )


@app.post("/settings/general")
def settings_general_save(
    request: Request,
    action: str = Form("save"),
    open_registration: str | None = Form(None),
    enable_enqueue_analysis: str | None = Form(None),
    cron_interval_seconds: int = Form(600),
    worker_poll_seconds: int = Form(5),
    user: User = Depends(require_role("admin", "supervisor")),
    session: Session = Depends(get_session),
):
    from app_config import set_open_registration, get_app_config
    if (action or "").strip() == "save":
        set_open_registration(session, open_registration == "1")
        cfg = get_app_config(session)
        cfg.enable_enqueue_analysis = bool(enable_enqueue_analysis == "1")
        cfg.cron_interval_seconds = max(10, int(cron_interval_seconds or 600))
        cfg.worker_poll_seconds = max(1, int(worker_poll_seconds or 5))
        cfg.updated_at = datetime.utcnow()
        session.add(cfg)
        session.commit()
    return _redirect("/settings/general?saved=1")


@app.get("/settings/general/export")
def settings_general_export(
    user: User = Depends(require_role("admin", "supervisor")),
    session: Session = Depends(get_session),
):
    from app_config import export_settings_to_file
    export_settings_to_file(session)
    return _redirect("/settings/general?export_ok=1")


@app.post("/settings/general/import")
def settings_general_import(
    request: Request,
    settings_file: UploadFile = File(...),
    user: User = Depends(require_role("admin", "supervisor")),
    session: Session = Depends(get_session),
):
    import time
    from pathlib import Path
    from fastapi.responses import RedirectResponse
    from app_config import import_settings_from_file
    try:
        raw = settings_file.file.read()
        text = raw.decode("utf-8") if isinstance(raw, bytes) else str(raw)
    except Exception as e:
        return RedirectResponse(url="/settings/general?import_error=" + quote(str(e)[:200]), status_code=302)
    path = Path("/tmp") / f"settings_import_{time.time()}.json"
    try:
        path.write_text(text, encoding="utf-8")
        import_settings_from_file(session, path)
    except Exception as e:
        return RedirectResponse(url="/settings/general?import_error=" + quote(str(e)[:200]), status_code=302)
    finally:
        try:
            path.unlink(missing_ok=True)
        except Exception:
            pass
    return _redirect("/settings/general?import_ok=1")


@app.get("/settings/bucket", response_class=HTMLResponse)
def bucket_settings_page(
    request: Request,
    user: User = Depends(require_role("admin", "supervisor")),
    session: Session = Depends(get_session),
    saved: int | None = None,
    run_result: str | None = None,
):
    from app_config import bucket_settings
    from models import CronState
    from sqlmodel import select
    from datetime import datetime, timedelta, timezone

    cfg = bucket_settings(session)

    st = session.exec(select(CronState).where(CronState.name == "bucket_poll_v1")).first()
    logs = (st.state or {}).get("logs") if st else []
    lines = []
    for it in (logs or []):
        try:
            lines.append(f"{it.get('ts','')} [{it.get('level','INFO')}] {it.get('msg','')}")
        except Exception:
            pass
    logs_text = "\n".join(lines)

    tz = timezone(timedelta(hours=8))
    default_date = (datetime.now(tz=tz) - timedelta(days=1)).date().isoformat()

    # ===== æŠ“å–/å¯¼å…¥ç¼ºå¤±çœ‹æ¿ =====
    # é»˜è®¤å±•ç¤ºæœ€è¿‘ 30 å¤©ï¼ˆåˆ°æ˜¨å¤©ä¸ºæ­¢ï¼‰
    qp = request.query_params
    try:
        days = int(qp.get("days") or "30")
    except Exception:
        days = 30
    days = max(7, min(120, days))

    from_str = (qp.get("from") or "").strip()
    to_str = (qp.get("to") or "").strip()

    try:
        if to_str:
            board_end = datetime.fromisoformat(to_str).date()
        else:
            board_end = (datetime.now(tz=tz) - timedelta(days=1)).date()
    except Exception:
        board_end = (datetime.now(tz=tz) - timedelta(days=1)).date()

    try:
        if from_str:
            board_start = datetime.fromisoformat(from_str).date()
        else:
            board_start = board_end - timedelta(days=days - 1)
    except Exception:
        board_start = board_end - timedelta(days=days - 1)

    if board_start > board_end:
        board_start, board_end = board_end, board_start

    board_dates = []
    cur = board_start
    while cur <= board_end:
        board_dates.append(cur.isoformat())
        cur += timedelta(days=1)

    # Always show both columns on the board.
    # Reason: a single bucket/prefix may contain mixed platforms (auto-detected at import time),
    # and we still want the admin board to reflect both, independent of which source toggles are ON.
    platforms = ["taobao", "douyin"]

    # 1) å…ˆè¯» ImportRunï¼ˆå¦‚æœæœ‰ï¼‰
    from models import ImportRun
    runs = session.exec(
        select(ImportRun).where(
            ImportRun.platform.in_(platforms),
            ImportRun.run_date >= board_start.isoformat(),
            ImportRun.run_date <= board_end.isoformat(),
        )
    ).all()
    run_map = {(r.platform, r.run_date): r for r in runs}

    # 2) å†ä» Conversation/Message æ¨æ–­ï¼ˆå…¼å®¹å†å²æ•°æ®ï¼‰
    from sqlalchemy import func

    # Use started_at first; fallback to uploaded_at
    date_expr = func.date(func.coalesce(Conversation.started_at, Conversation.uploaded_at))

    conv_counts = {}
    msg_counts = {}

    for p in platforms:
        rows = session.exec(
            select(date_expr, func.count(Conversation.id))
            .where(
                Conversation.platform == p,
                func.coalesce(Conversation.started_at, Conversation.uploaded_at) >= datetime.combine(board_start, datetime.min.time()),
                func.coalesce(Conversation.started_at, Conversation.uploaded_at) < datetime.combine(board_end + timedelta(days=1), datetime.min.time()),
            )
            .group_by(date_expr)
        ).all()
        for d, c in rows:
            if d is None:
                continue
            conv_counts[(p, str(d))] = int(c or 0)

        rows2 = session.exec(
            select(date_expr, func.count(Message.id))
            .join(Message, Message.conversation_id == Conversation.id)
            .where(
                Conversation.platform == p,
                func.coalesce(Conversation.started_at, Conversation.uploaded_at) >= datetime.combine(board_start, datetime.min.time()),
                func.coalesce(Conversation.started_at, Conversation.uploaded_at) < datetime.combine(board_end + timedelta(days=1), datetime.min.time()),
            )
            .group_by(date_expr)
        ).all()
        for d, c in rows2:
            if d is None:
                continue
            msg_counts[(p, str(d))] = int(c or 0)

    # 3) ç»„è£… cells
    board = {"dates": board_dates, "platforms": platforms, "cells": {}}
    for ds in board_dates:
        row = {}
        for p in platforms:
            r = run_map.get((p, ds))
            if r:
                det = r.details or {}
                row[p] = {
                    "status": r.status or "done",
                    "source": "run",
                    "conversations": int(det.get("conversations") or 0),
                    "messages": int(det.get("messages") or 0),
                }
            else:
                cc = conv_counts.get((p, ds), 0)
                mc = msg_counts.get((p, ds), 0)
                row[p] = {
                    "status": "done" if cc > 0 else "missing",
                    "source": "inferred",
                    "conversations": cc,
                    "messages": mc,
                }
        board["cells"][ds] = row

    return _template(
        request,
        "bucket_settings.html",
        user=user,
        s=cfg,
        saved=bool(saved),
        logs_text=logs_text,
        default_date=default_date,
        run_result=run_result or "",
        board=board,
        board_start=board_start.isoformat(),
        board_end=board_end.isoformat(),
        board_days=days,
    )


def _append_bucket_ui_log(session: Session, msg: str, level: str = "INFO") -> None:
    from models import CronState
    from sqlmodel import select
    from datetime import datetime, timezone, timedelta

    st = session.exec(select(CronState).where(CronState.name == "bucket_poll_v1")).first()
    if not st:
        st = CronState(name="bucket_poll_v1", state={})
        session.add(st)
        session.commit()
        session.refresh(st)

    s = dict(st.state or {})
    logs = list(s.get("logs") or [])
    tz = timezone(timedelta(hours=8))
    logs.append({"ts": datetime.now(tz=tz).isoformat(), "level": (level or "INFO").upper(), "msg": str(msg or "")})
    # respect config keep
    from app_config import get_app_config
    try:
        keep = int(get_app_config(session).bucket_log_keep or 800)
        keep = max(100, min(5000, keep))
    except Exception:
        keep = 800
    if len(logs) > keep:
        logs = logs[-keep:]
    s["logs"] = logs
    st.state = s
    st.updated_at = datetime.utcnow()
    session.add(st)
    session.commit()


@app.post("/settings/bucket")
def bucket_settings_update(
    request: Request,
    action: str = Form("save"),
    bucket_fetch_enabled: str | None = Form(None),
    taobao_bucket_import_enabled: str | None = Form(None),
    douyin_bucket_import_enabled: str | None = Form(None),
    bucket_daily_check_time: str = Form("10:15"),
    bucket_retry_interval_minutes: int = Form(60),
    bucket_log_keep: int = Form(800),
    bucket_backup_dir: str = Form(""),
    feishu_webhook_url: str = Form(""),
    taobao_bucket: str = Form(""),
    taobao_prefix: str = Form(""),
    taobao_endpoint: str = Form(""),
    taobao_region: str = Form(""),
    taobao_access_key: str = Form(""),
    taobao_secret_key: str = Form(""),
    douyin_bucket: str = Form(""),
    douyin_prefix: str = Form(""),
    douyin_endpoint: str = Form(""),
    douyin_region: str = Form(""),
    douyin_access_key: str = Form(""),
    douyin_secret_key: str = Form(""),
    date: str | None = Form(None),
    sources: list[str] = Form([]),
    user: User = Depends(require_role("admin", "supervisor")),
    session: Session = Depends(get_session),
):
    from app_config import set_bucket_settings
    from sqlmodel import select
    from models import CronState

    act = (action or "save").strip()

    if act == "save":
        taobao_cfg = {}
        if (taobao_bucket or "").strip():
            taobao_cfg["bucket"] = taobao_bucket.strip()
        if (taobao_prefix or "").strip():
            taobao_cfg["prefix"] = taobao_prefix.strip()
        if (taobao_endpoint or "").strip():
            taobao_cfg["endpoint"] = taobao_endpoint.strip()
        if (taobao_region or "").strip():
            taobao_cfg["region"] = taobao_region.strip()
        if (taobao_access_key or "").strip():
            taobao_cfg["access_key"] = taobao_access_key.strip()
        if (taobao_secret_key or "").strip():
            taobao_cfg["secret_key"] = taobao_secret_key.strip()
        
        douyin_cfg = {}
        if (douyin_bucket or "").strip():
            douyin_cfg["bucket"] = douyin_bucket.strip()
        if (douyin_prefix or "").strip():
            douyin_cfg["prefix"] = douyin_prefix.strip()
        if (douyin_endpoint or "").strip():
            douyin_cfg["endpoint"] = douyin_endpoint.strip()
        if (douyin_region or "").strip():
            douyin_cfg["region"] = douyin_region.strip()
        if (douyin_access_key or "").strip():
            douyin_cfg["access_key"] = douyin_access_key.strip()
        if (douyin_secret_key or "").strip():
            douyin_cfg["secret_key"] = douyin_secret_key.strip()

        set_bucket_settings(
            session,
            bucket_fetch_enabled=(bucket_fetch_enabled == "1"),
            taobao_bucket_import_enabled=(taobao_bucket_import_enabled == "1"),
            douyin_bucket_import_enabled=(douyin_bucket_import_enabled == "1"),
            bucket_daily_check_time=bucket_daily_check_time,
            bucket_retry_interval_minutes=bucket_retry_interval_minutes,
            bucket_log_keep=bucket_log_keep,
            bucket_backup_dir=(bucket_backup_dir or "").strip() or None,
            feishu_webhook_url=(feishu_webhook_url or "").strip() or None,
            taobao_bucket_config=taobao_cfg if taobao_cfg else None,
            douyin_bucket_config=douyin_cfg if douyin_cfg else None,
        )
        _append_bucket_ui_log(session, f"ç®¡ç†å‘˜ä¿å­˜æ¡¶é…ç½®ï¼šfetch={'ON' if bucket_fetch_enabled=='1' else 'OFF'} taobao={'ON' if taobao_bucket_import_enabled=='1' else 'OFF'} douyin={'ON' if douyin_bucket_import_enabled=='1' else 'OFF'} time={bucket_daily_check_time} retry={bucket_retry_interval_minutes}m")
        return _redirect("/settings/bucket?saved=1")

    if act == "clear_logs":
        st = session.exec(select(CronState).where(CronState.name == "bucket_poll_v1")).first()
        if st:
            st.state = {**(st.state or {}), "logs": []}
            session.add(st)
            session.commit()
        return _redirect("/settings/bucket?saved=1")

    if act == "run":
        from bucket_import import sync_bucket_once, get_bucket_config
        from datetime import datetime, timedelta, timezone
        from notify import send_feishu_webhook, get_feishu_webhook_url

        d = (date or "").strip()
        if not d:
            tz = timezone(timedelta(hours=8))
            d = (datetime.now(tz=tz) - timedelta(days=1)).date().isoformat()

        srcs = [str(x).strip().upper() for x in (sources or []) if str(x).strip()]
        # Require explicit selection to avoid accidental "douyinè¡¥æŠ“" also importing taobao.
        if not srcs:
            result_text = f"æ‰‹åŠ¨æŠ“å–æ—¥æœŸï¼š{d}\nè¯·é€‰æ‹©è¦æŠ“å–çš„å¹³å°ï¼ˆæ·˜å®/æŠ–éŸ³ï¼‰ã€‚"
            _append_bucket_ui_log(session, "æ‰‹åŠ¨æŠ“å–å–æ¶ˆï¼š" + result_text, level="WARN")
            from urllib.parse import quote
            return _redirect("/settings/bucket?run_result=" + quote(result_text))

        # Dedup (user may submit duplicated values).
        srcs = sorted(set(srcs))
        summaries = []
        all_unbound_accounts = []  # æ”¶é›†æ‰€æœ‰æœªç»‘å®šè´¦å·ä¿¡æ¯
        for src in srcs:
            src = (src or "").strip().upper()
            if src not in ("TAOBAO", "DOUYIN"):
                continue
            cfg = get_bucket_config(src, session)
            if not cfg.bucket:
                summaries.append(f"{src}: bucket æœªé…ç½®")
                continue
            # leyan structure: /leyan/YYYY-MM-DD/
            prefix = (cfg.prefix or "").strip("/")
            date_prefix = f"{prefix}/{d}/" if prefix else f"{d}/"
            platform = "taobao" if src == "TAOBAO" else "douyin"
            try:
                res = sync_bucket_once(session, cfg=cfg, prefix=date_prefix, platform=platform, imported_by_user_id=user.id, create_jobs_if_missing=False)
                seen = int(res.get("seen", 0) or 0)
                imported = int(res.get("imported", 0) or 0)
                errors = int(res.get("errors", 0) or 0)
                extra = ""
                byp = res.get("imported_files_by_platform") or {}
                if isinstance(byp, dict) and byp:
                    parts = []
                    for k in sorted(byp.keys()):
                        try:
                            parts.append(f"{k}={int(byp.get(k) or 0)}")
                        except Exception:
                            parts.append(f"{k}=?")
                    if parts:
                        extra = "ï¼ˆ" + ", ".join(parts) + "ï¼‰"
                summaries.append(f"{src}: å‘ç° {seen} ä¸ªæ–‡ä»¶ï¼Œå¯¼å…¥ {imported}ï¼Œå¤±è´¥ {errors}" + extra)
                
                # æ”¶é›†æœªç»‘å®šå®¢æœè´¦å·ä¿¡æ¯
                byp_unbound = res.get("unbound_agent_accounts_by_platform")
                byp_nicks = res.get("unbound_agent_nicks_by_platform") or {}
                if isinstance(byp_unbound, dict) and byp_unbound:
                    for plat, accs in byp_unbound.items():
                        plat_nicks = {}
                        if isinstance(byp_nicks, dict):
                            plat_nicks = (byp_nicks.get(plat) or {}) if isinstance(byp_nicks.get(plat), dict) else {}
                        for acc in (accs or []):
                            acc_s = str(acc).strip()
                            if not acc_s:
                                continue
                            nick = str(plat_nicks.get(acc_s, "") or "").strip()
                            all_unbound_accounts.append({"platform": plat or platform, "account": acc_s, "nick": nick})
                else:
                    unbound = res.get("unbound_agent_accounts") or []
                    unbound_nicks = res.get("unbound_agent_nicks") or {}
                    if unbound:
                        for acc in unbound:
                            acc_s = str(acc).strip()
                            if not acc_s:
                                continue
                            nick = str(unbound_nicks.get(acc_s, "") or "").strip()
                            all_unbound_accounts.append({"platform": platform, "account": acc_s, "nick": nick})
            except Exception as e:
                summaries.append(f"{src}: å¼‚å¸¸ {e}")

        result_text = f"æ‰‹åŠ¨æŠ“å–æ—¥æœŸï¼š{d}\n" + "\n".join(summaries)
        _append_bucket_ui_log(session, "æ‰‹åŠ¨æŠ“å–è§¦å‘ï¼š" + result_text)

        url = get_feishu_webhook_url(session)
        if url:
            try:
                send_feishu_webhook(url, title="ğŸ“¥ èŠå¤©è®°å½•æŠ“å–ï¼šæ‰‹åŠ¨è§¦å‘", text=result_text)
            except Exception:
                pass
        
        # å¦‚æœå‘ç°æœªç»‘å®šçš„å®¢æœè´¦å·ï¼Œå‘é€é€šçŸ¥
        if all_unbound_accounts and url:
            try:
                msg_lines = []
                for item in all_unbound_accounts:
                    acc = item["account"]
                    nick = item["nick"]
                    platform = item["platform"]
                    if nick:
                        msg_lines.append(f"â€¢ {acc} (æ˜µç§°: {nick}) - {platform}")
                    else:
                        msg_lines.append(f"â€¢ {acc} - {platform}")
                
                if msg_lines:
                    send_feishu_webhook(
                        url,
                        title="âš ï¸ å‘ç°æœªç»‘å®šå®¢æœè´¦å·",
                        text=f"æ—¥æœŸï¼š{d}\n\næœªç»‘å®šè´¦å·ï¼š\n" + "\n".join(msg_lines) + "\n\nè¯·åˆ°ã€è®¾ç½® > å®¢æœè´¦å·ç»‘å®šã€‘é¡µé¢è¿›è¡Œé…ç½®ã€‚"
                    )
            except Exception:
                pass

        from urllib.parse import quote
        return _redirect("/settings/bucket?run_result=" + quote(result_text))

    return _redirect("/settings/bucket")


# =========================
# æ ‡ç­¾ç³»ç»Ÿï¼šåˆ†ç±»/æ ‡ç­¾/åˆ¤å®šæ ‡å‡† + xlsx æ‰¹é‡å¯¼å…¥
# =========================


def _load_tag_tree(session: Session) -> tuple[list[TagCategory], dict[int, list[TagDefinition]]]:
    categories = session.exec(
        select(TagCategory).order_by(TagCategory.sort_order.asc(), TagCategory.id.asc())
    ).all()
    tags = session.exec(
        select(TagDefinition).order_by(TagDefinition.category_id.asc(), TagDefinition.sort_order.asc(), TagDefinition.id.asc())
    ).all()
    tags_by_cat: dict[int, list[TagDefinition]] = {}
    for t in tags:
        tags_by_cat.setdefault(int(t.category_id), []).append(t)
    return categories, tags_by_cat


def _normalize_header(v: str | None) -> str:
    """Normalize an xlsx header cell for robust matching.

    We keep only: a-z / 0-9 / ä¸­æ–‡ï¼Œé¿å…åƒ 'TagName(äºŒçº§æ ‡ç­¾)' è¿™ç§å¸¦æ‹¬å·çš„è¡¨å¤´åŒ¹é…å¤±è´¥ã€‚
    """
    raw = (v or "").strip().lower()
    raw = raw.replace(" ", "").replace("_", "")
    # remove punctuation/brackets/etc, keep only alnum + CJK
    raw = re.sub(r"[^0-9a-z\u4e00-\u9fff]+", "", raw)
    return raw


def _parse_tags_xlsx(file_bytes: bytes) -> list[dict[str, str]]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    ws = wb.active

    # Header row
    headers = []
    for c in range(1, ws.max_column + 1):
        headers.append(_normalize_header(str(ws.cell(1, c).value or "")))

    # Support both English and Chinese headers
    header_map = {
        "category": {"category", "categoryname", "ä¸€çº§åˆ†ç±»", "åˆ†ç±»", "ä¸€çº§"},
        "tag": {"tag", "tagname", "äºŒçº§æ ‡ç­¾", "æ ‡ç­¾", "äºŒçº§"},
        "standard": {"standard", "rule", "rubric", "åˆ¤å®šæ ‡å‡†", "å‘½ä¸­æ ‡å‡†", "aiæ ‡å‡†", "æ ‡å‡†"},
        "description": {"description", "desc", "è¯´æ˜", "æ–‡å­—è¯´æ˜", "å¤‡æ³¨"},
        "category_description": {"categorydescription", "åˆ†ç±»è¯´æ˜", "ä¸€çº§åˆ†ç±»è¯´æ˜"},
    }

    def _find_idx(keys: set[str]) -> int | None:
        norm_keys = {_normalize_header(k) for k in keys if k}
        for i, h in enumerate(headers):
            if not h:
                continue
            # exact match, prefix match, or contains match (supports 'TagName(äºŒçº§æ ‡ç­¾)' etc.)
            for k in norm_keys:
                if not k:
                    continue
                if h == k or h.startswith(k) or (k in h):
                    return i
        return None

    idx_category = _find_idx(header_map["category"])
    if idx_category is None:
        idx_category = 0
    idx_tag = _find_idx(header_map["tag"])
    if idx_tag is None:
        idx_tag = 1
    idx_standard = _find_idx(header_map["standard"])
    if idx_standard is None:
        idx_standard = 2
    idx_desc = _find_idx(header_map["description"])
    idx_cat_desc = _find_idx(header_map["category_description"])

    rows: list[dict[str, str]] = []
    for r in range(2, ws.max_row + 1):
        cat = str(ws.cell(r, idx_category + 1).value or "").strip()
        tag = str(ws.cell(r, idx_tag + 1).value or "").strip()
        std = str(ws.cell(r, idx_standard + 1).value or "").strip()
        desc = str(ws.cell(r, (idx_desc + 1)).value or "").strip() if idx_desc is not None else ""
        cat_desc = str(ws.cell(r, (idx_cat_desc + 1)).value or "").strip() if idx_cat_desc is not None else ""

        if not cat and not tag and not std and not desc:
            continue
        if not cat or not tag:
            # Skip incomplete lines
            continue
        rows.append(
            {
                "category": cat,
                "category_description": cat_desc,
                "tag": tag,
                "standard": std,
                "description": desc,
            }
        )

    return rows


def _tags_manage_url(*, cat: int | None = None, view: str | None = None, ok: str | None = None, error: str | None = None) -> str:
    """Build /settings/tags url with UI state preserved."""
    from urllib.parse import quote

    params: list[str] = []
    if cat is not None:
        params.append("cat=" + quote(str(int(cat))))
    if view:
        params.append("view=" + quote(str(view)))
    if ok:
        params.append("ok=" + quote(ok))
    if error:
        params.append("error=" + quote(error))
    return "/settings/tags" + ("?" + "&".join(params) if params else "")


@app.get("/settings/tags", response_class=HTMLResponse)
def tags_manage_page(
    request: Request,
    user: User = Depends(require_role("admin", "supervisor")),
    session: Session = Depends(get_session),
    cat: int | None = None,
    view: str | None = None,
    ok: str | None = None,
    error: str | None = None,
):
    categories, tags_by_cat = _load_tag_tree(session)

    # Page state
    view_mode = (view or "").strip() or "one"  # one | all
    selected_cat_id: int | None = None
    if cat is not None:
        try:
            selected_cat_id = int(cat)
        except Exception:
            selected_cat_id = None

    if view_mode != "all":
        # default to first active category, fallback to first category
        if selected_cat_id is None:
            first_active = next((c for c in categories if getattr(c, "is_active", True)), None)
            selected_cat_id = int(first_active.id) if first_active and first_active.id is not None else None
        if selected_cat_id is None and categories:
            selected_cat_id = int(categories[0].id)

    flash = ""
    flash_kind = ""
    if ok:
        flash = ok
        flash_kind = "ok"
    if error:
        flash = error
        flash_kind = "error"

    return _template(
        request,
        "tags_manage.html",
        user=user,
        categories=categories,
        tags_by_cat=tags_by_cat,
        selected_cat_id=selected_cat_id,
        view_mode=view_mode,
        flash=flash,
        flash_kind=flash_kind,
    )


@app.post("/settings/tags/category/reorder")
async def tags_category_reorder(
    request: Request,
    user: User = Depends(require_role("admin", "supervisor")),
    session: Session = Depends(get_session),
):
    try:
        payload = await request.json()
    except Exception:
        return JSONResponse(status_code=400, content={"ok": False, "error": "invalid_json"})

    ordered_ids = payload.get("ordered_ids") if isinstance(payload, dict) else None
    if not isinstance(ordered_ids, list) or not ordered_ids:
        return JSONResponse(status_code=400, content={"ok": False, "error": "invalid_order"})

    try:
        ids = [int(x) for x in ordered_ids]
    except Exception:
        return JSONResponse(status_code=400, content={"ok": False, "error": "invalid_id"})

    if len(set(ids)) != len(ids):
        return JSONResponse(status_code=400, content={"ok": False, "error": "duplicate_id"})

    existing = session.exec(select(TagCategory.id).where(TagCategory.id.in_(ids))).all()
    if len(existing) != len(ids):
        return JSONResponse(status_code=400, content={"ok": False, "error": "not_found"})

    now = datetime.utcnow()
    for idx, cid in enumerate(ids):
        c = session.get(TagCategory, cid)
        if not c:
            continue
        c.sort_order = (idx + 1) * 10
        c.updated_at = now
        session.add(c)
    session.commit()
    return JSONResponse(content={"ok": True})


@app.post("/settings/tags/tag/reorder")
async def tags_tag_reorder(
    request: Request,
    user: User = Depends(require_role("admin", "supervisor")),
    session: Session = Depends(get_session),
):
    try:
        payload = await request.json()
    except Exception:
        return JSONResponse(status_code=400, content={"ok": False, "error": "invalid_json"})

    if not isinstance(payload, dict):
        return JSONResponse(status_code=400, content={"ok": False, "error": "invalid_payload"})

    category_id = payload.get("category_id")
    ordered_ids = payload.get("ordered_ids")
    if not isinstance(ordered_ids, list) or not ordered_ids or not category_id:
        return JSONResponse(status_code=400, content={"ok": False, "error": "invalid_order"})

    try:
        cat_id = int(category_id)
        ids = [int(x) for x in ordered_ids]
    except Exception:
        return JSONResponse(status_code=400, content={"ok": False, "error": "invalid_id"})

    if len(set(ids)) != len(ids):
        return JSONResponse(status_code=400, content={"ok": False, "error": "duplicate_id"})

    existing = session.exec(
        select(TagDefinition.id).where(TagDefinition.category_id == cat_id, TagDefinition.id.in_(ids))
    ).all()
    if len(existing) != len(ids):
        return JSONResponse(status_code=400, content={"ok": False, "error": "not_found"})

    now = datetime.utcnow()
    for idx, tid in enumerate(ids):
        t = session.get(TagDefinition, tid)
        if not t:
            continue
        t.sort_order = (idx + 1) * 10
        t.updated_at = now
        session.add(t)
    session.commit()
    return JSONResponse(content={"ok": True})


@app.post("/settings/tags/category/create")
def tags_category_create(
    user: User = Depends(require_role("admin", "supervisor")),
    session: Session = Depends(get_session),
    name: str = Form(...),
    description: str = Form(""),
    ui_view: str = Form(""),
):
    name = name.strip()
    if not name:
        return _redirect(_tags_manage_url(view=ui_view, error="åˆ†ç±»åä¸èƒ½ä¸ºç©º"))
    existing = session.exec(select(TagCategory).where(TagCategory.name == name)).first()
    if existing:
        existing.description = description
        existing.is_active = True
        existing.updated_at = datetime.utcnow()
        session.add(existing)
    else:
        session.add(TagCategory(name=name, description=description, is_active=True, sort_order=0))
    session.commit()
    # jump to the new/updated category
    try:
        cid = int(existing.id if existing else session.exec(select(TagCategory).where(TagCategory.name == name)).first().id)  # type: ignore
    except Exception:
        cid = None
    return _redirect(_tags_manage_url(cat=cid, view=ui_view, ok="å·²ä¿å­˜"))


@app.post("/settings/tags/category/update")
def tags_category_update(
    user: User = Depends(require_role("admin", "supervisor")),
    session: Session = Depends(get_session),
    id: int = Form(...),
    name: str = Form(...),
    description: str = Form(""),
    is_active: str | None = Form(None),
    ui_view: str = Form(""),
):
    c = session.get(TagCategory, id)
    if not c:
        return _redirect(_tags_manage_url(cat=id, view=ui_view, error="åˆ†ç±»ä¸å­˜åœ¨"))
    c.name = name.strip() or c.name
    c.description = description
    c.is_active = bool(is_active)
    c.updated_at = datetime.utcnow()
    session.add(c)
    session.commit()
    return _redirect(_tags_manage_url(cat=int(c.id), view=ui_view, ok="å·²ä¿å­˜"))


@app.post("/settings/tags/category/delete")
def tags_category_delete(
    user: User = Depends(require_role("admin", "supervisor")),
    session: Session = Depends(get_session),
    id: int = Form(...),
    ui_view: str = Form(""),
):
    c = session.get(TagCategory, id)
    if not c:
        return _redirect(_tags_manage_url(cat=id, view=ui_view, error="åˆ†ç±»ä¸å­˜åœ¨"))
    c.is_active = False
    c.updated_at = datetime.utcnow()
    session.add(c)
    # Also disable child tags (soft)
    for t in session.exec(select(TagDefinition).where(TagDefinition.category_id == c.id)).all():
        t.is_active = False
        t.updated_at = datetime.utcnow()
        session.add(t)
    session.commit()
    # deleting category affects many tags, show full view for clarity
    return _redirect(_tags_manage_url(view="all", ok="å·²ä¿å­˜"))


@app.post("/settings/tags/category/remove")
def tags_category_remove(
    user: User = Depends(require_role("admin", "supervisor")),
    session: Session = Depends(get_session),
    id: int = Form(...),
    ui_view: str = Form(""),
):
    c = session.get(TagCategory, id)
    if not c:
        return _redirect(_tags_manage_url(cat=id, view=ui_view, error="åˆ†ç±»ä¸å­˜åœ¨"))

    # å½»åº•åˆ é™¤ï¼šåŒæ—¶åˆ é™¤å…¶ä¸‹æ‰€æœ‰æ ‡ç­¾ & å†å²å‘½ä¸­è®°å½•ï¼ˆä¸å¯æ¢å¤ï¼‰
    tag_ids = session.exec(select(TagDefinition.id).where(TagDefinition.category_id == int(c.id))).all()
    tag_ids = [int(x) for x in (tag_ids or []) if x is not None]

    deleted_hits = 0
    deleted_tags = 0

    if tag_ids:
        deleted_hits = int(session.exec(delete(ConversationTagHit).where(ConversationTagHit.tag_id.in_(tag_ids))).rowcount or 0)
        deleted_tags = int(session.exec(delete(TagDefinition).where(TagDefinition.id.in_(tag_ids))).rowcount or 0)

    session.exec(delete(TagCategory).where(TagCategory.id == int(c.id)))
    session.commit()

    return _redirect(_tags_manage_url(view="all", ok=f"å·²åˆ é™¤ï¼šåˆ†ç±»1ï¼Œæ ‡ç­¾{deleted_tags}ï¼Œå†å²å‘½ä¸­{deleted_hits}"))



@app.post("/settings/tags/tag/create")
def tags_tag_create(
    user: User = Depends(require_role("admin", "supervisor")),
    session: Session = Depends(get_session),
    category_id: int = Form(...),
    name: str = Form(...),
    standard: str = Form(""),
    description: str = Form(""),
    ui_view: str = Form(""),
):
    name = name.strip()
    if not name:
        return _redirect(_tags_manage_url(cat=category_id, view=ui_view, error="æ ‡ç­¾åä¸èƒ½ä¸ºç©º"))
    c = session.get(TagCategory, category_id)
    if not c:
        return _redirect(_tags_manage_url(cat=category_id, view=ui_view, error="æ‰€å±åˆ†ç±»ä¸å­˜åœ¨"))
    standard = normalize_standard_for_category(session, int(category_id), standard)
    existing = session.exec(
        select(TagDefinition).where(TagDefinition.category_id == category_id, TagDefinition.name == name)
    ).first()
    if existing:
        existing.standard = standard
        existing.description = description
        existing.is_active = True
        existing.updated_at = datetime.utcnow()
        session.add(existing)
    else:
        session.add(
            TagDefinition(
                category_id=category_id,
                name=name,
                standard=standard,
                description=description,
                is_active=True,
                sort_order=0,
            )
        )
    session.commit()
    return _redirect(_tags_manage_url(cat=int(category_id), view=ui_view, ok="å·²ä¿å­˜"))


@app.post("/settings/tags/tag/update")
def tags_tag_update(
    user: User = Depends(require_role("admin", "supervisor")),
    session: Session = Depends(get_session),
    id: int = Form(...),
    category_id: int | None = Form(None),
    name: str = Form(...),
    standard: str = Form(""),
    description: str = Form(""),
    is_active: str | None = Form(None),
    ui_cat: int | None = Form(None),
    ui_view: str = Form(""),
):
    try:
        t = update_tag_definition(
            session,
            tag_id=int(id),
            category_id=(int(category_id) if category_id is not None else None),
            name=name,
            standard=standard,
            description=description,
            is_active=bool(is_active),
        )
    except ValueError as e:
        return _redirect(_tags_manage_url(cat=ui_cat, view=ui_view, error=str(e)))

    session.commit()
    return _redirect(_tags_manage_url(cat=int(t.category_id), view=ui_view, ok="å·²ä¿å­˜"))


@app.post("/settings/tags/tag/delete")
def tags_tag_delete(
    user: User = Depends(require_role("admin", "supervisor")),
    session: Session = Depends(get_session),
    id: int = Form(...),
    ui_cat: int | None = Form(None),
    ui_view: str = Form(""),
):
    t = session.get(TagDefinition, id)
    if not t:
        return _redirect(_tags_manage_url(cat=ui_cat, view=ui_view, error="æ ‡ç­¾ä¸å­˜åœ¨"))
    t.is_active = False
    t.updated_at = datetime.utcnow()
    session.add(t)
    session.commit()
    return _redirect(_tags_manage_url(cat=int(t.category_id), view=ui_view, ok="å·²ä¿å­˜"))


@app.post("/settings/tags/tag/remove")
def tags_tag_remove(
    user: User = Depends(require_role("admin", "supervisor")),
    session: Session = Depends(get_session),
    id: int = Form(...),
    ui_cat: int | None = Form(None),
    ui_view: str = Form(""),
):
    t = session.get(TagDefinition, id)
    if not t:
        return _redirect(_tags_manage_url(cat=ui_cat, view=ui_view, error="æ ‡ç­¾ä¸å­˜åœ¨"))

    # å½»åº•åˆ é™¤ï¼šåŒæ—¶åˆ é™¤å†å²å‘½ä¸­è®°å½•ï¼ˆä¸å¯æ¢å¤ï¼‰
    deleted_hits = int(session.exec(delete(ConversationTagHit).where(ConversationTagHit.tag_id == int(t.id))).rowcount or 0)
    deleted_manual = int(session.exec(delete(ManualTagBinding).where(ManualTagBinding.tag_id == int(t.id))).rowcount or 0)
    session.exec(delete(TagDefinition).where(TagDefinition.id == int(t.id)))
    session.commit()

    return _redirect(_tags_manage_url(cat=int(t.category_id), view=ui_view, ok=f"å·²åˆ é™¤ï¼šæ ‡ç­¾1ï¼Œå†å²å‘½ä¸­{deleted_hits}ï¼Œæ‰‹åŠ¨æ ‡è®°{deleted_manual}"))


@app.post("/settings/tags/tag/merge")
def tags_tag_merge(
    user: User = Depends(require_role("admin", "supervisor")),
    session: Session = Depends(get_session),
    source_tag_id: int = Form(...),
    target_tag_id: int = Form(...),
    ui_cat: int | None = Form(None),
    ui_view: str = Form(""),
):
    """åˆå¹¶æ ‡ç­¾ï¼šå°† source_tag çš„æ‰€æœ‰å¯¹è¯æ ‡è®°è½¬ç§»åˆ° target_tagï¼Œç„¶ååˆ é™¤ source_tag"""
    
    # éªŒè¯ä¸¤ä¸ªæ ‡ç­¾æ˜¯å¦å­˜åœ¨
    source_tag = session.get(TagDefinition, source_tag_id)
    target_tag = session.get(TagDefinition, target_tag_id)
    
    if not source_tag:
        return _redirect(_tags_manage_url(cat=ui_cat, view=ui_view, error="æºæ ‡ç­¾ä¸å­˜åœ¨"))
    if not target_tag:
        return _redirect(_tags_manage_url(cat=ui_cat, view=ui_view, error="ç›®æ ‡æ ‡ç­¾ä¸å­˜åœ¨"))
    if source_tag_id == target_tag_id:
        return _redirect(_tags_manage_url(cat=ui_cat, view=ui_view, error="æºæ ‡ç­¾å’Œç›®æ ‡æ ‡ç­¾ä¸èƒ½ç›¸åŒ"))
    
    # 1. è¿ç§» ConversationTagHitï¼ˆAI è‡ªåŠ¨å‘½ä¸­ï¼‰
    # å…ˆæŸ¥æ‰¾æºæ ‡ç­¾çš„æ‰€æœ‰å‘½ä¸­è®°å½•
    source_hits = session.exec(
        select(ConversationTagHit).where(ConversationTagHit.tag_id == source_tag_id)
    ).all()
    
    migrated_hits = 0
    skipped_hits = 0
    
    for hit in source_hits:
        # æ£€æŸ¥ç›®æ ‡æ ‡ç­¾æ˜¯å¦å·²ç»æœ‰ç›¸åŒ analysis_id çš„è®°å½•
        existing = session.exec(
            select(ConversationTagHit).where(
                ConversationTagHit.analysis_id == hit.analysis_id,
                ConversationTagHit.tag_id == target_tag_id
            )
        ).first()
        
        if existing:
            # ç›®æ ‡æ ‡ç­¾å·²å­˜åœ¨è¯¥åˆ†æçš„å‘½ä¸­è®°å½•ï¼Œåˆ é™¤æºæ ‡ç­¾çš„è®°å½•
            session.delete(hit)
            skipped_hits += 1
        else:
            # è¿ç§»ï¼šä¿®æ”¹ tag_id
            hit.tag_id = target_tag_id
            migrated_hits += 1
    
    # 2. è¿ç§» ManualTagBindingï¼ˆæ‰‹åŠ¨æ ‡è®°ï¼‰
    source_bindings = session.exec(
        select(ManualTagBinding).where(ManualTagBinding.tag_id == source_tag_id)
    ).all()
    
    migrated_bindings = 0
    skipped_bindings = 0
    
    for binding in source_bindings:
        # æ£€æŸ¥ç›®æ ‡æ ‡ç­¾æ˜¯å¦å·²ç»æœ‰ç›¸åŒ conversation_id çš„æ‰‹åŠ¨æ ‡è®°
        existing = session.exec(
            select(ManualTagBinding).where(
                ManualTagBinding.conversation_id == binding.conversation_id,
                ManualTagBinding.tag_id == target_tag_id
            )
        ).first()
        
        if existing:
            # ç›®æ ‡æ ‡ç­¾å·²å­˜åœ¨è¯¥å¯¹è¯çš„æ‰‹åŠ¨æ ‡è®°ï¼Œåˆ é™¤æºæ ‡ç­¾çš„è®°å½•
            session.delete(binding)
            skipped_bindings += 1
        else:
            # è¿ç§»ï¼šä¿®æ”¹ tag_id
            binding.tag_id = target_tag_id
            migrated_bindings += 1
    
    # 3. åˆ é™¤æºæ ‡ç­¾
    session.delete(source_tag)
    session.commit()
    
    msg = f"å·²åˆå¹¶ï¼š{source_tag.name} â†’ {target_tag.name}ï¼›è¿ç§»å‘½ä¸­{migrated_hits}æ¡"
    if skipped_hits > 0:
        msg += f"ï¼ˆå»é‡{skipped_hits}æ¡ï¼‰"
    if migrated_bindings > 0:
        msg += f"ï¼Œè¿ç§»æ‰‹åŠ¨æ ‡è®°{migrated_bindings}æ¡"
    if skipped_bindings > 0:
        msg += f"ï¼ˆå»é‡{skipped_bindings}æ¡ï¼‰"
    
    return _redirect(_tags_manage_url(cat=int(target_tag.category_id), view=ui_view, ok=msg))


@app.get("/settings/tags/template.xlsx")
def tags_template_download(
    user: User = Depends(require_role("admin", "supervisor")),
):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "tags"
    ws.append(["CategoryName(ä¸€çº§åˆ†ç±»)", "CategoryDescription(åˆ†ç±»è¯´æ˜)", "TagName(äºŒçº§æ ‡ç­¾)", "Standard(åˆ¤å®šæ ‡å‡†/ç»™AI)", "Description(è¯´æ˜/ç»™äººçœ‹)"])
    ws.append(["å®¢æœæ¥å¾…", "", "æœªåŠæ—¶å›å¤", "å®¢æœè¶…è¿‡Xåˆ†é’Ÿæœªå›å¤/åå¤å‚¬ä¿ƒä»æ— åé¦ˆï¼Œåˆ™å‘½ä¸­", "ä¾‹ï¼šå®¢æˆ·å¤šæ¬¡è¿½é—®ï¼Œå®¢æœæ— æ˜ç¡®ç­”å¤"])
    ws.append(["å…¶ä»–æ ‡ç­¾", "", "æ— æ³•å½’ç±»", "æ— æ³•åŒ¹é…ç°æœ‰åˆ†ç±»ä¸”å¯¹å®¢æˆ·é€ æˆå½±å“æ—¶ï¼Œå¯æš‚å½’ä¸ºå…¶ä»–æ ‡ç­¾", ""])
    ws.append(["äº§å“è´¨é‡æŠ•è¯‰", "", "è‰²å·®/åšå·¥é—®é¢˜", "å‘½ä¸­åˆ¤å®šèŒƒå›´ï¼šç”¨æˆ·æ”¶åˆ°è´§ååé¦ˆï¼›å”®å‰/æœªæ”¶è´§ä»…å’¨è¯¢ä¸å‘½ä¸­ã€‚å‘½ä¸­æ ‡å‡†ï¼šå®¢æˆ·æ”¶åˆ°è´§åæ˜ç¡®åé¦ˆè‰²å·®ã€èµ·çƒã€å¼€çº¿ç­‰è´¨é‡é—®é¢˜ï¼Œåˆ™å‘½ä¸­ã€‚", ""])

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    filename = "tag_import_template.xlsx"
    headers = {
        "Content-Disposition": f"attachment; filename=\"{filename}\""
    }
    return Response(
        content=out.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@app.post("/settings/tags/import")
async def tags_import_xlsx(
    request: Request,
    user: User = Depends(require_role("admin", "supervisor")),
    session: Session = Depends(get_session),
    file: UploadFile = File(...),
    mode: str = Form("incremental"),
    duplicate: str = Form("overwrite"),
):
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        return _redirect(_tags_manage_url(view="all", error="è¯·ä¸Šä¼  xlsx æ–‡ä»¶"))
    b = await file.read()
    try:
        rows = _parse_tags_xlsx(b)
    except Exception:
        return _redirect(_tags_manage_url(view="all", error="xlsx è§£æå¤±è´¥ï¼Œè¯·ç”¨æ¨¡æ¿é‡æ–°å¯¼å‡º"))

    if not rows:
        return _redirect(_tags_manage_url(view="all", error="xlsx é‡Œæ²¡æœ‰å¯å¯¼å…¥æ•°æ®"))
    # è¦†ç›–å¯¼å…¥ï¼šä»¥è¡¨æ ¼ä¸ºå‡†ï¼ŒæŠŠâ€œæœªå‡ºç°åœ¨è¡¨æ ¼é‡Œçš„æ—§åˆ†ç±»/æ—§æ ‡ç­¾â€ç›´æ¥åˆ é™¤
    # æ³¨æ„ï¼šä¸ºäº†é¿å…äº§ç”Ÿæ‚¬ç©ºå¼•ç”¨ï¼Œè¿™é‡Œä¼šåŒæ—¶åˆ é™¤å¯¹åº”çš„å†å²å‘½ä¸­è®°å½•ï¼ˆConversationTagHitï¼‰
    deleted_categories = 0
    deleted_tags_total = 0
    deleted_hits_total = 0

    # å…ˆæŠŠ xlsx é‡Œå‡ºç°è¿‡çš„ (åˆ†ç±»å, æ ‡ç­¾å) / åˆ†ç±»å æ”¶é›†å‡ºæ¥
    desired_pairs: set[tuple[str, str]] = set()
    desired_categories: set[str] = set()
    for r in rows:
        cn = (r.get("category") or "").strip()
        tn = (r.get("tag") or "").strip()
        if cn:
            desired_categories.add(cn)
        if cn and tn:
            desired_pairs.add((cn, tn))

    if mode == "overwrite":
        # 1) åˆ é™¤ä¸åœ¨è¡¨æ ¼é‡Œçš„æ ‡ç­¾ï¼ˆä»¥åŠå®ƒä»¬çš„å†å²å‘½ä¸­ï¼‰
        existing = session.exec(
            select(TagDefinition.id, TagDefinition.name, TagCategory.name)
            .join(TagCategory, TagCategory.id == TagDefinition.category_id)
        ).all()

        tag_ids_to_delete: list[int] = []
        for tid, tname, cname in existing or []:
            cname2 = (cname or "").strip()
            tname2 = (tname or "").strip()
            if not cname2 or not tname2:
                continue
            if (cname2, tname2) not in desired_pairs:
                tag_ids_to_delete.append(int(tid))

        if tag_ids_to_delete:
            deleted_hits_total += int(session.exec(delete(ConversationTagHit).where(ConversationTagHit.tag_id.in_(tag_ids_to_delete))).rowcount or 0)
            deleted_tags_total += int(session.exec(delete(TagDefinition).where(TagDefinition.id.in_(tag_ids_to_delete))).rowcount or 0)
            session.commit()

        # 2) åˆ é™¤ä¸åœ¨è¡¨æ ¼é‡Œçš„åˆ†ç±»ï¼ˆä»¥åŠå…¶æ®‹ç•™çš„æ ‡ç­¾/å‘½ä¸­ï¼Œç†è®ºä¸Šä¸Šä¸€æ­¥å·²æ¸…å¹²å‡€ï¼Œè¿™é‡Œåšå…œåº•ï¼‰
        for c in session.exec(select(TagCategory)).all():
            if (c.name or "").strip() in desired_categories:
                continue
            ids = session.exec(select(TagDefinition.id).where(TagDefinition.category_id == int(c.id))).all()
            ids = [int(x) for x in (ids or []) if x is not None]
            if ids:
                deleted_hits_total += int(session.exec(delete(ConversationTagHit).where(ConversationTagHit.tag_id.in_(ids))).rowcount or 0)
                deleted_tags_total += int(session.exec(delete(TagDefinition).where(TagDefinition.id.in_(ids))).rowcount or 0)
            session.exec(delete(TagCategory).where(TagCategory.id == int(c.id)))
            deleted_categories += 1
        session.commit()

    # Upsert
    cat_by_name: dict[str, TagCategory] = {c.name: c for c in session.exec(select(TagCategory)).all()}
    imported = 0
    updated = 0
    skipped = 0

    for r in rows:
        cat_name = (r.get("category") or "").strip()
        tag_name = (r.get("tag") or "").strip()
        if not cat_name or not tag_name:
            continue

        c = cat_by_name.get(cat_name)
        if not c:
            c = TagCategory(name=cat_name, description=(r.get("category_description") or ""), is_active=True, sort_order=0)
            session.add(c)
            session.commit()
            session.refresh(c)
            cat_by_name[cat_name] = c
        else:
            # Optionally update category description when provided
            cat_desc = (r.get("category_description") or "").strip()
            if cat_desc:
                c.description = cat_desc
            c.is_active = True
            c.updated_at = datetime.utcnow()
            session.add(c)
            session.commit()

        existing = session.exec(
            select(TagDefinition).where(TagDefinition.category_id == int(c.id), TagDefinition.name == tag_name)
        ).first()
        normalized_std = normalize_standard_for_category(session, int(c.id), (r.get("standard") or ""))
        if existing:
            if duplicate in ("keep", "skip"):
                skipped += 1
                continue
            existing.standard = normalized_std
            existing.description = (r.get("description") or "")
            existing.is_active = True
            existing.updated_at = datetime.utcnow()
            session.add(existing)
            session.commit()
            updated += 1
        else:
            session.add(
                TagDefinition(
                    category_id=int(c.id),
                    name=tag_name,
                    standard=normalized_std,
                    description=(r.get("description") or ""),
                    is_active=True,
                    sort_order=0,
                )
            )
            session.commit()
            imported += 1


    # å¯¼å…¥åé»˜è®¤åˆ‡åˆ°"å…¨éƒ¨è§†å›¾"ï¼Œé¿å…ç”¨æˆ·è¯¯ä»¥ä¸ºæ²¡å¯¼å…¥æˆåŠŸ
    # é¢å¤–åšä¸€æ¬¡"æ•°æ®åº“å¯è§æ€§æ ¡éªŒ"ï¼šç»Ÿè®¡æœ¬æ¬¡ xlsx é‡Œå‡ºç°è¿‡çš„ (åˆ†ç±», æ ‡ç­¾) åœ¨åº“é‡Œèƒ½å¦æŸ¥åˆ°
    verify_tip = ""
    try:
        pairs = []
        for r in rows:
            cat_name = (r.get("category") or "").strip()
            tag_name = (r.get("tag") or "").strip()
            if cat_name and tag_name:
                pairs.append((cat_name, tag_name))
        pairs = list(dict.fromkeys(pairs))  # å»é‡ä¸”ä¿åº

        verified = 0
        if pairs:
            BATCH = 80
            for i in range(0, len(pairs), BATCH):
                batch = pairs[i:i + BATCH]
                conds = []
                for cn, tn in batch:
                    conds.append((TagCategory.name == cn) & (TagDefinition.name == tn))
                qv = (
                    select(func.count())
                    .select_from(TagDefinition)
                    .join(TagCategory, TagCategory.id == TagDefinition.category_id)
                    .where(or_(*conds))
                )
                verified += int(session.exec(qv).one() or 0)
        verify_tip = f"ï¼Œæ ¡éªŒå¯è§ {verified}/{len(pairs)}"
    except Exception:
        verify_tip = ""

    return _redirect(
        "/settings/tags?view=all&ok="
        + quote(
            f"å¯¼å…¥å®Œæˆï¼šè§£æ{len(rows)}è¡Œï¼Œæ–°å¢{imported}ï¼Œæ›´æ–°{updated}ï¼Œè·³è¿‡{skipped}"
            + (f"ï¼Œåˆ é™¤åˆ†ç±»{deleted_categories}ï¼Œåˆ é™¤æ ‡ç­¾{deleted_tags_total}ï¼Œåˆ é™¤å†å²å‘½ä¸­{deleted_hits_total}" if mode == "overwrite" else "")
            + verify_tip
        )
    )


@app.get("/settings/tag-suggestions", response_class=HTMLResponse)
def tag_suggestions_page(
    request: Request,
    user: User = Depends(require_role("admin", "supervisor")),
    session: Session = Depends(get_session),
    status_filter: str = "",
):
    status_norm = (status_filter or "").strip().lower()
    q = select(TagSuggestion).order_by(TagSuggestion.created_at.desc())
    if status_norm in ("pending", "approved", "rejected"):
        q = q.where(TagSuggestion.status == status_norm)
    suggestions = session.exec(q).all()

    conv_ids = [s.conversation_id for s in suggestions]
    convos = {c.id: c for c in session.exec(select(Conversation).where(Conversation.id.in_(conv_ids))).all()} if conv_ids else {}
    pending_count = session.exec(select(func.count(TagSuggestion.id)).where(TagSuggestion.status == "pending")).one() or 0

    return _template(
        request,
        "tag_suggestions.html",
        user=user,
        suggestions=suggestions,
        convos=convos,
        pending_count=int(pending_count),
        status_filter=status_filter or "",
        ok=request.query_params.get("ok"),
        error=request.query_params.get("error"),
    )


@app.post("/settings/tag-suggestions/{sid:int}/approve")
async def approve_tag_suggestion(
    request: Request,
    sid: int,
    user: User = Depends(require_role("admin", "supervisor")),
    session: Session = Depends(get_session),
):
    form = await request.form()
    ts = session.get(TagSuggestion, sid)
    if not ts or ts.status != "pending":
        return _redirect("/settings/tag-suggestions?error=å·¥å•ä¸å­˜åœ¨æˆ–å·²å¤„ç†")

    cat_name = (form.get("suggested_category") or ts.suggested_category or "").strip()
    tag_name = (form.get("suggested_tag_name") or ts.suggested_tag_name or "").strip()
    standard = (form.get("suggested_standard") or ts.suggested_standard or "").strip()
    description = (form.get("suggested_description") or ts.suggested_description or "").strip()
    if not cat_name or not tag_name:
        return _redirect("/settings/tag-suggestions?error=ä¸€çº§åˆ†ç±»ä¸äºŒçº§æ ‡ç­¾åä¸èƒ½ä¸ºç©º")

    cat = session.exec(select(TagCategory).where(TagCategory.name == cat_name)).first()
    if not cat:
        cat = TagCategory(name=cat_name, description="")
        session.add(cat)
        session.commit()
        session.refresh(cat)
    standard = normalize_standard_for_category(session, int(cat.id), standard)

    existing = session.exec(
        select(TagDefinition).where(TagDefinition.category_id == cat.id, TagDefinition.name == tag_name)
    ).first()
    if existing:
        tag = existing
    else:
        tag = TagDefinition(
            category_id=cat.id,
            name=tag_name,
            standard=standard,
            description=description,
        )
        session.add(tag)
        session.commit()
        session.refresh(tag)

    ts.status = "approved"
    ts.reviewed_by_user_id = user.id
    ts.reviewed_at = datetime.utcnow()
    ts.review_notes = (form.get("review_notes") or "").strip()
    ts.created_tag_id = tag.id
    session.add(ts)

    binding = session.exec(
        select(ManualTagBinding).where(
            ManualTagBinding.conversation_id == ts.conversation_id,
            ManualTagBinding.tag_id == tag.id,
        )
    ).first()
    if not binding:
        session.add(
            ManualTagBinding(
                conversation_id=ts.conversation_id,
                tag_id=tag.id,
                created_by_user_id=user.id,
                reason=f"å®¡æ ¸é€šè¿‡æ–°æ ‡ç­¾å»ºè®® #{ts.id}",
            )
        )
    session.commit()
    return _redirect("/settings/tag-suggestions?ok=å·²é€šè¿‡å¹¶åˆ›å»ºæ ‡ç­¾ã€ç»‘å®šå¯¹è¯")


@app.post("/settings/tag-suggestions/{sid:int}/reject")
async def reject_tag_suggestion(
    request: Request,
    sid: int,
    user: User = Depends(require_role("admin", "supervisor")),
    session: Session = Depends(get_session),
):
    form = await request.form()
    ts = session.get(TagSuggestion, sid)
    if not ts or ts.status != "pending":
        return _redirect("/settings/tag-suggestions?error=å·¥å•ä¸å­˜åœ¨æˆ–å·²å¤„ç†")

    ts.status = "rejected"
    ts.reviewed_by_user_id = user.id
    ts.reviewed_at = datetime.utcnow()
    ts.review_notes = (form.get("review_notes") or "").strip()
    session.add(ts)
    _upsert_rejected_tag_rule(
        session=session,
        user=user,
        category=ts.suggested_category,
        tag_name=ts.suggested_tag_name,
        notes=ts.review_notes,
    )
    session.commit()
    return _redirect("/settings/tag-suggestions?ok=å·²é©³å›")


def _upsert_rejected_tag_rule(
    *,
    session: Session,
    user: User,
    category: str,
    tag_name: str,
    notes: str = "",
) -> None:
    cat = (category or "").strip()
    name = (tag_name or "").strip()
    norm_key = normalize_key(cat, name)
    if not norm_key:
        return
    now = datetime.utcnow()
    existing = session.exec(
        select(RejectedTagRule).where(RejectedTagRule.norm_key == norm_key)
    ).first()
    if existing:
        if cat:
            existing.category = cat
        if name:
            existing.tag_name = name
        if notes:
            existing.notes = notes
        existing.is_active = True
        existing.updated_at = now
        existing.updated_by_user_id = user.id
        session.add(existing)
        return
    session.add(
        RejectedTagRule(
            category=cat,
            tag_name=name,
            aliases="",
            notes=notes or "",
            norm_key=norm_key,
            is_active=True,
            created_at=now,
            updated_at=now,
            created_by_user_id=user.id,
            updated_by_user_id=user.id,
        )
    )


@app.get("/settings/rejected-tags", response_class=HTMLResponse)
def rejected_tags_page(
    request: Request,
    user: User = Depends(require_role("admin", "supervisor")),
    session: Session = Depends(get_session),
    status_filter: str = "",
    q: str = "",
):
    status_norm = (status_filter or "").strip().lower()
    query = select(RejectedTagRule).order_by(RejectedTagRule.updated_at.desc())
    if status_norm == "active":
        query = query.where(RejectedTagRule.is_active == True)  # noqa: E712
    elif status_norm == "inactive":
        query = query.where(RejectedTagRule.is_active == False)  # noqa: E712
    qv = (q or "").strip()
    if qv:
        like = f"%{qv}%"
        query = query.where(
            or_(
                RejectedTagRule.category.ilike(like),
                RejectedTagRule.tag_name.ilike(like),
                RejectedTagRule.aliases.ilike(like),
                RejectedTagRule.notes.ilike(like),
            )
        )
    rules = session.exec(query).all()

    return _template(
        request,
        "rejected_tags.html",
        user=user,
        rules=rules,
        status_filter=status_filter or "",
        q=q or "",
        ok=request.query_params.get("ok"),
        error=request.query_params.get("error"),
    )


@app.post("/settings/rejected-tags/create")
async def rejected_tags_create(
    request: Request,
    user: User = Depends(require_role("admin", "supervisor")),
    session: Session = Depends(get_session),
):
    form = await request.form()
    category = (form.get("category") or "").strip()
    tag_name = (form.get("tag_name") or "").strip()
    aliases = (form.get("aliases") or "").strip()
    notes = (form.get("notes") or "").strip()
    is_active = (form.get("is_active") or "on") == "on"
    norm_key = normalize_key(category, tag_name)
    if not norm_key:
        return _redirect("/settings/rejected-tags?error=åˆ†ç±»ä¸æ ‡ç­¾åä¸èƒ½ä¸ºç©º")

    existing = session.exec(
        select(RejectedTagRule).where(RejectedTagRule.norm_key == norm_key)
    ).first()
    if existing:
        return _redirect("/settings/rejected-tags?error=å·²å­˜åœ¨ç›¸åŒçš„é©³å›è§„åˆ™")

    now = datetime.utcnow()
    session.add(
        RejectedTagRule(
            category=category,
            tag_name=tag_name,
            aliases=aliases,
            notes=notes,
            norm_key=norm_key,
            is_active=is_active,
            created_at=now,
            updated_at=now,
            created_by_user_id=user.id,
            updated_by_user_id=user.id,
        )
    )
    session.commit()
    return _redirect("/settings/rejected-tags?ok=å·²æ·»åŠ ")


@app.post("/settings/rejected-tags/{rid:int}/update")
async def rejected_tags_update(
    request: Request,
    rid: int,
    user: User = Depends(require_role("admin", "supervisor")),
    session: Session = Depends(get_session),
):
    form = await request.form()
    rule = session.get(RejectedTagRule, rid)
    if not rule:
        return _redirect("/settings/rejected-tags?error=è®°å½•ä¸å­˜åœ¨")

    category = (form.get("category") or "").strip()
    tag_name = (form.get("tag_name") or "").strip()
    aliases = (form.get("aliases") or "").strip()
    notes = (form.get("notes") or "").strip()
    is_active = (form.get("is_active") or "") == "on"

    new_norm = normalize_key(category, tag_name)
    if not new_norm:
        return _redirect("/settings/rejected-tags?error=åˆ†ç±»ä¸æ ‡ç­¾åä¸èƒ½ä¸ºç©º")

    if new_norm != rule.norm_key:
        exists = session.exec(
            select(RejectedTagRule).where(RejectedTagRule.norm_key == new_norm)
        ).first()
        if exists:
            return _redirect("/settings/rejected-tags?error=å·²å­˜åœ¨ç›¸åŒçš„é©³å›è§„åˆ™")

    rule.category = category
    rule.tag_name = tag_name
    rule.aliases = aliases
    rule.notes = notes
    rule.is_active = is_active
    rule.norm_key = new_norm
    rule.updated_at = datetime.utcnow()
    rule.updated_by_user_id = user.id
    session.add(rule)
    session.commit()
    return _redirect("/settings/rejected-tags?ok=å·²æ›´æ–°")


@app.post("/settings/rejected-tags/{rid:int}/remove")
async def rejected_tags_remove(
    request: Request,
    rid: int,
    user: User = Depends(require_role("admin", "supervisor")),
    session: Session = Depends(get_session),
):
    rule = session.get(RejectedTagRule, rid)
    if not rule:
        return _redirect("/settings/rejected-tags?error=è®°å½•ä¸å­˜åœ¨")
    session.delete(rule)
    session.commit()
    return _redirect("/settings/rejected-tags?ok=å·²åˆ é™¤")


def _upsert_agent_binding_and_sync(*, session: Session, platform: str, agent_account: str, user_id: int, created_by_user_id: int) -> AgentBinding:
    """Create/update a (platform, agent_account) -> user binding and sync historical conversations.

    è§„åˆ™ï¼š
    - (platform, agent_account) å¿…é¡»å”¯ä¸€ï¼šä¸€ä¸ªå¹³å°å®¢æœè´¦å·åªèƒ½ç»‘å®šåˆ°ä¸€ä¸ªç«™å†…è´¦å·
    - åŒä¸€ç«™å†…è´¦å·åœ¨åŒä¸€å¹³å°å…è®¸ç»‘å®šå¤šä¸ªå®¢æœè´¦å·
    """

    existing_platform_account = session.exec(
        select(AgentBinding).where(AgentBinding.platform == platform, AgentBinding.agent_account == agent_account)
    ).first()

    if existing_platform_account:
        existing_platform_account.user_id = user_id
        existing_platform_account.created_by_user_id = created_by_user_id
        existing_platform_account.created_at = datetime.utcnow()
        session.add(existing_platform_account)
        target = existing_platform_account
    else:
        target = AgentBinding(
            platform=platform,
            agent_account=agent_account,
            user_id=user_id,
            created_by_user_id=created_by_user_id,
        )
        session.add(target)

    # åŒæ­¥æ›´æ–°å·²æœ‰å¯¹è¯çš„å½’å±ï¼ˆå†å²å¯¹è¯ä¹Ÿä¼šç«‹å³â€œè®¤é¢†â€åˆ°è¯¥ç«™å†…è´¦å·ï¼‰
    convos = session.exec(
        select(Conversation).where(Conversation.platform == platform, Conversation.agent_account == agent_account)
    ).all()
    for c in convos:
        c.agent_user_id = user_id
        session.add(c)

    return target


@app.get("/bindings", response_class=HTMLResponse)
def bindings_page(
    request: Request,
    user: User = Depends(require_role("admin", "supervisor")),
    session: Session = Depends(get_session),
    message: str | None = None,
    error: str | None = None,
    show_deleted: int | None = None,
):
    from sqlalchemy import func, exists, and_

    bindings = session.exec(select(AgentBinding).order_by(AgentBinding.created_at.desc()).limit(300)).all()
    users = session.exec(select(User).order_by(User.role.asc(), User.name.asc())).all()
    # å¯æ¥å¾…è´¦å·ï¼šagent + supervisorï¼ˆä¸»ç®¡ä¹Ÿå¯â€œæ¥å¾…/è®¤é¢†â€å®¢æˆ·ï¼‰
    agents = [u for u in users if u.role in (Role.agent, Role.supervisor) and getattr(u, 'is_active', True)]
    users_by_id = {u.id: u for u in users}

    # === æœªç»‘å®šå®¢æœè´¦å·ï¼ˆæ¥è‡ªèŠå¤©è®°å½• Conversation.agent_accountï¼‰===
    # åªå±•ç¤ºâ€œå½“å‰åœ¨èŠå¤©é‡Œå‡ºç°è¿‡ï¼Œä½†ç³»ç»Ÿé‡Œè¿˜æ²¡å»ºç«‹ AgentBinding çš„è´¦å·â€ã€‚
    unbound_rows = session.exec(
        select(
            Conversation.platform.label("platform"),
            Message.agent_account.label("agent_account"),
            func.count(func.distinct(Conversation.id)).label("convo_cnt"),
            func.max(func.coalesce(Message.ts, Conversation.uploaded_at)).label("last_seen"),
            func.max(func.nullif(Message.agent_nick, "")).label("agent_nick"),
        )
        .join(Conversation, Conversation.id == Message.conversation_id)
        .where(
            Message.sender == "agent",
            Message.agent_account.is_not(None),
            Message.agent_account != "",
            ~exists(
                select(1).where(
                    and_(
                        AgentBinding.platform == Conversation.platform,
                        AgentBinding.agent_account == Message.agent_account,
                    )
                )
            ),
        )
        .group_by(Conversation.platform, Message.agent_account)
        .order_by(func.count(func.distinct(Conversation.id)).desc(), func.max(func.coalesce(Message.ts, Conversation.uploaded_at)).desc())
        .limit(600)
    ).all()

    # ç»Ÿä¸€ç»“æ„ï¼Œç»™æ¨¡æ¿ç”¨
    unbound_accounts = [
        {
            "platform": r[0] or "unknown",
            "agent_account": r[1] or "",
            "convo_cnt": int(r[2] or 0),
            "last_seen": r[3],
            "agent_nick": (r[4] or "").strip(),
        }
        for r in (unbound_rows or [])
        if (r and (r[1] or "").strip())
    ]

    return _template(
        request,
        "bindings.html",
        user=user,
        bindings=bindings,
        agents=agents,
        users_by_id=users_by_id,
        unbound_accounts=unbound_accounts,
        message=message,
        error=error,
    )


@app.post("/bindings/create")
def bindings_create(
    request: Request,
    platform: str = Form(...),
    agent_account: str = Form(...),
    user_id: int = Form(...),
    user: User = Depends(require_role("admin", "supervisor")),
    session: Session = Depends(get_session),
):
    platform = (platform or "").strip().lower()
    agent_account = (agent_account or "").strip()
    if not platform:
        return bindings_page(request, user=user, session=session, error="å¹³å°ä¸èƒ½ä¸ºç©ºï¼ˆä¾‹å¦‚ taobao / douyinï¼‰")
    if not agent_account:
        return bindings_page(request, user=user, session=session, error="å®¢æœè´¦å·ä¸èƒ½ä¸ºç©º")

    _upsert_agent_binding_and_sync(session=session, platform=platform, agent_account=agent_account, user_id=user_id, created_by_user_id=user.id)
    session.commit()
    return _redirect("/bindings?message=å·²ä¿å­˜ç»‘å®šï¼Œå¹¶åŒæ­¥åˆ°å†å²å¯¹è¯")


@app.post("/bindings/quick_create_and_bind")
def bindings_quick_create_and_bind(
    request: Request,
    platform: str = Form(...),
    agent_account: str = Form(...),
    username: str = Form(...),
    name: str = Form(...),
    password: str = Form(...),
    email: str | None = Form(None),
    user: User = Depends(require_role("admin", "supervisor")),
    session: Session = Depends(get_session),
):
    platform = (platform or "").strip().lower()
    agent_account = (agent_account or "").strip()
    username_norm = (username or "").strip()
    name_norm = (name or "").strip()
    email_norm = (email or "").strip().lower() if email else ""

    if not platform:
        return bindings_page(request, user=user, session=session, error="å¹³å°ä¸èƒ½ä¸ºç©ºï¼ˆä¾‹å¦‚ taobao / douyinï¼‰")
    if not agent_account:
        return bindings_page(request, user=user, session=session, error="å®¢æœè´¦å·ä¸èƒ½ä¸ºç©º")
    if not username_norm:
        return bindings_page(request, user=user, session=session, error="ç™»å½•ç”¨æˆ·åä¸èƒ½ä¸ºç©º")
    if len(password or "") < 6:
        return bindings_page(request, user=user, session=session, error="åˆå§‹å¯†ç è‡³å°‘6ä½")

    existed = session.exec(select(User).where(User.username == username_norm)).first()
    if existed:
        return bindings_page(request, user=user, session=session, error="è¯¥ç”¨æˆ·åå·²å­˜åœ¨ï¼Œè¯·æ¢ä¸€ä¸ª")

    # email åœ¨ç³»ç»Ÿé‡Œä»æ˜¯å¿…å¡«ï¼šæ²¡å¡«å°±è‡ªåŠ¨ç”Ÿæˆä¸€ä¸ªï¼ˆä»…ç”¨äºå ä½ï¼Œä¸å½±å“æ—¥å¸¸ä½¿ç”¨ï¼‰
    if not email_norm:
        email_norm = f"{username_norm}@local.marllen"
    # é¿å… email å†²çª
    if get_user_by_email(session, email_norm):
        from datetime import datetime
        suffix = datetime.utcnow().strftime('%Y%m%d%H%M%S')
        email_norm = f"{username_norm}+{suffix}@local.marllen"

    u = User(
        username=username_norm,
        email=email_norm,
        name=name_norm or username_norm,
        role=Role.agent,
        password_hash=hash_password(password),
    )
    session.add(u)
    session.commit()  # å…ˆæ‹¿åˆ° u.id

    _upsert_agent_binding_and_sync(session=session, platform=platform, agent_account=agent_account, user_id=u.id, created_by_user_id=user.id)
    session.commit()
    return _redirect(f"/bindings?message=å·²åˆ›å»ºå­è´¦å·å¹¶å®Œæˆç»‘å®šï¼š{u.username} Â· {u.name}")


@app.post("/bindings/{binding_id}/delete")
def bindings_delete(
    binding_id: int,
    user: User = Depends(require_role("admin", "supervisor")),
    session: Session = Depends(get_session),
):
    b = session.get(AgentBinding, binding_id)
    if b:
        platform = b.platform
        agent_account = b.agent_account
        bound_user_id = b.user_id
        session.delete(b)

        # è§£ç»‘åï¼ŒæŠŠå·²å½’å±åˆ°è¯¥è´¦å·çš„å†å²å¯¹è¯é‡Šæ”¾å‡ºæ¥ï¼ˆé¿å…ç»§ç»­æ˜¾ç¤ºä¸ºâ€œå·²ç»‘å®šâ€ï¼‰
        convos = session.exec(
            select(Conversation).where(
                Conversation.platform == platform,
                Conversation.agent_account == agent_account,
                Conversation.agent_user_id == bound_user_id,
            )
        ).all()
        for c in convos:
            c.agent_user_id = None
            session.add(c)

        session.commit()
    return _redirect("/bindings?message=å·²åˆ é™¤ç»‘å®šï¼Œå¹¶å·²é‡Šæ”¾å†å²å¯¹è¯")



@app.get("/admin/users", response_class=HTMLResponse)
def admin_users_page(
    request: Request,
    user: User = Depends(require_role("admin")),
    session: Session = Depends(get_session),
    message: str | None = None,
    error: str | None = None,
    show_deleted: int | None = None,
):
    stmt = select(User).order_by(User.created_at.desc()).limit(500)
    if not show_deleted and hasattr(User, 'is_active'):
        stmt = stmt.where(User.is_active == True)
    users = session.exec(stmt).all()

    from sqlalchemy import func

    # è´¦å·ç»´åº¦ç»Ÿè®¡ï¼šå¯¹è¯æ•°ï¼ˆè¢«è®¤é¢†çš„ Conversationï¼‰ä¸å‘é€æ¶ˆæ¯æ•°ï¼ˆéä¹°å®¶æ¶ˆæ¯ï¼‰
    convo_cnt_by_user = {
        int(uid): int(cnt or 0)
        for uid, cnt in (
            session.exec(
                select(Conversation.agent_user_id, func.count(Conversation.id))
                .where(Conversation.agent_user_id.is_not(None))
                .group_by(Conversation.agent_user_id)
            ).all()
            or []
        )
        if uid is not None
    }

    sent_msg_cnt_by_user = {
        int(uid): int(cnt or 0)
        for uid, cnt in (
            session.exec(
                select(Conversation.agent_user_id, func.count(Message.id))
                .join(Conversation, Message.conversation_id == Conversation.id)
                .where(
                    Conversation.agent_user_id.is_not(None),
                    Message.sender != "buyer",
                )
                .group_by(Conversation.agent_user_id)
            ).all()
            or []
        )
        if uid is not None
    }

    return _template(
        request,
        "admin_users.html",
        user=user,
        users=users,
        convo_cnt_by_user=convo_cnt_by_user,
        sent_msg_cnt_by_user=sent_msg_cnt_by_user,
        message=message,
        error=error,
    )


@app.post("/admin/users/create")
def admin_users_create(
    request: Request,
    username: str = Form(...),
    name: str = Form(...),
    email: str = Form(...),
    password: str = Form(...),
    role: str = Form("agent"),
    user: User = Depends(require_role("admin")),
    session: Session = Depends(get_session),
):
    username_norm = username.strip()
    if not username_norm:
        return admin_users_page(request, user=user, session=session, error="ç”¨æˆ·åä¸èƒ½ä¸ºç©º")
    existed = session.exec(select(User).where(User.username == username_norm)).first()
    if existed:
        return admin_users_page(request, user=user, session=session, error="è¯¥ç”¨æˆ·åå·²å­˜åœ¨")

    email_norm = email.strip().lower()
    if get_user_by_email(session, email_norm):
        return admin_users_page(request, user=user, session=session, error="è¯¥é‚®ç®±å·²å­˜åœ¨")
    if len(password) < 6:
        return admin_users_page(request, user=user, session=session, error="å¯†ç è‡³å°‘6ä½")
    try:
        role_enum = Role(role)
    except Exception:
        role_enum = Role.agent

    u = User(
        username=username_norm,
        email=email_norm,
        name=name.strip(),
        role=role_enum,
        password_hash=hash_password(password),
    )
    session.add(u)
    session.commit()
    return _redirect("/admin/users?message=å·²åˆ›å»ºç”¨æˆ·")


@app.post("/admin/users/{user_id}/reset_password")
def admin_users_reset_password(
    user_id: int,
    password: str = Form(...),
    user: User = Depends(require_role("admin")),
    session: Session = Depends(get_session),
):
    if len(password) < 6:
        return _redirect("/admin/users?error=å¯†ç è‡³å°‘6ä½")
    u = session.get(User, user_id)
    if not u:
        return _redirect("/admin/users?error=æ‰¾ä¸åˆ°ç”¨æˆ·")
    u.password_hash = hash_password(password)
    session.add(u)
    session.commit()
    return _redirect("/admin/users?message=å·²é‡ç½®å¯†ç ")


@app.post("/admin/users/{user_id}/update")
def admin_users_update(
    user_id: int,
    username: str = Form(...),
    name: str = Form(...),
    email: str = Form(...),
    password: str = Form(""),
    role: str = Form(""),
    user: User = Depends(require_role("admin")),
    session: Session = Depends(get_session),
):
    u = session.get(User, user_id)
    if not u:
        return _redirect("/admin/users?error=æ‰¾ä¸åˆ°ç”¨æˆ·")
    if getattr(u, "is_active", True) is False:
        return _redirect("/admin/users?error=è¯¥è´¦å·å·²åˆ é™¤ï¼Œæ— æ³•ç¼–è¾‘")

    username_norm = (username or "").strip()
    if not username_norm:
        return _redirect("/admin/users?error=ç”¨æˆ·åä¸èƒ½ä¸ºç©º")

    email_norm = (email or "").strip().lower()
    if not email_norm:
        return _redirect("/admin/users?error=é‚®ç®±ä¸èƒ½ä¸ºç©º")

    # uniqueness checks (exclude self)
    existed_u = session.exec(select(User).where(User.username == username_norm, User.id != user_id)).first()
    if existed_u and getattr(existed_u, "is_active", True) is True:
        return _redirect("/admin/users?error=è¯¥ç”¨æˆ·åå·²å­˜åœ¨")
    existed_e = session.exec(select(User).where(User.email == email_norm, User.id != user_id)).first()
    if existed_e and getattr(existed_e, "is_active", True) is True:
        return _redirect("/admin/users?error=è¯¥é‚®ç®±å·²å­˜åœ¨")

    u.username = username_norm
    u.name = (name or "").strip() or u.name
    u.email = email_norm

    role_raw = (role or "").strip()
    if role_raw:
        try:
            new_role = Role(role_raw)
        except Exception:
            return _redirect("/admin/users?error=è§’è‰²éæ³•")

        err = validate_role_change(session=session, target_user=u, new_role=new_role, acting_user=user)
        if err:
            return _redirect(f"/admin/users?error={quote(err)}")
        u.role = new_role

    pwd = (password or "").strip()
    if pwd:
        if len(pwd) < 6:
            return _redirect("/admin/users?error=å¯†ç è‡³å°‘6ä½")
        u.password_hash = hash_password(pwd)

    session.add(u)
    session.commit()
    # If admin changed their own role to non-admin, avoid redirecting to admin-only page.
    role_val_after = getattr(u.role, "value", None) or str(u.role)
    if user.id == u.id and role_val_after != Role.admin.value:
        return _redirect("/conversations?message=å·²æ›´æ–°è´¦å·ä¿¡æ¯ï¼ˆå½“å‰è´¦å·æƒé™å·²å˜æ›´ï¼‰")
    return _redirect("/admin/users?message=å·²æ›´æ–°è´¦å·ä¿¡æ¯")


@app.post("/admin/users/{user_id}/delete")
def admin_users_delete(
    user_id: int,
    user: User = Depends(require_role("admin")),
    session: Session = Depends(get_session),
):
    # å®‰å…¨ï¼šä¸å…è®¸åˆ é™¤è‡ªå·±
    if user.id == user_id:
        return _redirect("/admin/users?error=ä¸èƒ½åˆ é™¤å½“å‰ç™»å½•è´¦å·")

    u = session.get(User, user_id)
    if not u:
        return _redirect("/admin/users?error=æ‰¾ä¸åˆ°ç”¨æˆ·")

    if getattr(u, "is_active", True) is False:
        return _redirect("/admin/users?message=è¯¥è´¦å·å·²åˆ é™¤")

    # å®‰å…¨ï¼šè‡³å°‘ä¿ç•™ä¸€ä¸ªå¯ç”¨ç®¡ç†å‘˜
    try:
        admins = session.exec(select(User).where(User.role == Role.admin, User.is_active == True)).all()
        if getattr(u.role, "value", None) == "admin" and len(admins) <= 1:
            return _redirect("/admin/users?error=è‡³å°‘éœ€è¦ä¿ç•™ä¸€ä¸ªç®¡ç†å‘˜è´¦å·")
    except Exception:
        pass

    # è½¯åˆ é™¤ï¼šåœç”¨è´¦å· + é‡Šæ”¾ç”¨æˆ·å/é‚®ç®±ï¼Œé¿å…åç»­æ— æ³•å¤ç”¨
    suffix = datetime.utcnow().strftime("%Y%m%d%H%M%S")
    u.is_active = False
    u.deleted_at = datetime.utcnow()
    try:
        u.username = f"deleted_{u.id}_{suffix}"
        u.email = f"deleted_{u.id}_{suffix}@deleted.local"
    except Exception:
        pass
    session.add(u)

    # åˆ é™¤è¯¥è´¦å·çš„å®¢æœç»‘å®šï¼ˆé¿å…ç»§ç»­â€œè®¤é¢†â€å¯¹è¯ï¼‰
    try:
        bs = session.exec(select(AgentBinding).where(AgentBinding.user_id == user_id)).all()
        for b in bs:
            session.delete(b)
    except Exception:
        pass
    # é‡Šæ”¾è¯¥è´¦å·å·²å½’å±çš„å†å²å¯¹è¯ï¼ˆæ–¹ä¾¿åç»­é‡æ–°ç»‘å®š/é‡æ–°åˆ†é…ï¼‰
    try:
        convos = session.exec(select(Conversation).where(Conversation.agent_user_id == user_id)).all()
        for c in convos:
            c.agent_user_id = None
            session.add(c)
    except Exception:
        pass


    session.commit()
    return _redirect("/admin/users?message=å·²åˆ é™¤è´¦å·ï¼ˆå†å²æ•°æ®ä»ä¿ç•™ï¼‰")


@app.get("/conversations", response_class=HTMLResponse)
def conversations_list(
    request: Request,
    user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
    agent_id: list[str] | None = Query(None),
    ext_agent_account: list[str] | None = Query(None),
    platform: str = "",
    match: str = "",
    time_mode: str | None = None,
    start: str = "",
    end: str = "",
    min_rounds: str = "",
    max_rounds: str = "",
    only_flagged: str | None = None,
    reception_scenario: list[str] | None = Query(None),
    satisfaction_change: list[str] | None = Query(None),
    tag_id: list[str] | None = Query(None),
    cid: str = "",
    has_analysis: str = "",
    page: int = 1,
):
    """Conversations list (landing page).

    Filters:
    - agent_id: internal user id (admin/supervisor only)
    - start/end: date range on customer message time (fallback to started_at/uploaded_at)
    - only_flagged: '1' to show analyses flagged for review
    """

    from sqlalchemy import func

    page_size = 100
    try:
        page = int(page or 1)
    except Exception:
        page = 1
    page = max(1, page)

    def _parse_int_list(values: list[str] | None) -> list[int]:
        out: list[int] = []
        for v in values or []:
            s = str(v or "").strip()
            if not s:
                continue
            try:
                out.append(int(s))
            except Exception:
                continue
        return out

    agent_ids = _parse_int_list(agent_id)
    ext_agent_account_norms = [str(v or "").strip() for v in (ext_agent_account or []) if str(v or "").strip()]
    platform_norm = (platform or "").strip().lower()

    # éœ€æ±‚æ›´æ–°ï¼šæ‰€æœ‰å·²ç™»å½•å®¢æœéƒ½å¯ä»¥æŸ¥çœ‹å…¨éƒ¨å¯¹è¯ã€‚
    # åŒæ—¶ä¿ç•™ï¼ˆå¯é€‰çš„ï¼‰æŒ‰ç«™å†…å®¢æœè´¦å·è¿‡æ»¤åŠŸèƒ½ï¼šç”¨äºâ€œèŠå¤©æ¥å¾…ç•Œé¢â€æŒ‰å®¢æœç­›é€‰ã€‚
    conv_where: list = []
    if platform_norm:
        if platform_norm == "unknown":
            conv_where.append(
                or_(
                    Conversation.platform.is_(None),
                    Conversation.platform == "",
                    func.lower(Conversation.platform) == "unknown",
                )
            )
        else:
            # ä½¿ç”¨ lower() å…¼å®¹å†å²æ•°æ®å¤§å°å†™ä¸ä¸€è‡´
            conv_where.append(func.lower(Conversation.platform) == platform_norm)
    if agent_ids:
        # include both explicit assignment and (platform,agent_account) bindings as fallback
        bs = session.exec(select(AgentBinding).where(AgentBinding.user_id.in_(agent_ids))).all()
        bindings_by_user: dict[int, list[AgentBinding]] = {}
        for b in bs or []:
            try:
                uid = int(b.user_id)
            except Exception:
                continue
            bindings_by_user.setdefault(uid, []).append(b)

        multi_or_terms = []
        for uid in agent_ids:
            or_terms = [Conversation.agent_user_id == uid]
            for b in bindings_by_user.get(uid, []):
                # primary agent_account match (legacy)
                platform_norm = str(b.platform or "").strip().lower()
                agent_acc = str(b.agent_account or "").strip()
                if platform_norm and agent_acc:
                    # ä½¿ç”¨ lower() å…¼å®¹å†å²æ•°æ®å¤§å°å†™ä¸ä¸€è‡´
                    or_terms.append(and_(func.lower(Conversation.platform) == platform_norm, Conversation.agent_account == agent_acc))
                # multi-agent: any agent account appearing in messages
                if agent_acc:
                    or_terms.append(
                        exists(
                            select(1)
                            .select_from(Message)
                            .where(
                                Message.conversation_id == Conversation.id,
                                Message.sender == "agent",
                                Message.agent_account == agent_acc,
                            )
                        )
                    )
            multi_or_terms.append(or_(*or_terms))

        if multi_or_terms:
            conv_where.append(or_(*multi_or_terms))

    # ç«™å¤–å®¢æœè´¦å·ç­›é€‰ï¼šç”¨äºç²¾ç¡®å®šä½æŸä¸ªæ·˜å®/æŠ–éŸ³ç­‰â€œå¤–éƒ¨è´¦å·â€å‚ä¸çš„å¯¹è¯
    # å…¼å®¹ä¸¤ç§æ•°æ®æ¥æºï¼š
    # 1) Conversation.agent_accountï¼ˆæ—§æ•°æ®/ä¸»å®¢æœï¼‰
    # 2) Message.agent_accountï¼ˆå¤šå®¢æœåœºæ™¯ï¼Œæ¯æ¡æ¶ˆæ¯è‡ªå·±çš„å¤–éƒ¨è´¦å·ï¼‰
    if ext_agent_account_norms:
        conv_where.append(
            or_(
                Conversation.agent_account.in_(ext_agent_account_norms),
                exists(
                    select(1)
                    .select_from(Message)
                    .where(
                        Message.conversation_id == Conversation.id,
                        Message.sender == "agent",
                        Message.agent_account.in_(ext_agent_account_norms),
                    )
                ),
            )
        )

    # We display & sort by "å®¢æˆ·æ¶ˆæ¯æ—¶é—´" (max buyer ts), fallback to started_at/uploaded_at.
    customer_ts_sq = (
        select(
            Message.conversation_id.label("cid"),
            func.max(Message.ts).label("customer_ts"),
        )
        .where(Message.sender == "buyer", Message.ts.is_not(None))
        .group_by(Message.conversation_id)
    ).subquery()

    # "è½®æ•°" åœ¨åˆ—è¡¨é‡Œåªå±•ç¤ºæ¶ˆæ¯æ¡æ•°ï¼ˆå®¢æˆ· + å®¢æœï¼‰ï¼Œä¸ç®— systemã€‚
    msg_cnt_sq = (
        select(
            Message.conversation_id.label("cid"),
            func.count(Message.id).label("msg_cnt"),
        )
        .where(Message.sender != "system")
        .group_by(Message.conversation_id)
    ).subquery()

    display_ts = func.coalesce(customer_ts_sq.c.customer_ts, Conversation.started_at, Conversation.uploaded_at)

    time_mode_norm = (time_mode or "").strip()
    start_norm = (start or "").strip()
    end_norm = (end or "").strip()

    # Backward compatible: old URLs only had start/end.
    if not time_mode_norm:
        time_mode_norm = "custom" if (start_norm or end_norm) else "all"

    # Prevent "custom" with empty dates from implicitly becoming last 7 days.
    if time_mode_norm == "custom" and not (start_norm or end_norm):
        time_mode_norm = "all"

    allowed_modes = {"all", "last_year", "last_quarter", "last_month", "last_week", "yesterday", "custom"}
    if time_mode_norm not in allowed_modes:
        time_mode_norm = "custom" if (start_norm or end_norm) else "all"

    period = _build_tag_report_period(
        time_mode_norm if time_mode_norm != "custom" else "custom",
        start_norm or None,
        end_norm or None,
        False,
    )

    cur_start = period.get("cur_start")
    cur_end_excl = period.get("cur_end_excl")
    if cur_start:
        conv_where.append(display_ts >= cur_start)
    if cur_end_excl:
        conv_where.append(display_ts < cur_end_excl)

    # Tag/flag filters are applied via ConversationAnalysis (any analysis record).
    match_norm = (match or "").strip()
    only_flagged_bool = (only_flagged == "1")
    reception_scenario_norms = [str(v or "").strip() for v in (reception_scenario or []) if str(v or "").strip()]
    satisfaction_change_norms = [str(v or "").strip() for v in (satisfaction_change or []) if str(v or "").strip()]
    tag_id_norms = [str(v or "").strip() for v in (tag_id or []) if str(v or "").strip()]
    cid_norm = (cid or "").strip()

    # rounds (non-system message count) range filter
    def _parse_nonneg_int(s: str) -> int | None:
        s = (s or "").strip()
        if not s:
            return None
        try:
            v = int(float(s))
        except Exception:
            return None
        return max(0, v)

    min_rounds_i = _parse_nonneg_int(min_rounds)
    max_rounds_i = _parse_nonneg_int(max_rounds)
    if min_rounds_i is not None and max_rounds_i is not None and max_rounds_i < min_rounds_i:
        min_rounds_i, max_rounds_i = max_rounds_i, min_rounds_i
    rounds_expr = func.coalesce(msg_cnt_sq.c.msg_cnt, 0)
    if min_rounds_i is not None:
        conv_where.append(rounds_expr >= min_rounds_i)
    if max_rounds_i is not None:
        conv_where.append(rounds_expr <= max_rounds_i)
    
    # Tag ID filter via ConversationTagHit AND ManualTagBinding
    # This filter is independent of ConversationAnalysis conditions
    if tag_id_norms:
        tag_ids: list[int] = []
        try:
            for s in tag_id_norms:
                try:
                    tag_ids.append(int(s))
                except Exception:
                    continue
            tag_ids = [tid for tid in tag_ids if tid > 0]
        except Exception:
            tag_ids = []

        if tag_ids:
            # Get latest analysis per conversation that has this tag (AI auto hit)
            latest_subq = (
                select(func.max(ConversationAnalysis.id).label("analysis_id"))
                .group_by(ConversationAnalysis.conversation_id)
            ).subquery()
            
            ai_tag_hit_subq = (
                select(ConversationAnalysis.conversation_id)
                .select_from(ConversationTagHit)
                .join(ConversationAnalysis, ConversationAnalysis.id == ConversationTagHit.analysis_id)
                .join(latest_subq, latest_subq.c.analysis_id == ConversationAnalysis.id)
                .where(ConversationTagHit.tag_id.in_(tag_ids))
                .distinct()
            )
            
            # Also check ManualTagBinding (manually added tags by admin/supervisor)
            manual_tag_subq = (
                select(ManualTagBinding.conversation_id)
                .where(ManualTagBinding.tag_id.in_(tag_ids))
                .distinct()
            )
            
            # Combine both: conversation has tag either from AI or manual binding
            conv_where.append(or_(
                Conversation.id.in_(ai_tag_hit_subq),
                Conversation.id.in_(manual_tag_subq)
            ))
    
    if only_flagged_bool or reception_scenario_norms or satisfaction_change_norms:
        a_where = []
        if only_flagged_bool:
            a_where.append(ConversationAnalysis.flag_for_review == True)
        if reception_scenario_norms:
            a_where.append(ConversationAnalysis.reception_scenario.in_(reception_scenario_norms))
        if satisfaction_change_norms:
            a_where.append(ConversationAnalysis.satisfaction_change.in_(satisfaction_change_norms))
        
        subq = select(ConversationAnalysis.conversation_id).where(*a_where).distinct()
        conv_where.append(Conversation.id.in_(subq))
    
    # CID filter (exact match by conversation ID)
    if cid_norm:
        try:
            cid_int = int(cid_norm)
            conv_where.append(Conversation.id == cid_int)
        except Exception:
            pass

    # Has analysis filter (æ˜¯å¦å·²AIè´¨æ£€)
    has_analysis_norm = (has_analysis or "").strip()
    if has_analysis_norm == "1":
        # åªæ˜¾ç¤ºå·²è´¨æ£€çš„å¯¹è¯
        analysis_subq = select(ConversationAnalysis.conversation_id).distinct()
        conv_where.append(Conversation.id.in_(analysis_subq))
    elif has_analysis_norm == "0":
        # åªæ˜¾ç¤ºæœªè´¨æ£€çš„å¯¹è¯
        analysis_subq = select(ConversationAnalysis.conversation_id).distinct()
        conv_where.append(~Conversation.id.in_(analysis_subq))

    # Message match filter: hit if ANY message text in that day conversation contains the keyword.
    if match_norm:
        like = f"%{match_norm}%"
        conv_where.append(
            exists(
                select(1)
                .select_from(Message)
                .where(
                    Message.conversation_id == Conversation.id,
                    Message.sender != "system",
                    Message.text.ilike(like),
                )
            )
        )

    # total count for pagination
    # æ³¨æ„ï¼šå¦‚æœä½¿ç”¨äº†æ—¶é—´ç­›é€‰ï¼ˆdisplay_tsï¼‰ï¼Œè®¡æ•°æŸ¥è¯¢ä¹Ÿéœ€è¦ join customer_ts_sq
    count_q = select(func.count(Conversation.id))
    if cur_start or cur_end_excl:
        # æ—¶é—´ç­›é€‰ä¾èµ– display_tsï¼Œéœ€è¦ join customer_ts_sq
        count_q = count_q.outerjoin(customer_ts_sq, customer_ts_sq.c.cid == Conversation.id)
    if min_rounds_i is not None or max_rounds_i is not None:
        count_q = count_q.outerjoin(msg_cnt_sq, msg_cnt_sq.c.cid == Conversation.id)
    if conv_where:
        count_q = count_q.where(*conv_where)
    
    total = session.exec(count_q).one()
    try:
        total = int(total or 0)
    except Exception:
        total = 0
    total_pages = max(1, (total + page_size - 1) // page_size)
    if page > total_pages:
        page = total_pages

    q = (
        select(
            Conversation,
            display_ts.label("display_ts"),
            func.coalesce(msg_cnt_sq.c.msg_cnt, 0).label("rounds"),
        )
        .outerjoin(customer_ts_sq, customer_ts_sq.c.cid == Conversation.id)
        .outerjoin(msg_cnt_sq, msg_cnt_sq.c.cid == Conversation.id)
        .order_by(display_ts.desc(), Conversation.id.desc())
    )
    if conv_where:
        q = q.where(*conv_where)

    rows = session.exec(q.offset((page - 1) * page_size).limit(page_size)).all()
    convos = [r[0] for r in rows]
    display_ts_by_conv = {r[0].id: r[1] for r in rows if r and r[0] and r[0].id is not None}
    rounds_by_conv = {r[0].id: int(r[2] or 0) for r in rows if r and r[0] and r[0].id is not None}

    conv_ids = [c.id for c in convos if c.id is not None]
    latest_by_conv = {}
    if conv_ids:
        rows = session.exec(
            select(ConversationAnalysis)
            .where(ConversationAnalysis.conversation_id.in_(conv_ids))
            .order_by(ConversationAnalysis.conversation_id.asc(), ConversationAnalysis.id.desc())
        ).all()
        for r in rows:
            if r.conversation_id not in latest_by_conv:
                latest_by_conv[r.conversation_id] = r

    users = session.exec(select(User)).all()
    users_by_id = {u.id: u for u in users}
    # å¯æ¥å¾…è´¦å·ï¼šå®¢æœ + ä¸»ç®¡ï¼ˆä¸»ç®¡ä¹Ÿå¯æ¥å¾…/ç»‘å®šï¼‰
    agents = [u for u in users if u.role in (Role.agent, Role.supervisor) and getattr(u, 'is_active', True)]

    # ç«™å¤–å®¢æœè´¦å·ä¸‹æ‹‰é€‰é¡¹ï¼šæ¥è‡ªæ¶ˆæ¯é‡Œçš„ agent_account + å¯¹è¯ä¸»å­—æ®µ agent_account
    ext_accounts: set[str] = set()
    try:
        rows_acc = session.exec(
            select(Message.agent_account)
            .where(Message.sender == "agent", Message.agent_account.is_not(None), Message.agent_account != "")
            .distinct()
        ).all()
        for r in (rows_acc or []):
            if isinstance(r, tuple):
                v = r[0]
            else:
                v = r
            v = str(v or "").strip()
            if v:
                ext_accounts.add(v)
    except Exception:
        pass

    try:
        rows_conv_acc = session.exec(
            select(Conversation.agent_account)
            .where(Conversation.agent_account.is_not(None), Conversation.agent_account != "")
            .distinct()
        ).all()
        for r in (rows_conv_acc or []):
            if isinstance(r, tuple):
                v = r[0]
            else:
                v = r
            v = str(v or "").strip()
            if v:
                ext_accounts.add(v)
    except Exception:
        pass

    external_agent_accounts = sorted(ext_accounts)

    # å¹³å°ä¸‹æ‹‰é€‰é¡¹ï¼ˆå¯¹è¯åˆ—è¡¨ç­›é€‰ï¼‰
    platform_labels = {
        "taobao": "æ·˜å®",
        "douyin": "æŠ–éŸ³",
        "unknown": "æœªçŸ¥",
    }
    platform_opt_set: set[str] = {"taobao", "douyin", "unknown"}
    try:
        rows_plat = session.exec(
            select(Conversation.platform)
            .where(Conversation.platform.is_not(None), Conversation.platform != "")
            .distinct()
        ).all()
        for r in rows_plat or []:
            v = r[0] if isinstance(r, tuple) else r
            v = str(v or "").strip().lower()
            if v:
                platform_opt_set.add(v)
    except Exception:
        pass
    preferred = ["taobao", "douyin", "unknown"]
    platform_options = [p for p in preferred if p in platform_opt_set] + sorted([p for p in platform_opt_set if p not in preferred])

    # === å®¢æœæ˜¾ç¤ºï¼ˆæ”¯æŒåŒä¸€å¯¹è¯å¤šå®¢æœï¼‰===
    # åˆ—è¡¨é¡µå±•ç¤ºâ€œè¿™ä¸ªå¯¹è¯é‡Œå‡ºç°è¿‡å“ªäº›å®¢æœâ€ã€‚ä¼˜å…ˆç«™å†…ç»‘å®šåï¼Œå…¶æ¬¡å¤–éƒ¨æ˜µç§°/è´¦å·ã€‚
    # å¤šä¸ªå®¢æœç”¨ " / " è¿æ¥ï¼›è¶…è¿‡ 3 ä¸ªåˆ™æ˜¾ç¤ºå‰ 3 ä¸ª + " â€¦(+N)".

    bindings = session.exec(select(AgentBinding)).all()
    binding_map = {
        (str(b.platform or "").strip().lower(), str(b.agent_account or "").strip()): int(b.user_id)
        for b in (bindings or [])
        if b and b.platform and b.agent_account and b.user_id
    }

    platform_by_conv: dict[int, str] = {
        int(c.id): str(c.platform or "").strip().lower()
        for c in convos
        if c and c.id is not None
    }

    display_agent_label_by_conv: dict[int, str] = {}

    # å…ˆæ”¶é›†æ¯ä¸ªå¯¹è¯é‡Œâ€œçœŸå®å‡ºç°è¿‡çš„å®¢æœå‚ä¸è€…â€
    participants_by_conv: dict[int, list[tuple[int | None, str, str, int]]] = {}
    if conv_ids:
        rows = session.exec(
            select(
                Message.conversation_id,
                Message.agent_user_id,
                Message.agent_account,
                Message.agent_nick,
                func.count(Message.id).label("cnt"),
            )
            .where(
                Message.conversation_id.in_(conv_ids),
                Message.sender == "agent",
                (
                    (Message.agent_user_id.is_not(None))
                    | ((Message.agent_account.is_not(None)) & (Message.agent_account != ""))
                    | ((Message.agent_nick.is_not(None)) & (Message.agent_nick != ""))
                ),
            )
            .group_by(
                Message.conversation_id,
                Message.agent_user_id,
                Message.agent_account,
                Message.agent_nick,
            )
        ).all()

        for (cid, uid, acc, nick, cnt) in rows:
            try:
                cid_int = int(cid)
            except Exception:
                continue
            participants_by_conv.setdefault(cid_int, []).append(
                (
                    int(uid) if uid is not None else None,
                    str(acc or "").strip(),
                    str(nick or "").strip(),
                    int(cnt or 0),
                )
            )

    def _user_label(uid: int) -> str | None:
        u = users_by_id.get(uid)
        if not u:
            return None
        return f"{u.username} Â· {u.name}" if u.username else (u.name or u.email)

    def _resolve_label(*, conv_id: int, uid: int | None, acc: str, nick: str) -> str:
        """Resolve label for display.

        é‡è¦ï¼šä»¥â€œå®æ—¶ç»‘å®šâ€ä¸ºå‡†ã€‚
        - å…ˆç”¨ AgentBinding(platform, agent_account) -> user æ˜ å°„
        - å†å›é€€åˆ°æ¶ˆæ¯é‡Œå­˜çš„ agent_user_idï¼ˆå†å²åŒæ­¥å­—æ®µï¼‰
        - æœ€åå›é€€å¤–éƒ¨æ˜µç§°/è´¦å·
        """
        platform = platform_by_conv.get(int(conv_id), "")
        if acc:
            bound_uid = binding_map.get((platform, acc))
            if bound_uid:
                lab = _user_label(int(bound_uid))
                if lab:
                    return lab

        if uid:
            lab = _user_label(uid)
            if lab:
                return lab

        return nick or acc or "-"

    for c in convos:
        if not c or c.id is None:
            continue
        cid = int(c.id)

        parts = participants_by_conv.get(cid) or []
        if parts:
            parts_sorted = sorted(parts, key=lambda x: (-x[3], str(x[0] or ""), x[1], x[2]))

            labels: list[str] = []
            seen: set[str] = set()
            for (uid, acc, nick, _cnt) in parts_sorted:
                lab = _resolve_label(conv_id=cid, uid=uid, acc=acc, nick=nick)
                if lab and lab not in seen and lab != "-":
                    labels.append(lab)
                    seen.add(lab)

            if labels:
                if len(labels) > 3:
                    display_agent_label_by_conv[cid] = " / ".join(labels[:3]) + f" â€¦(+{len(labels) - 3})"
                else:
                    display_agent_label_by_conv[cid] = " / ".join(labels)
                continue

        # æ²¡æ”¶é›†åˆ°å‚ä¸å®¢æœæ—¶ï¼Œå›é€€åˆ°â€œä¸»å®¢æœâ€é€»è¾‘ï¼ˆå…¼å®¹æ—§æ•°æ®ï¼‰
        if c.agent_user_id and users_by_id.get(c.agent_user_id):
            u = users_by_id[c.agent_user_id]
            display_agent_label_by_conv[cid] = f"{u.username} Â· {u.name}" if u.username else (u.name or u.email)
            continue

        uid2 = binding_map.get((platform_by_conv.get(cid, ""), str(c.agent_account or "").strip())) if c.agent_account else None
        if uid2 and users_by_id.get(uid2):
            u = users_by_id[int(uid2)]
            display_agent_label_by_conv[cid] = f"{u.username} Â· {u.name}" if u.username else (u.name or u.email)
            continue

        display_agent_label_by_conv[cid] = (c.agent_account or "-")

    # conversations list page does not need per-message labels; keep an empty mapping
    # to avoid NameError from legacy template args.
    sender_label_by_msgid = {}

    # Get tag categories and tags for filter dropdown
    tag_categories = session.exec(
        select(TagCategory)
        .where(TagCategory.is_active == True)
        .order_by(TagCategory.sort_order.asc(), TagCategory.id.asc())
    ).all()
    
    tag_definitions = session.exec(
        select(TagDefinition)
        .where(TagDefinition.is_active == True)
        .order_by(TagDefinition.category_id.asc(), TagDefinition.sort_order.asc(), TagDefinition.id.asc())
    ).all()

    # æŸ¥è¯¢æ¯ä¸ªå¯¹è¯çš„æ ‡ç­¾ä¿¡æ¯ï¼ˆé€šè¿‡æœ€æ–°åˆ†æï¼‰
    # conversations åˆ—è¡¨é¡µåªå±•ç¤º AI å‘½ä¸­çš„æ ‡ç­¾ï¼ˆConversationTagHitï¼‰ï¼Œå¹¶æºå¸¦ reason ä¾›å‰ç«¯æ‚¬æµ®/é•¿æŒ‰æŸ¥çœ‹ã€‚
    tags_by_conv: dict[int, list[dict[str, str]]] = {}
    if conv_ids and latest_by_conv:
        analysis_ids = [a.id for a in latest_by_conv.values() if a and a.id is not None]
        if analysis_ids:
            tag_hits = session.exec(
                select(ConversationTagHit, TagDefinition)
                .join(TagDefinition, TagDefinition.id == ConversationTagHit.tag_id)
                .where(ConversationTagHit.analysis_id.in_(analysis_ids))
                .order_by(
                    TagDefinition.category_id.asc(),
                    TagDefinition.sort_order.asc(),
                    TagDefinition.id.asc(),
                    ConversationTagHit.id.asc(),
                )
            ).all()
            
            # æ„å»º analysis_id -> conversation_id æ˜ å°„
            analysis_to_conv: dict[int, int] = {}
            for conv_id, analysis in latest_by_conv.items():
                if analysis and analysis.id is not None:
                    analysis_to_conv[analysis.id] = conv_id
            
            # æ„å»ºæ¯ä¸ªå¯¹è¯çš„æ ‡ç­¾åˆ—è¡¨
            seen_tag_ids_by_conv: dict[int, set[int]] = {}
            for hit, tag_def in tag_hits:
                conv_id = analysis_to_conv.get(hit.analysis_id)
                if not conv_id:
                    continue
                if not tag_def or not tag_def.id or not tag_def.name:
                    continue

                if conv_id not in tags_by_conv:
                    tags_by_conv[conv_id] = []
                    seen_tag_ids_by_conv[conv_id] = set()

                if tag_def.id in seen_tag_ids_by_conv[conv_id]:
                    continue
                seen_tag_ids_by_conv[conv_id].add(tag_def.id)

                tags_by_conv[conv_id].append(
                    {
                        "name": str(tag_def.name),
                        "ai_reason": str((hit.reason or "")).strip(),
                    }
                )

    return _template(
        request,
        "conversations.html",
        user=user,
        convos=convos,
        display_ts_by_conv=display_ts_by_conv,
        rounds_by_conv=rounds_by_conv,
        latest_by_conv=latest_by_conv,
        users_by_id=users_by_id,
        agents=agents,
        external_agent_accounts=external_agent_accounts,
        display_agent_label_by_conv=display_agent_label_by_conv,
        tag_categories=tag_categories,
        tag_definitions=tag_definitions,
        tags_by_conv=tags_by_conv,
        # èŠå¤©æ¥å¾…é¡µéœ€è¦â€œæŒ‰å®¢æœè´¦å·ç­›é€‰â€ï¼Œæ‰€ä»¥æ‰€æœ‰è§’è‰²éƒ½å›æ˜¾å½“å‰é€‰æ‹©ã€‚
        agent_ids=agent_ids,
        ext_agent_accounts_selected=ext_agent_account_norms,
        match=match_norm,
        time_mode=time_mode_norm,
        start=str(period.get("display_start") or ""),
        end=str(period.get("display_end") or ""),
        period_label=str(period.get("cur_label") or ""),
        min_rounds=str(min_rounds_i) if min_rounds_i is not None else "",
        max_rounds=str(max_rounds_i) if max_rounds_i is not None else "",
        only_flagged=only_flagged_bool,
        reception_scenarios=reception_scenario_norms,
        satisfaction_changes=satisfaction_change_norms,
        tag_ids=tag_id_norms,
        cid=cid_norm,
        page=page,
        total=total,
        total_pages=total_pages,
        page_size=page_size,
        has_analysis=has_analysis_norm,
        platform=platform_norm,
        platform_options=platform_options,
        platform_labels=platform_labels,
        time_modes=[
            ("all", "æ±‡æ€»"),
            ("last_year", "å¹´"),
            ("last_quarter", "å­£åº¦"),
            ("last_month", "æœˆ"),
            ("last_week", "å‘¨"),
            ("yesterday", "æ—¥"),
            ("custom", "è‡ªå®šä¹‰"),
        ],
        base_qs=_build_conversations_base_qs(
            agent_ids=agent_ids,
            ext_agent_accounts=ext_agent_account_norms,
            platform=platform_norm,
            match=match_norm,
            time_mode=time_mode_norm,
            start=str(period.get("display_start") or ""),
            end=str(period.get("display_end") or ""),
            min_rounds=str(min_rounds_i) if min_rounds_i is not None else "",
            max_rounds=str(max_rounds_i) if max_rounds_i is not None else "",
            only_flagged=only_flagged_bool,
            reception_scenarios=reception_scenario_norms,
            satisfaction_changes=satisfaction_change_norms,
            tag_ids=tag_id_norms,
            cid=cid_norm,
            has_analysis=has_analysis_norm,
        ),
    )


def _build_conversations_base_qs(
    agent_ids: list[int],
    ext_agent_accounts: list[str],
    platform: str,
    match: str,
    time_mode: str,
    start: str,
    end: str,
    only_flagged: bool,
    min_rounds: str = "",
    max_rounds: str = "",
    reception_scenarios: list[str] | None = None,
    satisfaction_changes: list[str] | None = None,
    tag_ids: list[str] | None = None,
    cid: str = "",
    has_analysis: str = "",
) -> str:
    parts: list[str] = []
    if agent_ids:
        for uid in agent_ids:
            parts.append(f"agent_id={uid}")
    if ext_agent_accounts:
        from urllib.parse import quote_plus
        for acc in ext_agent_accounts:
            parts.append(f"ext_agent_account={quote_plus(acc)}")
    if platform and platform.strip():
        from urllib.parse import quote_plus
        parts.append(f"platform={quote_plus(platform.strip())}")
    if match:
        from urllib.parse import quote_plus
        parts.append(f"match={quote_plus(match)}")
    if time_mode and time_mode != "all":
        parts.append(f"time_mode={time_mode}")
    if start:
        parts.append(f"start={start}")
    if end:
        parts.append(f"end={end}")
    if min_rounds:
        parts.append(f"min_rounds={min_rounds}")
    if max_rounds:
        parts.append(f"max_rounds={max_rounds}")
    if only_flagged:
        parts.append("only_flagged=1")
    if reception_scenarios:
        from urllib.parse import quote_plus
        for val in reception_scenarios:
            parts.append(f"reception_scenario={quote_plus(val)}")
    if satisfaction_changes:
        from urllib.parse import quote_plus
        for val in satisfaction_changes:
            parts.append(f"satisfaction_change={quote_plus(val)}")
    if tag_ids:
        for tid in tag_ids:
            parts.append(f"tag_id={tid}")
    if cid:
        parts.append(f"cid={cid}")
    if has_analysis:
        parts.append(f"has_analysis={has_analysis}")
    return "&".join(parts)


@app.get("/conversations/{conversation_id}", response_class=HTMLResponse)
def conversation_detail(
    request: Request,
    conversation_id: int,
    user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
):
    conv = session.get(Conversation, conversation_id)
    if not conv:
        return _template(request, "error.html", user=user, message="æ‰¾ä¸åˆ°å¯¹è¯")

        # permissions: agents can view their own conversations; supervisors/admin can view all
    if not _can_view_conversation(user, session, conv):
        return _template(request, "error.html", user=user, message="æ— æƒé™æŸ¥çœ‹è¯¥å¯¹è¯")

    agent = session.get(User, conv.agent_user_id) if conv.agent_user_id else None

    # ç»™æ¨¡æ¿ç”¨çš„ï¼šconversation_id -> â€œåº”è¯¥æ˜¾ç¤ºçš„å®¢æœåå­—â€
    # ä¸åˆ—è¡¨é¡µä¿æŒä¸€è‡´ï¼šä»¥â€œå®æ—¶ç»‘å®šâ€ä¸ºå‡†ï¼Œå…¶æ¬¡å›é€€å†å²å­—æ®µï¼Œæœ€åå…œåº•å¤–éƒ¨å®¢æœè´¦å·ã€‚
    display_agent_label_by_conv: dict[int, str] = {}

    users_by_id_tmp = {u.id: u for u in session.exec(select(User)).all()}

    # 1) ä»¥ AgentBinding å®æ—¶åæŸ¥
    if conv.platform and conv.agent_account:
        from sqlalchemy import func
        platform_norm = str(conv.platform or "").strip().lower()
        b = session.exec(
            select(AgentBinding).where(
                func.lower(AgentBinding.platform) == platform_norm,
                AgentBinding.agent_account == str(conv.agent_account or "").strip(),
            )
        ).first()
        if b and users_by_id_tmp.get(b.user_id):
            u = users_by_id_tmp[b.user_id]
            display_agent_label_by_conv[conv.id] = f"{u.username} Â· {u.name}" if u.username else (u.name or u.email)

    # 2) å›é€€ï¼šå†å²å­—æ®µï¼ˆç”¨äºæ²¡æœ‰ç»‘å®šã€æˆ–æ—§æ•°æ®ï¼‰
    if conv.id not in display_agent_label_by_conv and conv.agent_user_id and users_by_id_tmp.get(conv.agent_user_id):
        u = users_by_id_tmp[conv.agent_user_id]
        display_agent_label_by_conv[conv.id] = f"{u.username} Â· {u.name}" if u.username else (u.name or u.email)

    # 3) å…œåº•ï¼šå¤–éƒ¨å®¢æœè´¦å·
    if conv.id not in display_agent_label_by_conv:
        display_agent_label_by_conv[conv.id] = (conv.agent_account or "-")

    # === Build "full user history" across days (UserThread) ===

    def _platform_norm(p: str | None) -> str:
        return str(p or "unknown").strip().lower() or "unknown"

    # Best-effort: ensure this conversation is attached to a thread (platform + buyer_id).
    try:
        if (not getattr(conv, "user_thread_id", None)) and (conv.buyer_id or "").strip():
            pnorm = _platform_norm(conv.platform)
            buyer_id_norm = str(conv.buyer_id or "").strip()
            if buyer_id_norm:
                ut = session.exec(
                    select(UserThread).where(UserThread.platform == pnorm, UserThread.buyer_id == buyer_id_norm)
                ).first()
                if not ut:
                    ut = UserThread(platform=pnorm, buyer_id=buyer_id_norm, meta={})
                    session.add(ut)
                    session.commit()
                    session.refresh(ut)
                conv.user_thread_id = int(ut.id) if ut.id is not None else None
                session.add(conv)
                session.commit()
    except Exception:
        # If schema/backfill is still running, fall back to day-only view.
        pass

    # Thread conversations (ordered by started/uploaded; within each, messages are ordered by ts/id).
    thread_convs: list[Conversation] = []
    try:
        if getattr(conv, "user_thread_id", None):
            thread_convs = session.exec(
                select(Conversation)
                .where(Conversation.user_thread_id == conv.user_thread_id)
                .order_by(
                    Conversation.started_at.asc().nulls_last(),
                    Conversation.uploaded_at.asc().nulls_last(),
                    Conversation.id.asc(),
                )
            ).all()
    except Exception:
        thread_convs = []

    if not thread_convs:
        thread_convs = [conv]

    thread_conv_ids = [c.id for c in thread_convs if c.id is not None]

    # Display date per conversation (max buyer ts, fallback to started/uploaded)
    buyer_ts_map: dict[int, datetime] = {}
    try:
        rows = session.exec(
            select(Message.conversation_id, func.max(Message.ts))
            .where(
                Message.conversation_id.in_(thread_conv_ids),
                Message.sender == "buyer",
                Message.ts.is_not(None),
            )
            .group_by(Message.conversation_id)
        ).all()
        for cid, ts_max in rows:
            if cid is None or ts_max is None:
                continue
            buyer_ts_map[int(cid)] = ts_max
    except Exception:
        buyer_ts_map = {}

    def _display_ts_for_conv(c: Conversation) -> datetime | None:
        if c.id is not None and int(c.id) in buyer_ts_map:
            return buyer_ts_map[int(c.id)]
        return c.started_at or c.uploaded_at

    def _date_label_for_conv(c: Conversation) -> str:
        dt = _display_ts_for_conv(c)
        if not dt:
            return "æœªçŸ¥æ—¥æœŸ"
        try:
            return dt.strftime("%Y-%m-%d")
        except Exception:
            return str(dt)[:10]

    # Messages grouped per conversation
    msgs_by_conv: dict[int, list[Message]] = {}
    for c in thread_convs:
        if not c.id:
            continue
        cid = int(c.id)
        msgs_by_conv[cid] = session.exec(
            select(Message)
            .where(Message.conversation_id == cid)
            .order_by(Message.ts.asc().nulls_last(), Message.id.asc())
        ).all()

    # Flatten with separators for UI rendering
    history_items: list[dict] = []
    global_idx = 0
    today_local_to_global: dict[int, int] = {}
    today_anchor_global: int | None = None

    for c in thread_convs:
        if not c.id:
            continue
        cid = int(c.id)
        label = _date_label_for_conv(c)
        history_items.append(
            {
                "type": "sep",
                "conv_id": cid,
                "label": label,
                "is_today": cid == int(conv.id),
            }
        )

        local_today_idx = 0
        for m in (msgs_by_conv.get(cid) or []):
            is_today = (cid == int(conv.id))
            today_local_idx: int | None = None
            if is_today:
                today_local_idx = local_today_idx
                today_local_to_global[local_today_idx] = global_idx
                if today_anchor_global is None:
                    today_anchor_global = global_idx
                local_today_idx += 1

            history_items.append(
                {
                    "type": "msg",
                    "conv_id": cid,
                    "msg": m,
                    "global_idx": global_idx,
                    "is_today": is_today,
                    "today_local_idx": today_local_idx,
                }
            )
            global_idx += 1

    # Day-scoped msgs for analysis/task logic
    msgs_today = msgs_by_conv.get(int(conv.id), []) if conv.id is not None else []
    msgs = msgs_today  # keep legacy name for downstream helpers

    # åªæŸ¥è¯¢æœ€æ–°çš„ä¸€æ¡ AI åˆ†æï¼ˆä¼˜åŒ–æ€§èƒ½ï¼‰
    latest_analysis = session.exec(
        select(ConversationAnalysis)
        .where(ConversationAnalysis.conversation_id == conv.id)
        .order_by(ConversationAnalysis.id.desc())
        .limit(1)
    ).first()
    # ä¿æŒ analyses å˜é‡ç”¨äºå‘åå…¼å®¹
    analyses = [latest_analysis] if latest_analysis else []
    
    tasks = session.exec(
        select(TrainingTask)
        .where(TrainingTask.conversation_id == conv.id)
        .order_by(TrainingTask.created_at.desc())
    ).all()

    # Convert evidence highlights (local idx within today's conversation) -> global idx for full history rendering
    highlight_yellow_set = set()
    highlight_green_set = set()
    highlight_red_set = set()
    if latest_analysis:
        ev = latest_analysis.evidence or {}
        for h in (ev.get("highlights") or []):
            try:
                li = int(h.get("message_index"))
                gi = today_local_to_global.get(li)
                if gi is not None:
                    highlight_red_set.add(int(gi))
            except Exception:
                pass
        for h in (ev.get("customer_issue_highlights") or []):
            try:
                li = int(h.get("message_index"))
                gi = today_local_to_global.get(li)
                if gi is not None:
                    highlight_yellow_set.add(int(gi))
            except Exception:
                pass
        for h in (ev.get("must_read_highlights") or []):
            try:
                li = int(h.get("message_index"))
                gi = today_local_to_global.get(li)
                if gi is not None:
                    highlight_green_set.add(int(gi))
            except Exception:
                pass

    # æ ‡ç­¾å‘½ä¸­ï¼ˆæœ€æ–°ä¸€æ¡AIè´¨æ£€ï¼‰ï¼Œå« evidence ç”¨äºç‚¹å‡»é«˜äº®
    tag_groups: list[dict[str, object]] = []
    if analyses and analyses[0].id is not None:
        try:
            rows = session.exec(
                select(ConversationTagHit, TagDefinition, TagCategory)
                .join(TagDefinition, TagDefinition.id == ConversationTagHit.tag_id)
                .join(TagCategory, TagCategory.id == TagDefinition.category_id)
                .where(
                    ConversationTagHit.analysis_id == analyses[0].id,
                    TagDefinition.is_active == True,
                    TagCategory.is_active == True,
                )
                .order_by(TagCategory.sort_order.asc(), TagCategory.id.asc(), TagDefinition.sort_order.asc(), TagDefinition.id.asc())
            ).all()

            grouped: dict[int, dict[str, object]] = {}
            for hit, tag, cat in rows:
                cid = int(cat.id or 0)
                g = grouped.get(cid)
                if not g:
                    g = {"category_id": cid, "category": cat.name or "", "items": []}
                    grouped[cid] = g
                ev = hit.evidence or {}
                ev_list = ev if isinstance(ev, list) else (ev.get("evidence") if isinstance(ev, dict) else []) or []
                if not isinstance(ev_list, list):
                    ev_list = []
                evidence_global: list[dict[str, object]] = []
                for e in ev_list:
                    if not isinstance(e, dict):
                        continue
                    start_loc = e.get("start_index") if e.get("start_index") is not None else e.get("message_index")
                    end_loc = e.get("end_index") if e.get("end_index") is not None else e.get("message_index")
                    if start_loc is None and end_loc is None:
                        continue
                    try:
                        si = int(start_loc) if start_loc is not None else 0
                        ei = int(end_loc) if end_loc is not None else si
                    except Exception:
                        continue
                    indices = []
                    for li in range(si, ei + 1):
                        gi = today_local_to_global.get(li)
                        if gi is not None:
                            indices.append(int(gi))
                    if indices:
                        evidence_global.append({"indices": indices})
                grouped[cid]["items"].append({
                    "hit_id": int(hit.id or 0),
                    "tag_id": int(tag.id or 0),
                    "tag": tag.name or "",
                    "standard": (tag.standard or "") or (tag.description or "") or "æœªé…ç½®",
                    "reason": (hit.reason or "").strip(),
                    "evidence_global": evidence_global,
                    "evidence_json": json.dumps(evidence_global),
                })
            tag_groups = list(grouped.values())
        except Exception:
            tag_groups = []

    # æ‰‹åŠ¨ç»‘å®šæ ‡ç­¾ï¼ˆå¯¹è¯çº§ç®¡ç†ï¼‰
    manual_tag_bindings: list[dict[str, object]] = []
    if conv.id is not None:
        rows = session.exec(
            select(ManualTagBinding, TagDefinition, TagCategory)
            .join(TagDefinition, TagDefinition.id == ManualTagBinding.tag_id)
            .join(TagCategory, TagCategory.id == TagDefinition.category_id)
            .where(ManualTagBinding.conversation_id == conv.id)
            .order_by(
                TagCategory.sort_order.asc(),
                TagCategory.id.asc(),
                TagDefinition.sort_order.asc(),
                TagDefinition.id.asc(),
                ManualTagBinding.id.asc(),
            )
        ).all()
        for mb, t, c in rows:
            # Process evidence for manual tags
            # Note: Evidence conversion happens later after msgs_by_conv is loaded
            ev = mb.evidence or {}
            ev_list = ev if isinstance(ev, list) else (ev.get("evidence") if isinstance(ev, dict) else []) or []
            
            evidence_global: list[dict[str, object]] = []
            if isinstance(ev_list, list):
                for e in ev_list:
                    if not isinstance(e, dict):
                        continue
                    start_loc = e.get("start_index") if e.get("start_index") is not None else e.get("message_index")
                    end_loc = e.get("end_index") if e.get("end_index") is not None else e.get("message_index")
                    if start_loc is None and end_loc is None:
                        continue
                    try:
                        si = int(start_loc) if start_loc is not None else 0
                        ei = int(end_loc) if end_loc is not None else si
                    except Exception:
                        continue
                    indices = []
                    for li in range(si, ei + 1):
                        gi = today_local_to_global.get(li)
                        if gi is not None:
                            indices.append(int(gi))
                    if (not indices) and (global_idx > 0):
                        # Back-compat: some manual bindings may have stored GLOBAL indices directly.
                        # If the indices are out of today's local range but within the global range, treat them as global.
                        today_len = len(msgs_today) if isinstance(msgs_today, list) else 0
                        if (si >= today_len or ei >= today_len) and (si < global_idx or ei < global_idx):
                            gs = max(0, min(si, ei))
                            ge = min(global_idx - 1, max(si, ei))
                            if ge >= gs:
                                indices = list(range(gs, ge + 1))
                    if indices:
                        evidence_global.append({"indices": indices})

            manual_tag_bindings.append({
                "binding_id": mb.id,
                "tag_id": t.id,
                "category_id": int(c.id or 0),
                "tag": t.name or "",
                "category": c.name or "",
                "standard": (t.standard or "") or (t.description or "") or "æœªé…ç½®",
                "reason": mb.reason or "",
                "evidence": ev_list if isinstance(ev_list, list) else [],
                "evidence_global": evidence_global,
                "evidence_json": json.dumps(evidence_global),
            })

    # å…¨éƒ¨æ ‡ç­¾ï¼ˆä¾›æ·»åŠ ä¸‹æ‹‰ï¼‰
    tags_for_add: list[dict[str, object]] = []
    _cats, _tags_by_cat = _load_tag_tree(session)
    for c in _cats:
        for t in _tags_by_cat.get(int(c.id or 0), []):
            if t.is_active and c.is_active:
                tags_for_add.append({"id": t.id, "name": t.name or "", "category": c.name or ""})

    # æ„å»ºå‘½ä¸­æ ‡ç­¾æ˜ å°„ï¼ˆtag_id -> hit infoï¼‰ï¼Œç”¨äºæ ‡ç­¾ä½“ç³»æ¨¡å—å±•ç¤º
    hit_tags_map: dict[int, dict[str, object]] = {}
    for g in tag_groups:
        for it in g.get("items", []):
            tid = it.get("tag_id")
            if tid:
                hit_tags_map[tid] = {
                    "reason": it.get("reason", ""),
                    "evidence_json": it.get("evidence_json", "[]"),
                }

    # åˆå¹¶ï¼šæ ‡ç­¾ä½“ç³»ï¼ˆAIå‘½ä¸­ï¼‰ + æ‰‹åŠ¨ç»‘å®š => å¯¹è¯æ ‡ç­¾ç®¡ç†é¢æ¿
    manual_tag_ids: set[int] = set()
    for b in manual_tag_bindings:
        try:
            tid = int(b.get("tag_id") or 0)
        except Exception:
            tid = 0
        if tid:
            manual_tag_ids.add(tid)

    combined_by_cat: dict[int, dict[str, object]] = {}

    def _ensure_group(cid: int, name: str) -> dict[str, object]:
        g = combined_by_cat.get(cid)
        if not g:
            g = {"category_id": cid, "category": name, "items": []}
            combined_by_cat[cid] = g
        return g

    # Manual first (manual overrides AI when same tag_id exists)
    for b in manual_tag_bindings:
        try:
            tid = int(b.get("tag_id") or 0)
        except Exception:
            tid = 0
        if not tid:
            continue

        try:
            cid = int(b.get("category_id") or 0)
        except Exception:
            cid = 0
        cat_name = str(b.get("category") or "")

        manual_reason = str(b.get("reason") or "").strip()
        hit_meta = hit_tags_map.get(tid) or {}
        hit_reason = str(hit_meta.get("reason") or "").strip()

        display_reason = manual_reason or hit_reason
        reason_from = "manual" if manual_reason else ("ai" if hit_reason else "")

        evidence_json = str(b.get("evidence_json") or "[]")
        try:
            has_manual_evidence = bool(json.loads(evidence_json))
        except Exception:
            has_manual_evidence = False
        if (not has_manual_evidence) and hit_meta.get("evidence_json"):
            evidence_json = str(hit_meta.get("evidence_json") or "[]")

        _ensure_group(cid, cat_name)["items"].append(
            {
                "tag_id": tid,
                "tag": b.get("tag") or "",
                "source": "manual",
                "reason": display_reason,
                "reason_from": reason_from,
                "standard": b.get("standard") or "",
                "evidence_json": evidence_json,
                "binding_id": b.get("binding_id"),
            }
        )

    # AI hits (skip if overridden by manual binding)
    for g in tag_groups:
        try:
            cid = int(g.get("category_id") or 0)
        except Exception:
            cid = 0
        cat_name = str(g.get("category") or "")
        for it in g.get("items", []):
            try:
                tid = int(it.get("tag_id") or 0)
            except Exception:
                tid = 0
            if not tid:
                continue
            if tid in manual_tag_ids:
                continue
            _ensure_group(cid, cat_name)["items"].append(
                {
                    "tag_id": tid,
                    "tag": it.get("tag") or "",
                    "source": "ai",
                    "reason": it.get("reason") or "",
                    "reason_from": "ai",
                    "standard": it.get("standard") or "",
                    "evidence_json": it.get("evidence_json") or "[]",
                    "hit_id": it.get("hit_id"),
                }
            )

    tag_panel_groups: list[dict[str, object]] = []
    seen_cat: set[int] = set()
    for c in _cats:
        cid = int(c.id or 0)
        g = combined_by_cat.get(cid)
        if not g or not g.get("items"):
            continue
        tag_panel_groups.append(g)
        seen_cat.add(cid)
    # Any leftover categories (deleted/inactive etc.)
    for cid, g in combined_by_cat.items():
        if cid in seen_cat:
            continue
        if g.get("items"):
            tag_panel_groups.append(g)

    # task assignees
    users_by_id = {u.id: u for u in session.exec(select(User)).all()}

    # per-message sender label (multi-agent; across history)
    sender_label_by_msgid: dict[int, str] = {}

    # build binding maps per platform for quick lookup
    binding_map_by_platform: dict[str, dict[str, int]] = {}
    try:
        from sqlalchemy import func as sa_func
        platforms = sorted({_platform_norm(c.platform) for c in thread_convs})
        for p in platforms:
            mp: dict[str, int] = {}
            for b in session.exec(select(AgentBinding).where(sa_func.lower(AgentBinding.platform) == p)).all():
                acc = str(getattr(b, "agent_account", "") or "").strip()
                if acc and getattr(b, "user_id", None):
                    mp[acc] = int(b.user_id)
            binding_map_by_platform[p] = mp
    except Exception:
        binding_map_by_platform = {}

    conv_platform_map: dict[int, str] = {}
    for c in thread_convs:
        if c.id is None:
            continue
        conv_platform_map[int(c.id)] = _platform_norm(c.platform)

    def _user_label(uid: int) -> str | None:
        u = users_by_id.get(uid)
        if not u:
            return None
        return f"{u.username} Â· {u.name}" if u.username else (u.name or u.email)

    for item in history_items:
        if item.get("type") != "msg":
            continue
        m = item.get("msg")
        if not m or m.id is None:
            continue
        mid = int(m.id)

        sender = (m.sender or "").lower()
        if sender == "buyer":
            sender_label_by_msgid[mid] = "buyer"
            continue
        if sender == "agent":
            cid = int(item.get("conv_id") or 0)
            p = conv_platform_map.get(cid, "unknown")
            acc = str(getattr(m, "agent_account", "") or "").strip()
            uid = binding_map_by_platform.get(p, {}).get(acc) or getattr(m, "agent_user_id", None)
            if uid:
                lab = _user_label(int(uid))
                if lab:
                    sender_label_by_msgid[mid] = lab
                    continue
            nick = str(getattr(m, "agent_nick", "") or "").strip()
            sender_label_by_msgid[mid] = nick or acc or "-"
            continue

        sender_label_by_msgid[mid] = m.sender or "system"

    # Header display: show all agent participants in today's conversation
    agents_display = ""
    try:
        labels = []
        seen = set()
        for m in msgs_today:
            if (m.sender or "").lower() != "agent":
                continue
            mid = int(m.id) if m.id is not None else None
            lab = sender_label_by_msgid.get(mid) if mid else None
            if not lab or lab == "-" or lab in seen:
                continue
            labels.append(lab)
            seen.add(lab)
        if labels:
            if len(labels) > 4:
                agents_display = " / ".join(labels[:4]) + f" â€¦(+{len(labels) - 4})"
            else:
                agents_display = " / ".join(labels)
    except Exception:
        agents_display = ""

    page_error = (request.query_params.get("error") or "").strip()

    return _template(
        request,
        "conversation.html",
        user=user,
        conv=conv,
        agent=agent,
        agents_display=agents_display,
        history_items=history_items,
        today_anchor_global=today_anchor_global,
        today_local_to_global=today_local_to_global,
        msgs=msgs,
        analyses=analyses,
        latest_analysis=latest_analysis,
        tasks=tasks,
        users_by_id=users_by_id,
        display_agent_label_by_conv=display_agent_label_by_conv,
        sender_label_by_msgid=sender_label_by_msgid,
        page_error=page_error,
        highlight_red_set=highlight_red_set,
        highlight_yellow_set=highlight_yellow_set,
        highlight_green_set=highlight_green_set,
        tag_groups=tag_groups,
        manual_tag_bindings=manual_tag_bindings,
        tag_panel_groups=tag_panel_groups,
        tags_for_add=tags_for_add,
        all_tag_categories=_cats,
        all_tags_by_cat=_tags_by_cat,
        hit_tags_map=hit_tags_map,
    )


def _build_leyan_jsonl(conv: Conversation, msgs: list[Message]) -> str:
    """Export messages back to a leyan-like JSONL (best-effort).

    Think of it like putting loose items back into the original packaging: we keep the shape (key fields + body
    as a JSON string) so you can reuse it downstream.
    """
    import json
    from datetime import timezone, timedelta

    tz = timezone(timedelta(hours=8))

    seller_id = str((conv.meta or {}).get("seller_id") or "")
    buyer_id = str(conv.buyer_id or "")
    buyer_nick = str((conv.meta or {}).get("buyer_nick") or "")
    assistant_nick = str((conv.meta or {}).get("assistant_nick") or (conv.agent_account or ""))
    assistant_id = str((conv.meta or {}).get("assistant_id") or "")

    out_lines: list[str] = []
    for m in msgs:
        sent_at = None
        if m.ts:
            dt = m.ts
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=tz)
            sent_at = dt.astimezone(tz).isoformat()

        sender = (m.sender or "system").lower()
        if sender == "buyer":
            sent_by = "buyer"
        elif sender == "agent":
            sent_by = "assistant"
        else:
            sent_by = "system"

        # Determine body format
        body_obj: dict
        attachments = m.attachments or []
        img = next((a for a in attachments if (a or {}).get("type") == "image" and (a or {}).get("url")), None)
        card = next((a for a in attachments if (a or {}).get("type") == "system_card" and (a or {}).get("url")), None)
        if img:
            body_obj = {"format": "IMAGE", "image_url": img.get("url")}
        elif card:
            body_obj = {"format": "SYSTEM_CARD", "card_url": card.get("url")}
        else:
            body_obj = {"format": "TEXT", "text": m.text or ""}

        assistant_id_local = assistant_id
        assistant_nick_local = assistant_nick
        if sent_by == "assistant":
            assistant_id_local = str(getattr(m, "agent_account", "") or "") or assistant_id
            assistant_nick_local = str(getattr(m, "agent_nick", "") or "") or assistant_nick

        row = {
            "message_id": str(getattr(m, "external_message_id", "") or m.id),
            "conversation_id": str(conv.external_id or conv.id),
            "seller_id": seller_id,
            "buyer_id": buyer_id,
            "assistant_id": assistant_id_local,
            "buyer_nick": buyer_nick,
            "assistant_nick": assistant_nick_local,
            "sent_at": sent_at,
            "sent_by": sent_by,
            "body": json.dumps(body_obj, ensure_ascii=False),
        }
        out_lines.append(json.dumps(row, ensure_ascii=False))

    return "\n".join(out_lines)



@app.get("/api/conversations/{conversation_id}/preview")
def conversation_preview_api(
    conversation_id: int,
    limit: int = 120,
    user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
):
    """Return a lightweight conversation preview for CID modal."""
    conv = session.get(Conversation, conversation_id)
    if not conv:
        raise HTTPException(status_code=404, detail="æ‰¾ä¸åˆ°å¯¹è¯")

    if not _can_view_conversation(user, session, conv):
        raise HTTPException(status_code=403, detail="æ— æƒé™æŸ¥çœ‹è¯¥å¯¹è¯")

    # clamp
    try:
        limit = int(limit)
    except Exception:
        limit = 120
    limit = max(20, min(500, limit))

    total = session.exec(select(func.count()).select_from(Message).where(Message.conversation_id == conv.id)).one()
    is_truncated = False

    # Show latest N messages for responsiveness
    q = (
        select(Message)
        .where(Message.conversation_id == conv.id)
        .order_by(Message.ts.desc().nulls_last(), Message.id.desc())
        .limit(limit)
    )
    msgs_desc = session.exec(q).all()
    msgs = list(reversed(msgs_desc))

    if total and total > limit:
        is_truncated = True

    def _fmt_ts(dt: datetime | None):
        if not dt:
            return None
        try:
            return dt.isoformat()
        except Exception:
            return str(dt)

    
    # sender display (for CID modal)
    users_by_id = {u.id: u for u in session.exec(select(User)).all()}
    binding_map: dict[str, int] = {}
    try:
        for b in session.exec(select(AgentBinding).where(AgentBinding.platform == (conv.platform or "").strip().lower())).all():
            binding_map[str(b.agent_account or "").strip()] = int(b.user_id)
    except Exception:
        binding_map = {}

    def _user_label(uid: int) -> str | None:
        u = users_by_id.get(uid)
        if not u:
            return None
        return f"{u.username} Â· {u.name}" if u.username else (u.name or u.email)

    def _sender_display(m: Message) -> str:
        sender = (m.sender or "").lower()
        if sender == "buyer":
            return "buyer"
        if sender == "agent":
            uid = getattr(m, "agent_user_id", None) or binding_map.get(str(getattr(m, "agent_account", "") or "").strip())
            if uid:
                lab = _user_label(int(uid))
                if lab:
                    return lab
            nick = str(getattr(m, "agent_nick", "") or "").strip()
            acc = str(getattr(m, "agent_account", "") or "").strip()
            return nick or acc or "agent"
        return m.sender or "system"

    messages_out = []
    for m in msgs:
        messages_out.append(
            {
                "id": m.id,
                "sender": m.sender,
                "sender_display": _sender_display(m),
                "ts": _fmt_ts(m.ts),
                "text": m.text or "",
                "attachments": m.attachments or [],
            }
        )

    return {
        "conversation": {
            "id": conv.id,
            "platform": conv.platform,
            "buyer_id": conv.buyer_id,
            "agent_account": conv.agent_account,
            "started_at": _fmt_ts(conv.started_at),
            "ended_at": _fmt_ts(conv.ended_at),
        },
        "total_messages": int(total or 0),
        "is_truncated": is_truncated,
        "messages": messages_out,
    }


# ===== Marllen assistant (floating AI butler) =====


def _fmt_dt_iso(dt: datetime | None) -> str | None:
    if not dt:
        return None
    try:
        # IMPORTANT: our DB timestamps are stored as naive UTC (datetime.utcnow()).
        # If we return a timezone-less ISO string, browsers parse it as local time,
        # which breaks elapsed-time UX (e.g. Shanghai shows +480 minutes).
        from datetime import timezone

        if getattr(dt, "tzinfo", None) is None:
            dt = dt.replace(tzinfo=timezone.utc)
        else:
            dt = dt.astimezone(timezone.utc)
        s = dt.isoformat()
        if s.endswith("+00:00"):
            s = s[:-6] + "Z"
        return s
    except Exception:
        return str(dt)

def _status_val(x) -> str:
    return str(getattr(x, "value", None) or x or "")

def _assistant_message_out(m: AssistantMessage) -> dict:
    role = (getattr(m, "role", None) or "user").strip().lower()
    out: dict = {
        "id": int(m.id or 0),
        "role": role,
        "content": m.content,
        "created_at": _fmt_dt_iso(m.created_at),
    }

    if role == "assistant":
        try:
            from rendering import render_markdown_safe
            out["content_html"] = render_markdown_safe(m.content or "")
        except Exception:
            out["content_html"] = None

        meta = m.meta or {}
        safe_meta: dict = {}
        for k in ("type", "source", "charts", "model", "reasoning_effort", "format"):
            if k in meta:
                safe_meta[k] = meta.get(k)
        if safe_meta:
            out["meta"] = safe_meta

    return out


class _AssistantSendIn(BaseModel):
    text: str = ""


@app.get("/api/marllen-assistant/threads")
def marllen_assistant_threads_api(
    limit: int = 50,
    user: User = Depends(require_role("admin", "supervisor", "agent")),
    session: Session = Depends(get_session),
):
    uid = int(user.id or 0)
    threads = list_marllen_threads(session, owner_user_id=uid, limit=limit)
    active_thread_id = None
    if threads:
        t0 = threads[0]
        if t0 and getattr(t0, "id", None) and not getattr(t0, "is_archived", False):
            active_thread_id = int(t0.id or 0)

    return {
        "assistant_name": MARLLEN_ASSISTANT_NAME,
        "active_thread_id": active_thread_id,
        "threads": [
            {
                "id": int(t.id or 0),
                "title": t.title,
                "is_archived": bool(getattr(t, "is_archived", False)),
                "created_at": _fmt_dt_iso(getattr(t, "created_at", None)),
                "updated_at": _fmt_dt_iso(getattr(t, "updated_at", None)),
            }
            for t in (threads or [])
        ],
    }


@app.get("/api/marllen-assistant/threads/{thread_id}")
def marllen_assistant_thread_detail_api(
    thread_id: int,
    limit: int = 80,
    user: User = Depends(require_role("admin", "supervisor", "agent")),
    session: Session = Depends(get_session),
):
    uid = int(user.id or 0)
    t = session.get(AssistantThread, int(thread_id))
    if not t or int(getattr(t, "owner_user_id", 0) or 0) != uid:
        raise HTTPException(status_code=404, detail="çº¿ç¨‹ä¸å­˜åœ¨")

    msgs = list_marllen_thread_messages(session, thread_id=int(thread_id), limit=limit)
    job = get_marllen_pending_job(session, thread_id=int(thread_id))

    return {
        "assistant_name": MARLLEN_ASSISTANT_NAME,
        "thread": {
            "id": int(getattr(t, "id", 0) or 0),
            "title": t.title,
            "is_archived": bool(getattr(t, "is_archived", False)),
            "created_at": _fmt_dt_iso(getattr(t, "created_at", None)),
            "updated_at": _fmt_dt_iso(getattr(t, "updated_at", None)),
        },
        "messages": [_assistant_message_out(m) for m in (msgs or [])],
        "pending_job": (
            {
                "id": int(job.id or 0),
                "status": _status_val(job.status),
                "created_at": _fmt_dt_iso(job.created_at),
                "started_at": _fmt_dt_iso(job.started_at),
                "finished_at": _fmt_dt_iso(job.finished_at),
                "attempts": int(getattr(job, "attempts", 0) or 0),
                "last_error": (job.last_error or "")[:2000],
            }
            if job
            else None
        ),
    }


@app.get("/api/marllen-assistant/thread")
def marllen_assistant_thread_api(
    limit: int = 80,
    user: User = Depends(require_role("admin", "supervisor", "agent")),
    session: Session = Depends(get_session),
):
    uid = int(user.id or 0)
    thread = get_or_create_active_thread(session, owner_user_id=uid)
    tid = int(thread.id or 0)

    msgs = list_marllen_thread_messages(session, thread_id=tid, limit=limit)
    job = get_marllen_pending_job(session, thread_id=tid)

    return {
        "assistant_name": MARLLEN_ASSISTANT_NAME,
        "thread": {
            "id": tid,
            "title": thread.title,
            "updated_at": _fmt_dt_iso(thread.updated_at),
        },
        "messages": [_assistant_message_out(m) for m in (msgs or [])],
        "pending_job": (
            {
                "id": int(job.id or 0),
                "status": _status_val(job.status),
                "created_at": _fmt_dt_iso(job.created_at),
                "started_at": _fmt_dt_iso(job.started_at),
                "finished_at": _fmt_dt_iso(job.finished_at),
                "attempts": int(getattr(job, "attempts", 0) or 0),
                "last_error": (job.last_error or "")[:2000],
            }
            if job
            else None
        ),
    }


@app.post("/api/marllen-assistant/threads/new")
def marllen_assistant_new_thread_api(
    user: User = Depends(require_role("admin", "supervisor", "agent")),
    session: Session = Depends(get_session),
):
    uid = int(user.id or 0)
    thread = archive_and_create_new_thread(session, owner_user_id=uid)
    tid = int(thread.id or 0)
    msgs = list_marllen_thread_messages(session, thread_id=tid, limit=80)

    return {
        "assistant_name": MARLLEN_ASSISTANT_NAME,
        "thread": {
            "id": tid,
            "title": thread.title,
            "updated_at": _fmt_dt_iso(thread.updated_at),
        },
        "messages": [_assistant_message_out(m) for m in (msgs or [])],
        "pending_job": None,
    }


@app.post("/api/marllen-assistant/threads/{thread_id}/messages")
def marllen_assistant_send_message_api(
    request: Request,
    thread_id: int,
    payload: _AssistantSendIn,
    user: User = Depends(require_role("admin", "supervisor", "agent")),
    session: Session = Depends(get_session),
):
    uid = int(user.id or 0)
    t = session.get(AssistantThread, int(thread_id))
    if not t or int(getattr(t, "owner_user_id", 0) or 0) != uid:
        raise HTTPException(status_code=404, detail="çº¿ç¨‹ä¸å­˜åœ¨")

    # Enforce "one question at a time" to keep conversation order coherent.
    inflight = session.exec(
        select(AssistantJob)
        .where(
            AssistantJob.thread_id == int(thread_id),
            AssistantJob.status.in_(["pending", "running"]),
        )
        .order_by(AssistantJob.created_at.desc(), AssistantJob.id.desc())
        .limit(1)
    ).first()
    if inflight:
        raise HTTPException(status_code=409, detail="ä¸Šä¸€æ¡é—®é¢˜è¿˜åœ¨å¤„ç†ä¸­ï¼Œè¯·ç­‰å¾…å›ç­”åå†æé—®")

    text = (payload.text or "").strip()
    if not text:
        raise HTTPException(status_code=400, detail="é—®é¢˜ä¸èƒ½ä¸ºç©º")
    if len(text) > 8000:
        raise HTTPException(status_code=400, detail="é—®é¢˜å¤ªé•¿ï¼ˆæœ€å¤š 8000 å­—ï¼‰")

    referer = (request.headers.get("referer") or "").strip()
    if len(referer) > 500:
        referer = referer[:500]

    msg = AssistantMessage(
        thread_id=int(thread_id),
        role="user",
        content=text,
        meta={"referer": referer},
    )
    session.add(msg)

    # Keep one active thread: sending to an archived thread re-activates it.
    try:
        activate_marllen_thread(session, owner_user_id=uid, thread=t)
        maybe_auto_title_marllen_thread(t, text)
    except Exception:
        # best-effort; do not block send
        pass

    session.commit()
    session.refresh(msg)

    job = AssistantJob(
        thread_id=int(thread_id),
        created_by_user_id=uid,
        status="pending",  # type: ignore
        user_message_id=int(msg.id or 0),
        extra={"referer": referer},
    )
    session.add(job)
    session.commit()
    session.refresh(job)

    return {
        "thread_id": int(thread_id),
        "user_message_id": int(msg.id or 0),
        "job_id": int(job.id or 0),
    }


@app.get("/api/marllen-assistant/jobs/{job_id}")
def marllen_assistant_job_api(
    job_id: int,
    user: User = Depends(require_role("admin", "supervisor", "agent")),
    session: Session = Depends(get_session),
):
    uid = int(user.id or 0)
    job = session.get(AssistantJob, int(job_id))
    if not job:
        raise HTTPException(status_code=404, detail="ä»»åŠ¡ä¸å­˜åœ¨")

    t = session.get(AssistantThread, int(job.thread_id))
    if not t or int(getattr(t, "owner_user_id", 0) or 0) != uid:
        raise HTTPException(status_code=404, detail="ä»»åŠ¡ä¸å­˜åœ¨")

    assistant_msg = None
    if getattr(job, "assistant_message_id", None):
        assistant_msg = session.get(AssistantMessage, int(job.assistant_message_id))

    return {
        "job": {
            "id": int(job.id or 0),
            "thread_id": int(job.thread_id or 0),
            "status": _status_val(job.status),
            "created_at": _fmt_dt_iso(job.created_at),
            "started_at": _fmt_dt_iso(job.started_at),
            "finished_at": _fmt_dt_iso(job.finished_at),
            "attempts": int(getattr(job, "attempts", 0) or 0),
            "last_error": (job.last_error or "")[:2000],
        },
        "assistant_message": (
            _assistant_message_out(assistant_msg)
            if assistant_msg
            else None
        ),
    }


@app.post("/api/marllen-assistant/jobs/{job_id}/cancel")
def marllen_assistant_job_cancel_api(
    job_id: int,
    user: User = Depends(require_role("admin", "supervisor", "agent")),
    session: Session = Depends(get_session),
):
    """Force-stop a stuck assistant job so the UI can recover.

    We reuse the existing JobStatus ("error") to avoid schema changes.
    """
    uid = int(user.id or 0)
    job = session.get(AssistantJob, int(job_id))
    if not job:
        raise HTTPException(status_code=404, detail="ä»»åŠ¡ä¸å­˜åœ¨")

    t = session.get(AssistantThread, int(job.thread_id))
    if not t or int(getattr(t, "owner_user_id", 0) or 0) != uid:
        raise HTTPException(status_code=404, detail="ä»»åŠ¡ä¸å­˜åœ¨")

    st = _status_val(job.status)
    if st not in ("done", "error"):
        job.status = "error"  # type: ignore
        job.finished_at = datetime.utcnow()
        job.last_error = "canceled by user"
        try:
            job.extra = {**(job.extra or {}), "canceled": 1, "canceled_at": datetime.utcnow().isoformat(), "canceled_by_user_id": uid}
        except Exception:
            pass
        session.add(job)
        session.commit()
        session.refresh(job)

    return {
        "job": {
            "id": int(job.id or 0),
            "thread_id": int(job.thread_id or 0),
            "status": _status_val(job.status),
            "created_at": _fmt_dt_iso(job.created_at),
            "started_at": _fmt_dt_iso(job.started_at),
            "finished_at": _fmt_dt_iso(job.finished_at),
            "attempts": int(getattr(job, "attempts", 0) or 0),
            "last_error": (job.last_error or "")[:2000],
        }
    }


@app.get("/conversations/{conversation_id}/export", response_class=HTMLResponse)
def conversation_export_page(
    request: Request,
    conversation_id: int,
    user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
):
    conv = session.get(Conversation, conversation_id)
    if not conv:
        return _template(request, "error.html", user=user, message="æ‰¾ä¸åˆ°å¯¹è¯")

    if not _can_view_conversation(user, session, conv):
        return _template(request, "error.html", user=user, message="æ— æƒé™æŸ¥çœ‹è¯¥å¯¹è¯")

    msgs = session.exec(select(Message).where(Message.conversation_id == conv.id).order_by(Message.ts.asc().nulls_last(), Message.id.asc())).all()
    jsonl_text = _build_leyan_jsonl(conv, msgs)

    images = []
    for m in msgs:
        for a in (m.attachments or []):
            if (a or {}).get("type") == "image" and (a or {}).get("url"):
                images.append({"url": a.get("url")})

    return _template(
        request,
        "conversation_export.html",
        user=user,
        conv=conv,
        jsonl_text=jsonl_text,
        images=images,
    )


@app.get("/conversations/{conversation_id}/export.jsonl")
def conversation_export_jsonl(
    conversation_id: int,
    user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
):
    from fastapi.responses import PlainTextResponse

    conv = session.get(Conversation, conversation_id)
    if not conv:
        return PlainTextResponse("not found", status_code=404)
    if not _can_view_conversation(user, session, conv):
        return PlainTextResponse("forbidden", status_code=403)


    msgs = session.exec(select(Message).where(Message.conversation_id == conv.id).order_by(Message.ts.asc().nulls_last(), Message.id.asc())).all()
    content = _build_leyan_jsonl(conv, msgs)
    headers = {
        "Content-Disposition": f"attachment; filename=conversation_{conversation_id}.jsonl"
    }
    return PlainTextResponse(content, media_type="text/plain; charset=utf-8", headers=headers)




@app.get("/conversations/{conversation_id}/export.txt")
def conversation_export_txt(
    conversation_id: int,
    user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
):
    from fastapi.responses import PlainTextResponse

    conv = session.get(Conversation, conversation_id)
    if not conv:
        return PlainTextResponse("not found", status_code=404)
    if not _can_view_conversation(user, session, conv):
        return PlainTextResponse("forbidden", status_code=403)

    msgs = session.exec(
        select(Message)
        .where(Message.conversation_id == conv.id)
        .order_by(Message.ts.asc().nulls_last(), Message.id.asc())
    ).all()

    users_by_id = {u.id: u for u in session.exec(select(User)).all()}

    # binding map for this platform
    binding_map: dict[str, int] = {}
    try:
        platform_norm = str(conv.platform or "unknown").strip().lower()
        for b in session.exec(select(AgentBinding).where(func.lower(AgentBinding.platform) == platform_norm)).all():
            acc = str(getattr(b, "agent_account", "") or "").strip()
            if acc and getattr(b, "user_id", None):
                binding_map[acc] = int(b.user_id)
    except Exception:
        binding_map = {}

    def _label(m: Message) -> str:
        sender = (m.sender or "").lower()
        if sender == "buyer":
            return "buyer"
        if sender == "agent":
            acc = str(getattr(m, "agent_account", "") or "").strip()
            uid = binding_map.get(acc) or getattr(m, "agent_user_id", None)
            if uid and users_by_id.get(int(uid)):
                u = users_by_id[int(uid)]
                return f"{u.username} Â· {u.name}" if u.username else (u.name or u.email)
            nick = str(getattr(m, "agent_nick", "") or "").strip()
            return nick or acc or "-"
        return m.sender or "system"

    def _msg_text(m: Message) -> str:
        t = (m.text or "").strip()
        if t:
            return t
        # fallback for attachments
        atts = m.attachments or []
        if any((a or {}).get("type") == "image" for a in atts):
            return "[å›¾ç‰‡]"
        if atts:
            return "[é™„ä»¶]"
        return ""

    lines: list[str] = []
    for m in msgs:
        ts = ""
        if m.ts:
            try:
                ts = m.ts.strftime("%Y-%m-%d %H:%M:%S")
            except Exception:
                ts = str(m.ts)
        lines.append(f"[{ts}] {_label(m)}: {_msg_text(m)}".rstrip())

    content = "\n".join(lines).strip() + "\n"
    headers = {"Content-Disposition": f"attachment; filename=conversation_{conversation_id}.txt"}
    return PlainTextResponse(content, media_type="text/plain; charset=utf-8", headers=headers)


@app.get("/conversations/{conversation_id}/export-user.txt")
def conversation_export_user_txt(
    conversation_id: int,
    user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
):
    from fastapi.responses import PlainTextResponse

    conv = session.get(Conversation, conversation_id)
    if not conv:
        return PlainTextResponse("not found", status_code=404)
    if not _can_view_conversation(user, session, conv):
        return PlainTextResponse("forbidden", status_code=403)

    # Find thread conversations (best-effort)
    thread_convs: list[Conversation] = []
    try:
        if getattr(conv, "user_thread_id", None):
            thread_convs = session.exec(
                select(Conversation)
                .where(Conversation.user_thread_id == conv.user_thread_id)
                .order_by(
                    Conversation.started_at.asc().nulls_last(),
                    Conversation.uploaded_at.asc().nulls_last(),
                    Conversation.id.asc(),
                )
            ).all()
    except Exception:
        thread_convs = []

    if not thread_convs:
        thread_convs = [conv]

    thread_conv_ids = [c.id for c in thread_convs if c.id is not None]

    # Display date per conversation (max buyer ts, fallback to started/uploaded)
    buyer_ts_map: dict[int, datetime] = {}
    try:
        rows = session.exec(
            select(Message.conversation_id, func.max(Message.ts))
            .where(
                Message.conversation_id.in_(thread_conv_ids),
                Message.sender == "buyer",
                Message.ts.is_not(None),
            )
            .group_by(Message.conversation_id)
        ).all()
        for cid, ts_max in rows:
            if cid is None or ts_max is None:
                continue
            buyer_ts_map[int(cid)] = ts_max
    except Exception:
        buyer_ts_map = {}

    def _display_ts_for_conv(c: Conversation) -> datetime | None:
        if c.id is not None and int(c.id) in buyer_ts_map:
            return buyer_ts_map[int(c.id)]
        return c.started_at or c.uploaded_at

    def _date_label_for_conv(c: Conversation) -> str:
        dt = _display_ts_for_conv(c)
        if not dt:
            return "æœªçŸ¥æ—¥æœŸ"
        try:
            return dt.strftime("%Y-%m-%d")
        except Exception:
            return str(dt)[:10]

    users_by_id = {u.id: u for u in session.exec(select(User)).all()}

    # binding maps per platform
    binding_map_by_platform: dict[str, dict[str, int]] = {}
    try:
        platforms = sorted({str(c.platform or "unknown").strip().lower() or "unknown" for c in thread_convs})
        for p in platforms:
            mp: dict[str, int] = {}
            for b in session.exec(select(AgentBinding).where(func.lower(AgentBinding.platform) == p)).all():
                acc = str(getattr(b, "agent_account", "") or "").strip()
                if acc and getattr(b, "user_id", None):
                    mp[acc] = int(b.user_id)
            binding_map_by_platform[p] = mp
    except Exception:
        binding_map_by_platform = {}

    def _label(c: Conversation, m: Message) -> str:
        sender = (m.sender or "").lower()
        if sender == "buyer":
            return "buyer"
        if sender == "agent":
            p = str(c.platform or "unknown").strip().lower() or "unknown"
            acc = str(getattr(m, "agent_account", "") or "").strip()
            uid = binding_map_by_platform.get(p, {}).get(acc) or getattr(m, "agent_user_id", None)
            if uid and users_by_id.get(int(uid)):
                u = users_by_id[int(uid)]
                return f"{u.username} Â· {u.name}" if u.username else (u.name or u.email)
            nick = str(getattr(m, "agent_nick", "") or "").strip()
            return nick or acc or "-"
        return m.sender or "system"

    def _msg_text(m: Message) -> str:
        t = (m.text or "").strip()
        if t:
            return t
        atts = m.attachments or []
        if any((a or {}).get("type") == "image" for a in atts):
            return "[å›¾ç‰‡]"
        if atts:
            return "[é™„ä»¶]"
        return ""

    lines: list[str] = []
    for c in thread_convs:
        if not c.id:
            continue
        label = _date_label_for_conv(c)
        lines.append(f"===== {label} (CID={c.id}) =====")
        msgs = session.exec(
            select(Message)
            .where(Message.conversation_id == int(c.id))
            .order_by(Message.ts.asc().nulls_last(), Message.id.asc())
        ).all()
        for m in msgs:
            ts = ""
            if m.ts:
                try:
                    ts = m.ts.strftime("%Y-%m-%d %H:%M:%S")
                except Exception:
                    ts = str(m.ts)
            lines.append(f"[{ts}] {_label(c, m)}: {_msg_text(m)}".rstrip())
        lines.append("")  # blank line between days

    content = "\n".join(lines).strip() + "\n"
    headers = {"Content-Disposition": f"attachment; filename=user_thread_{conversation_id}.txt"}
    return PlainTextResponse(content, media_type="text/plain; charset=utf-8", headers=headers)



@app.post("/conversations/{conversation_id}/reanalyze")
def conversation_reanalyze(
    request: Request,
    conversation_id: int,
    user: User = Depends(require_role("admin", "supervisor")),
    session: Session = Depends(get_session),
):
    conv = session.get(Conversation, conversation_id)
    if not conv:
        return _template(request, "error.html", user=user, message="æ‰¾ä¸åˆ°å¯¹è¯")

    # Always enqueue a new job (overwrite mode). If there's already a pending/running job for this conversation,
    # we won't create duplicates.
    exists = session.exec(
        select(AIAnalysisJob).where(
            AIAnalysisJob.conversation_id == conversation_id,
            AIAnalysisJob.status.in_(["pending", "running"]),
        )
    ).first()
    if not exists:
        session.add(
            AIAnalysisJob(
                conversation_id=conversation_id,
                extra={"overwrite": True, "requested_by": user.id, "requested_at": datetime.utcnow().isoformat()},
            )
        )
        session.commit()

    return _redirect(f"/conversations/{conversation_id}")


@app.post("/conversations/{conversation_id}/tags/add")
def add_conversation_tag(
    request: Request,
    conversation_id: int,
    user: User = Depends(require_role("admin", "supervisor")),
    session: Session = Depends(get_session),
    tag_id: str = Form(""),
    reason: str = Form(""),
    start_index: str = Form(""),
    end_index: str = Form(""),
):
    conv = session.get(Conversation, conversation_id)
    if not conv:
        return _template(request, "error.html", user=user, message="æ‰¾ä¸åˆ°å¯¹è¯")
    tag_id_norm = (tag_id or "").strip()
    try:
        tag_id_int = int(tag_id_norm)
    except Exception:
        tag_id_int = 0
    if not tag_id_int:
        return _redirect(f"/conversations/{conversation_id}?error=è¯·é€‰æ‹©æ ‡ç­¾")
    tag = session.get(TagDefinition, int(tag_id_int))
    if not tag or not getattr(tag, "is_active", True):
        return _redirect(f"/conversations/{conversation_id}?error=æ ‡ç­¾ä¸å­˜åœ¨æˆ–å·²åœç”¨")

    # å¿…é¡»é€‰æ‹©å¼•ç”¨èŒƒå›´ï¼ˆä»…å½“æ—¥ï¼‰
    if not (start_index or "").strip() or not (end_index or "").strip():
        return _redirect(f"/conversations/{conversation_id}?error=è¯·å…ˆåœ¨å·¦ä¾§é€‰æ‹©å½“æ—¥æ¶ˆæ¯ä½œä¸ºå¼•ç”¨èŒƒå›´")
    try:
        si = int((start_index or "").strip())
        ei = int((end_index or "").strip())
    except Exception:
        return _redirect(f"/conversations/{conversation_id}?error=å¼•ç”¨èŒƒå›´ä¸åˆæ³•")

    if si < 0 or ei < 0:
        return _redirect(f"/conversations/{conversation_id}?error=å¼•ç”¨èŒƒå›´ä¸åˆæ³•")
    if ei < si:
        si, ei = ei, si

    # Validate indices against today's conversation messages (local indices, 0-based)
    msg_count = len(
        session.exec(
            select(Message.id)
            .where(Message.conversation_id == conversation_id)
        ).all()
    )
    if msg_count <= 0:
        return _redirect(f"/conversations/{conversation_id}?error=å½“æ—¥å¯¹è¯æš‚æ— æ¶ˆæ¯ï¼Œæ— æ³•å¼•ç”¨")
    if si >= msg_count or ei >= msg_count:
        return _redirect(f"/conversations/{conversation_id}?error=å¼•ç”¨èŒƒå›´è¶…å‡ºå½“æ—¥æ¶ˆæ¯èŒƒå›´")

    existing = session.exec(
        select(ManualTagBinding).where(
            ManualTagBinding.conversation_id == conversation_id,
            ManualTagBinding.tag_id == int(tag_id_int),
        )
    ).first()
    if not existing:
        evidence_list = [{
            "message_index": si,
            "start_index": si,
            "end_index": ei,
        }]
        evidence = {"evidence": evidence_list}
        
        session.add(
            ManualTagBinding(
                conversation_id=conversation_id,
                tag_id=int(tag_id_int),
                created_by_user_id=user.id,
                reason=(reason or "").strip()[:500],
                evidence=evidence,
            )
        )
        session.commit()
    return _redirect(f"/conversations/{conversation_id}")


@app.post("/conversations/{conversation_id}/tags/remove")
def remove_conversation_tag(
    conversation_id: int,
    user: User = Depends(require_role("admin", "supervisor")),
    session: Session = Depends(get_session),
    tag_id: int = Form(...),
):
    conv = session.get(Conversation, conversation_id)
    if not conv:
        return _redirect("/conversations")
    session.exec(delete(ManualTagBinding).where(
        ManualTagBinding.conversation_id == conversation_id,
        ManualTagBinding.tag_id == tag_id,
    ))
    session.commit()
    return _redirect(f"/conversations/{conversation_id}")


@app.post("/conversations/{conversation_id}/tags/ai/remove")
def remove_ai_tag(
    conversation_id: int,
    user: User = Depends(require_role("admin", "supervisor")),
    session: Session = Depends(get_session),
    hit_id: int = Form(...),
):
    """åˆ é™¤AIè‡ªåŠ¨ç”Ÿæˆçš„æ ‡ç­¾å‘½ä¸­è®°å½•"""
    conv = session.get(Conversation, conversation_id)
    if not conv:
        return _redirect("/conversations")
    
    # éªŒè¯hit_idæ˜¯å¦å±äºè¯¥å¯¹è¯çš„åˆ†æç»“æœ
    hit = session.get(ConversationTagHit, hit_id)
    if not hit:
        return _redirect(f"/conversations/{conversation_id}")
    
    analysis = session.get(ConversationAnalysis, hit.analysis_id)
    if not analysis or analysis.conversation_id != conversation_id:
        return _redirect(f"/conversations/{conversation_id}")
    
    session.exec(delete(ConversationTagHit).where(ConversationTagHit.id == hit_id))
    session.commit()
    return _redirect(f"/conversations/{conversation_id}")


def _qc_conversations_in_range(
    session: Session,
    start_date: str,
    end_date: str,
    threshold_messages: int,
) -> list[tuple[Conversation, int]]:
    """Return [(Conversation, msg_count), ...] for date range with msg_count >= threshold (excl. system)."""
    from sqlalchemy import func

    # Parse as local date strings (YYYY-MM-DD). If parsing fails, fall back to "today".
    try:
        start_dt = datetime.strptime(start_date, "%Y-%m-%d")
    except Exception:
        start_dt = datetime.strptime(_today_shanghai_str(), "%Y-%m-%d")

    try:
        end_dt_inclusive = datetime.strptime(end_date, "%Y-%m-%d")
    except Exception:
        end_dt_inclusive = start_dt

    if end_dt_inclusive < start_dt:
        start_dt, end_dt_inclusive = end_dt_inclusive, start_dt

    end_dt = end_dt_inclusive + timedelta(days=1)

    convs = session.exec(
        select(Conversation)
        .where(
            ((Conversation.started_at >= start_dt) & (Conversation.started_at < end_dt))
            | (
                (Conversation.started_at.is_(None))
                & (Conversation.uploaded_at >= start_dt)
                & (Conversation.uploaded_at < end_dt)
            )
        )
        .order_by(Conversation.id.asc())
    ).all()

    if not convs:
        return []

    conv_ids = [int(c.id) for c in convs if c.id]
    if not conv_ids:
        return []

    rows = session.exec(
        select(Message.conversation_id, func.count(Message.id))
        .where(
            Message.conversation_id.in_(conv_ids),
            Message.sender != "system",
        )
        .group_by(Message.conversation_id)
    ).all()
    msg_map = {int(cid): int(cnt) for (cid, cnt) in rows}

    out: list[tuple[Conversation, int]] = []
    for c in convs:
        if not c.id:
            continue
        cnt = msg_map.get(int(c.id), 0)
        if cnt >= threshold_messages:
            out.append((c, cnt))
    return out


@app.get("/qc/daily", response_class=HTMLResponse)
def qc_daily_page(
    request: Request,
    date: str = "",
    start_date: str = "",
    end_date: str = "",
    min_messages: str = "",
    has_analysis: str = "",
    page: str = "",
    per_page: str = "",
    user: User = Depends(require_role("admin", "supervisor")),
    session: Session = Depends(get_session),
):
    # Back-compat: if "date" provided, treat it as a single-day range.
    single = (date or "").strip()
    start_s = (start_date or "").strip() or single or _today_shanghai_str()
    end_s = (end_date or "").strip() or single or start_s

    try:
        threshold = int((min_messages or "").strip() or 5)
    except Exception:
        threshold = 5
    threshold = max(1, min(500, threshold))

    # åˆ†é¡µå‚æ•°
    try:
        current_page = int((page or "").strip() or 1)
    except Exception:
        current_page = 1
    current_page = max(1, current_page)

    try:
        items_per_page = int((per_page or "").strip() or 50)
    except Exception:
        items_per_page = 50
    items_per_page = max(10, min(200, items_per_page))

    pairs = _qc_conversations_in_range(session, start_date=start_s, end_date=end_s, threshold_messages=threshold)
    convos = [p[0] for p in pairs]
    rounds_by_conv = {p[0].id: p[1] for p in pairs if p[0].id}

    conv_ids = [int(c.id) for c in convos if c.id]
    has_analysis_norm = (has_analysis or "").strip()
    if conv_ids and has_analysis_norm in ("1", "0"):
        analyzed_ids = set(
            session.exec(
                select(ConversationAnalysis.conversation_id)
                .where(ConversationAnalysis.conversation_id.in_(conv_ids))
                .distinct()
            ).all()
        )
        if has_analysis_norm == "1":
            pairs = [p for p in pairs if int(p[0].id) in analyzed_ids]
        else:
            pairs = [p for p in pairs if int(p[0].id) not in analyzed_ids]
        convos = [p[0] for p in pairs]
        rounds_by_conv = {p[0].id: p[1] for p in pairs if p[0].id}
        conv_ids = [int(c.id) for c in convos if c.id]

    # è®¡ç®—æ€»æ•°å’Œåˆ†é¡µ
    total_count = len(convos)
    total_message_count = sum(rounds_by_conv.get(c.id, 0) for c in convos if c.id)
    total_pages = (total_count + items_per_page - 1) // items_per_page if total_count > 0 else 1
    current_page = min(current_page, total_pages)

    # åˆ†é¡µåˆ‡ç‰‡
    start_idx = (current_page - 1) * items_per_page
    end_idx = start_idx + items_per_page
    convos_page = convos[start_idx:end_idx]
    conv_ids_page = [int(c.id) for c in convos_page if c.id]
    page_message_count = sum(rounds_by_conv.get(c.id, 0) for c in convos_page if c.id)

    pending = set()
    if conv_ids_page:
        pending = set(
            session.exec(
                select(AIAnalysisJob.conversation_id).where(
                    AIAnalysisJob.conversation_id.in_(conv_ids_page),
                    AIAnalysisJob.status.in_(["pending", "running"]),
                )
            ).all()
        )

    return _template(
        request,
        "qc_daily.html",
        user=user,
        start_date=start_s,
        end_date=end_s,
        threshold_messages=threshold,
        has_analysis=has_analysis_norm,
        convos=convos_page,
        rounds_by_conv=rounds_by_conv,
        pending_ids=pending,
        current_page=current_page,
        total_pages=total_pages,
        total_count=total_count,
        total_message_count=total_message_count,
        page_message_count=page_message_count,
        per_page=items_per_page,
        ok=request.query_params.get("ok"),
        error=request.query_params.get("error"),
    )


@app.post("/qc/daily/batch")
async def qc_daily_batch_analyze(
    request: Request,
    user: User = Depends(require_role("admin", "supervisor")),
    session: Session = Depends(get_session),
):
    form = await request.form()
    start_s = (form.get("start_date") or "").strip() or _today_shanghai_str()
    end_s = (form.get("end_date") or "").strip() or start_s
    try:
        threshold = int((form.get("min_messages") or "").strip() or 5)
    except Exception:
        threshold = 5
    threshold = max(1, min(500, threshold))

    # åˆ†é¡µå‚æ•°
    page_s = (form.get("page") or "").strip() or "1"
    per_page_s = (form.get("per_page") or "").strip() or "50"

    raw_list = form.getlist("conversation_ids")
    ids: list[int] = []
    for raw in raw_list:
        s = (raw or "").strip()
        if not s:
            continue
        try:
            ids.append(int(s))
        except ValueError:
            continue

    if not ids:
        return _redirect(f"/qc/daily?start_date={start_s}&end_date={end_s}&min_messages={threshold}&page={page_s}&per_page={per_page_s}&error=è¯·è‡³å°‘å‹¾é€‰ä¸€æ¡å¯¹è¯")

    existing = set(
        session.exec(
            select(AIAnalysisJob.conversation_id).where(
                AIAnalysisJob.conversation_id.in_(ids),
                AIAnalysisJob.status.in_(["pending", "running"]),
            )
        ).all()
    )
    created = 0
    for cid in ids:
        if cid in existing:
            continue
        session.add(AIAnalysisJob(conversation_id=cid, extra={"source": "qc_daily_batch"}))
        created += 1
        existing.add(cid)
    if created:
        session.commit()

    return _redirect(f"/qc/daily?start_date={start_s}&end_date={end_s}&min_messages={threshold}&page={page_s}&per_page={per_page_s}&ok=å·²å…¥é˜Ÿ{created}æ¡è´¨æ£€ä»»åŠ¡")


@app.get("/tasks", response_class=HTMLResponse)
def tasks_list(
    request: Request,
    user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
):
    if user.role in [Role.admin, Role.supervisor]:
        tasks = session.exec(select(TrainingTask).order_by(TrainingTask.created_at.desc()).limit(200)).all()
    else:
        tasks = session.exec(select(TrainingTask).where(TrainingTask.assigned_to_user_id == user.id).order_by(TrainingTask.created_at.desc()).limit(200)).all()

    conv_ids = [t.conversation_id for t in tasks]
    convos = {c.id: c for c in session.exec(select(Conversation).where(Conversation.id.in_(conv_ids))).all()} if conv_ids else {}
    users = {u.id: u for u in session.exec(select(User)).all()}

    return _template(request, "tasks.html", user=user, tasks=tasks, convos=convos, users=users)


@app.post("/tasks/new")
def create_task(
    request: Request,
    conversation_id: int = Form(...),
    assigned_to_user_id: int = Form(...),
    notes: str = Form(""),
    user: User = Depends(require_role("admin", "supervisor")),
    session: Session = Depends(get_session),
):
    conv = session.get(Conversation, conversation_id)
    if not conv:
        return _template(request, "error.html", user=user, message="æ‰¾ä¸åˆ°å¯¹è¯")

    a = session.exec(
        select(ConversationAnalysis)
        .where(ConversationAnalysis.conversation_id == conv.id)
        .order_by(ConversationAnalysis.id.desc())
    ).first()
    focus = _pick_default_focus_tag(a)

    t = TrainingTask(
        conversation_id=conversation_id,
        created_by_user_id=user.id,
        assigned_to_user_id=assigned_to_user_id,
        notes=notes,
        focus_negative_tag=focus,
        status=TaskStatus.open,
    )
    session.add(t)
    session.commit()
    return _redirect(f"/tasks/{t.id}")


@app.get("/tasks/{task_id}", response_class=HTMLResponse)
def task_detail(
    request: Request,
    task_id: int,
    user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
):
    task = session.get(TrainingTask, task_id)
    if not task:
        return _template(request, "error.html", user=user, message="æ‰¾ä¸åˆ°ä»»åŠ¡")

    # permissions
    if user.role == Role.agent and task.assigned_to_user_id != user.id:
        return _template(request, "error.html", user=user, message="æ— æƒé™æŸ¥çœ‹è¯¥ä»»åŠ¡")

    conv = session.get(Conversation, task.conversation_id)
    msgs = session.exec(select(Message).where(Message.conversation_id == conv.id).order_by(Message.ts.asc().nulls_last(), Message.id.asc())).all()
    # åªæŸ¥è¯¢æœ€æ–°çš„ä¸€æ¡ AI åˆ†æ
    latest_analysis = session.exec(select(ConversationAnalysis).where(ConversationAnalysis.conversation_id == conv.id).order_by(ConversationAnalysis.id.desc()).limit(1)).first()
    analyses = [latest_analysis] if latest_analysis else []
    attempts = session.exec(select(TrainingAttempt).where(TrainingAttempt.task_id == task.id).order_by(TrainingAttempt.created_at.desc())).all()

    reflections = session.exec(select(TrainingReflection).where(TrainingReflection.task_id == task.id).order_by(TrainingReflection.created_at.desc())).all()

    sim = session.exec(select(TrainingSimulation).where(TrainingSimulation.task_id == task.id, TrainingSimulation.user_id == user.id)).first()
    sim_messages = []
    if sim:
        sim_messages = session.exec(select(TrainingSimulationMessage).where(TrainingSimulationMessage.simulation_id == sim.id).order_by(TrainingSimulationMessage.created_at.asc(), TrainingSimulationMessage.id.asc())).all()

    users = {u.id: u for u in session.exec(select(User)).all()}

    return _template(
        request,
        "task_detail.html",
        user=user,
        task=task,
        conv=conv,
        msgs=msgs,
        analyses=analyses,
        latest_analysis=latest_analysis,
        attempts=attempts,
        reflections=reflections,
        simulation=sim,
        simulation_messages=sim_messages,
        users=users,
    )


@app.post("/tasks/{task_id}/submit")
def submit_attempt(
    request: Request,
    task_id: int,
    attempt_text: str = Form(...),
    user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
):
    task = session.get(TrainingTask, task_id)
    if not task:
        return _template(request, "error.html", user=user, message="æ‰¾ä¸åˆ°ä»»åŠ¡")

    if user.role == Role.agent and task.assigned_to_user_id != user.id:
        return _template(request, "error.html", user=user, message="æ— æƒé™æäº¤")

    session.add(TrainingAttempt(task_id=task.id, attempt_by_user_id=user.id, attempt_text=attempt_text.strip()))
    task.status = TaskStatus.submitted
    session.add(task)
    session.commit()

    return _redirect(f"/tasks/{task.id}")


@app.post("/tasks/{task_id}/review")
def review_attempt(
    request: Request,
    task_id: int,
    attempt_id: int = Form(...),
    score: int = Form(...),
    review_notes: str = Form(""),
    user: User = Depends(require_role("admin", "supervisor")),
    session: Session = Depends(get_session),
):
    task = session.get(TrainingTask, task_id)
    if not task:
        return _template(request, "error.html", user=user, message="æ‰¾ä¸åˆ°ä»»åŠ¡")

    attempt = session.get(TrainingAttempt, attempt_id)
    if not attempt or attempt.task_id != task.id:
        return _template(request, "error.html", user=user, message="æ‰¾ä¸åˆ°æäº¤è®°å½•")

    attempt.review_score = int(score)
    attempt.review_notes = review_notes.strip()
    attempt.reviewed_by_user_id = user.id
    from datetime import datetime, timedelta

    attempt.reviewed_at = datetime.utcnow()
    session.add(attempt)

    task.status = TaskStatus.reviewed
    session.add(task)
    session.commit()

    return _redirect(f"/tasks/{task.id}")


@app.post("/tasks/{task_id}/close")
def close_task(
    request: Request,
    task_id: int,
    user: User = Depends(require_role("admin", "supervisor")),
    session: Session = Depends(get_session),
):
    task = session.get(TrainingTask, task_id)
    if not task:
        return _template(request, "error.html", user=user, message="æ‰¾ä¸åˆ°ä»»åŠ¡")
    task.status = TaskStatus.closed
    session.add(task)
    session.commit()
    return _redirect("/tasks")

# =========================
# Reports (å›´ç»•4ä¸ªæ ¸å¿ƒï¼šæ­£å‘/è´Ÿå‘/å¤ç›˜åŸ¹è®­/VOC)
# =========================

@app.get("/reports", response_class=HTMLResponse)
def reports(
    request: Request,
    user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
    start: str | None = None,
    end: str | None = None,
    platform: str | None = None,
    agent_id: str | None = None,
    tag: str | None = None,
):
    """æƒé™å†…çš„æŠ¥è¡¨ã€‚

    ç»´åº¦ï¼šæ—¶é—´èŒƒå›´ã€å›¢é˜Ÿ/è´¦å·ã€å¹³å°ã€‚
    """

    from datetime import datetime, timedelta

    def _parse_date(d: str | None):
        if not d:
            return None
        d = d.strip()
        if not d:
            return None
        try:
            parts = d.split("-")
            if len(parts) != 3:
                return None
            y, m, day = int(parts[0]), int(parts[1]), int(parts[2])
            return datetime(y, m, day)
        except Exception:
            return None

    def _week_bounds(d: datetime) -> tuple[datetime, datetime]:
        week_start = d - timedelta(days=d.weekday())
        week_end = week_start + timedelta(days=6)
        return week_start, week_end

    def _shift_year_safe(d: datetime, years: int) -> datetime:
        try:
            return d.replace(year=d.year + years)
        except ValueError:
            # Handle Feb 29 for non-leap years by falling back to Feb 28
            return d.replace(year=d.year + years, day=28)
    
    start_dt = _parse_date(start)
    end_dt = _parse_date(end)
    end_dt_excl = (end_dt + timedelta(days=1)) if end_dt else None

    stmt = select(Conversation)

    # scope
    if user.role == Role.agent:
        stmt = stmt.where(Conversation.agent_user_id == user.id)
    else:
        agent_id_int: int | None = None
        if agent_id is not None:
            s = (agent_id or "").strip()
            if s:
                try:
                    agent_id_int = int(s)
                except Exception:
                    agent_id_int = None
        if agent_id_int:
            bs = session.exec(select(AgentBinding).where(AgentBinding.user_id == agent_id_int)).all()
            or_terms = [Conversation.agent_user_id == agent_id_int]
            for b in bs:
                # primary agent_account match (legacy)
                or_terms.append(and_(Conversation.platform == b.platform, Conversation.agent_account == b.agent_account))
                # multi-agent: any agent account appearing in messages
                or_terms.append(
                    exists(
                        select(1)
                        .select_from(Message)
                        .where(
                            Message.conversation_id == Conversation.id,
                            Message.sender == "agent",
                            Message.agent_account == b.agent_account,
                        )
                    )
                )
            stmt = stmt.where(or_(*or_terms))

    if platform and platform.strip():
        stmt = stmt.where(Conversation.platform == platform.strip())

    if start_dt:
        stmt = stmt.where(Conversation.uploaded_at >= start_dt)
    if end_dt_excl:
        stmt = stmt.where(Conversation.uploaded_at < end_dt_excl)

    stmt = stmt.order_by(Conversation.uploaded_at.desc(), Conversation.id.desc()).limit(5000)
    convos = session.exec(stmt).all()

    convo_ids = [c.id for c in convos if c.id]

    # latest analysis per conversation
    latest_by_conv: dict[int, ConversationAnalysis | None] = {cid: None for cid in convo_ids}
    if convo_ids:
        rows = session.exec(
            select(ConversationAnalysis)
            .where(ConversationAnalysis.conversation_id.in_(convo_ids))
            .order_by(ConversationAnalysis.conversation_id.asc(), ConversationAnalysis.id.desc())
        ).all()
        for r in rows:
            if r.conversation_id in latest_by_conv and latest_by_conv[r.conversation_id] is None:
                latest_by_conv[r.conversation_id] = r

    tag_q = (tag or "").strip().lower()

    def _tags_text(a: ConversationAnalysis | None) -> str:
        if not a:
            return ""
        return " ".join(
            [
                a.dialog_type or "",
                a.day_summary or "",
                a.tag_update_suggestion or "",
            ]
        ).lower()

    # apply tag filter (if any)
    if tag_q:
        convos = [c for c in convos if tag_q in _tags_text(latest_by_conv.get(c.id))]

    analyses = [latest_by_conv.get(c.id) for c in convos]

    # ç®€åŒ–æ­£å‘/è´Ÿå‘åˆ¤æ–­ï¼šä»…æ ¹æ® flag_for_review
    def _has_positive(a: ConversationAnalysis | None) -> bool:
        # æœ‰åˆ†æä¸”æœªæ ‡è®°éœ€å¤æ ¸åˆ™è§†ä¸ºæ­£å‘
        return a is not None and not a.flag_for_review

    def _has_negative(a: ConversationAnalysis | None) -> bool:
        # æ ‡è®°éœ€å¤æ ¸åˆ™è§†ä¸ºè´Ÿå‘
        return a is not None and a.flag_for_review

    positive_convos = []
    negative_convos = []

    for c in convos:
        a = latest_by_conv.get(c.id)
        if _has_positive(a):
            positive_convos.append(c)
        if _has_negative(a):
            negative_convos.append(c)

    # Training stats
    task_stmt = select(TrainingTask)
    if user.role == Role.agent:
        task_stmt = task_stmt.where(TrainingTask.assigned_to_user_id == user.id)
    else:
        if agent_id:
            task_stmt = task_stmt.where(TrainingTask.assigned_to_user_id == agent_id)
    if convo_ids:
        task_stmt = task_stmt.where(TrainingTask.conversation_id.in_(convo_ids))
    tasks = session.exec(task_stmt.order_by(TrainingTask.created_at.desc()).limit(2000)).all()

    task_ids = [t.id for t in tasks if t.id]
    reflections = session.exec(
        select(TrainingReflection).where(TrainingReflection.task_id.in_(task_ids)).order_by(TrainingReflection.created_at.desc())
    ).all() if task_ids else []

    passed_cnt = sum(1 for r in reflections if r.ai_passed is True)
    reviewed_cnt = sum(1 for t in tasks if t.status in [TaskStatus.reviewed, TaskStatus.closed])

    platforms = sorted({c.platform for c in convos if c.platform})

    users = session.exec(select(User).order_by(User.role.asc(), User.name.asc())).all()
    # ä¸»ç®¡ä¹Ÿå…è®¸ç»‘å®šæ¥å¾…å®¢æˆ·
    agents = [u for u in users if u.role in (Role.agent, Role.supervisor) and getattr(u, 'is_active', True)]
    users_by_id = {u.id: u for u in users}

    return _template(
        request,
        "reports.html",
        user=user,
        convos=convos,
        latest_by_conv=latest_by_conv,
        users_by_id=users_by_id,
        agents=agents,
        platforms=platforms,
        filters={
            "start": start or "",
            "end": end or "",
            "platform": (platform or "").strip(),
            "agent_id": agent_id_int if user.role != Role.agent else user.id,
            "tag": tag or "",
        },
        stats={
            "total": len(convos),
            "positive": len(positive_convos),
            "negative": len(negative_convos),
            "tasks": len(tasks),
            "tasks_reviewed": reviewed_cnt,
            "reflections": len(reflections),
            "reflections_passed": passed_cnt,
        },
        positive_convos=positive_convos[:50],
        negative_convos=negative_convos[:50],
        tasks=tasks[:50],
    )


def _tag_report_parse_date(d: str | None) -> datetime | None:
    if not d:
        return None
    d = d.strip()
    if not d:
        return None
    try:
        y, m, day = [int(x) for x in d.split("-")]
        return datetime(y, m, day)
    except Exception:
        return None


def _tag_report_normalize_range(start_dt: datetime | None, end_dt: datetime | None) -> tuple[datetime | None, datetime | None]:
    if not start_dt and not end_dt:
        return None, None
    if start_dt and not end_dt:
        end_dt = start_dt
    if end_dt and not start_dt:
        start_dt = end_dt
    if start_dt and end_dt and start_dt > end_dt:
        start_dt, end_dt = end_dt, start_dt
    return start_dt, end_dt


def _tag_report_week_bounds(d: datetime) -> tuple[datetime, datetime]:
    monday = d - timedelta(days=d.weekday())
    sunday = monday + timedelta(days=6)
    return datetime(monday.year, monday.month, monday.day), datetime(sunday.year, sunday.month, sunday.day)


def _tag_report_month_bounds(d: datetime) -> tuple[datetime, datetime]:
    start = datetime(d.year, d.month, 1)
    if d.month == 12:
        end = datetime(d.year, 12, 31)
    else:
        end = datetime(d.year, d.month + 1, 1) - timedelta(days=1)
    return start, end


def _tag_report_quarter_bounds(d: datetime) -> tuple[datetime, datetime]:
    quarter = (d.month - 1) // 3 + 1
    start_month = (quarter - 1) * 3 + 1
    start = datetime(d.year, start_month, 1)
    end_month = start_month + 2
    if end_month == 12:
        end = datetime(d.year, 12, 31)
    else:
        end = datetime(d.year, end_month + 1, 1) - timedelta(days=1)
    return start, end


def _tag_report_year_bounds(d: datetime) -> tuple[datetime, datetime]:
    return datetime(d.year, 1, 1), datetime(d.year, 12, 31)


def _tag_report_shift_year_safe(d: datetime, years: int) -> datetime:
    try:
        return d.replace(year=d.year + years)
    except ValueError:
        # Handle Feb 29 for non-leap years by falling back to Feb 28
        return d.replace(year=d.year + years, day=28)


def _build_tag_report_period(
    time_mode: str | None,
    start: str | None,
    end: str | None,
    yoy: bool,
) -> dict[str, object]:
    mode = (time_mode or "last_week").strip() or "last_week"
    start_dt = _tag_report_parse_date(start)
    end_dt = _tag_report_parse_date(end)
    now = datetime.utcnow()
    today = datetime(now.year, now.month, now.day)

    def _date_str(d: datetime | None) -> str:
        return d.date().isoformat() if d else ""

    start_dt, end_dt = _tag_report_normalize_range(start_dt, end_dt)
    anchor = start_dt or end_dt

    if mode == "all":
        cur_start = None
        cur_end_excl = None
        cur_label = "æ±‡æ€»ï¼ˆå…¨éƒ¨ï¼‰"
        prev_start = None
        prev_end_excl = None
        prev_label = ""
        yoy_start = None
        yoy_end_excl = None
        yoy_label = ""
    elif mode == "yesterday":
        if anchor:
            cur_start = anchor
            cur_end_excl = anchor + timedelta(days=1)
            cur_label = f"æ—¥ ({_date_str(cur_start)})"
        else:
            cur_start = today - timedelta(days=1)
            cur_end_excl = today
            cur_label = f"æ˜¨å¤© ({_date_str(cur_start)})"

        prev_start = cur_start - timedelta(days=1)
        prev_end_excl = cur_start
        prev_label = f"å‰ä¸€æ—¥ ({_date_str(prev_start)})"

        if yoy:
            yoy_start = _tag_report_shift_year_safe(cur_start, -1)
            yoy_end_excl = yoy_start + timedelta(days=1)
            yoy_label = f"å»å¹´åŒæ—¥ ({_date_str(yoy_start)})"
        else:
            yoy_start = None
            yoy_end_excl = None
            yoy_label = ""
    elif mode == "last_week":
        if anchor:
            week_start, week_end = _tag_report_week_bounds(anchor)
            cur_label = f"å‘¨ ({_date_str(week_start)} ~ {_date_str(week_end)})"
        else:
            days_since_monday = today.weekday()
            this_monday = today - timedelta(days=days_since_monday)
            week_start = this_monday - timedelta(days=7)
            week_end = this_monday - timedelta(days=1)
            cur_label = f"ä¸Šå‘¨ ({_date_str(week_start)} ~ {_date_str(week_end)})"

        cur_start = week_start
        cur_end_excl = week_end + timedelta(days=1)

        prev_start = week_start - timedelta(days=7)
        prev_end_excl = week_start
        prev_label = f"{_date_str(prev_start)} ~ {_date_str(prev_end_excl - timedelta(days=1))}"

        if yoy:
            yoy_start = _tag_report_shift_year_safe(week_start, -1)
            yoy_end = _tag_report_shift_year_safe(week_end, -1)
            yoy_end_excl = yoy_end + timedelta(days=1)
            yoy_label = f"{_date_str(yoy_start)} ~ {_date_str(yoy_end)}"
        else:
            yoy_start = None
            yoy_end_excl = None
            yoy_label = ""
    elif mode == "last_month":
        if anchor:
            month_start, month_end = _tag_report_month_bounds(anchor)
            cur_label = f"æœˆ ({_date_str(month_start)} ~ {_date_str(month_end)})"
        else:
            this_month_first = datetime(today.year, today.month, 1)
            last_month_last = this_month_first - timedelta(days=1)
            month_start, month_end = _tag_report_month_bounds(last_month_last)
            cur_label = f"ä¸Šæœˆ ({_date_str(month_start)} ~ {_date_str(month_end)})"

        cur_start = month_start
        cur_end_excl = month_end + timedelta(days=1)

        prev_anchor = month_start - timedelta(days=1)
        prev_start, prev_end = _tag_report_month_bounds(prev_anchor)
        prev_end_excl = prev_end + timedelta(days=1)
        prev_label = f"{_date_str(prev_start)} ~ {_date_str(prev_end)}"

        if yoy:
            yoy_anchor = datetime(month_start.year - 1, month_start.month, 1)
            yoy_start, yoy_end = _tag_report_month_bounds(yoy_anchor)
            yoy_end_excl = yoy_end + timedelta(days=1)
            yoy_label = f"{_date_str(yoy_start)} ~ {_date_str(yoy_end)}"
        else:
            yoy_start = None
            yoy_end_excl = None
            yoy_label = ""
    elif mode == "last_quarter":
        if anchor:
            quarter_start, quarter_end = _tag_report_quarter_bounds(anchor)
            cur_label = f"å­£åº¦ ({_date_str(quarter_start)} ~ {_date_str(quarter_end)})"
        else:
            current_quarter = (today.month - 1) // 3 + 1
            if current_quarter == 1:
                quarter_anchor = datetime(today.year - 1, 12, 31)
            else:
                quarter_anchor = datetime(today.year, (current_quarter - 1) * 3, 1) - timedelta(days=1)
            quarter_start, quarter_end = _tag_report_quarter_bounds(quarter_anchor)
            cur_label = f"ä¸Šå­£åº¦ ({_date_str(quarter_start)} ~ {_date_str(quarter_end)})"

        cur_start = quarter_start
        cur_end_excl = quarter_end + timedelta(days=1)

        prev_anchor = quarter_start - timedelta(days=1)
        prev_start, prev_end = _tag_report_quarter_bounds(prev_anchor)
        prev_end_excl = prev_end + timedelta(days=1)
        prev_label = f"{_date_str(prev_start)} ~ {_date_str(prev_end)}"

        if yoy:
            yoy_anchor = datetime(quarter_start.year - 1, quarter_start.month, 1)
            yoy_start, yoy_end = _tag_report_quarter_bounds(yoy_anchor)
            yoy_end_excl = yoy_end + timedelta(days=1)
            yoy_label = f"{_date_str(yoy_start)} ~ {_date_str(yoy_end)}"
        else:
            yoy_start = None
            yoy_end_excl = None
            yoy_label = ""
    elif mode == "last_year":
        if anchor:
            year_start, year_end = _tag_report_year_bounds(anchor)
            cur_label = f"å¹´ ({year_start.year})"
        else:
            year_start = datetime(today.year - 1, 1, 1)
            year_end = datetime(today.year - 1, 12, 31)
            cur_label = f"å»å¹´ ({year_start.year})"

        cur_start = year_start
        cur_end_excl = year_end + timedelta(days=1)

        prev_start = datetime(year_start.year - 1, 1, 1)
        prev_end = datetime(year_start.year - 1, 12, 31)
        prev_end_excl = prev_end + timedelta(days=1)
        prev_label = f"{prev_start.year}"

        if yoy:
            yoy_start = prev_start
            yoy_end_excl = prev_end_excl
            yoy_label = f"{prev_start.year}"
        else:
            yoy_start = None
            yoy_end_excl = None
            yoy_label = ""
    else:
        if not start_dt and not end_dt:
            end_dt = today - timedelta(days=1)
            start_dt = end_dt - timedelta(days=6)
        start_dt, end_dt = _tag_report_normalize_range(start_dt, end_dt)

        cur_start = start_dt
        cur_end_excl = (end_dt + timedelta(days=1)) if end_dt else None
        cur_label = f"{_date_str(start_dt)} ~ {_date_str(end_dt)}" if (start_dt and end_dt) else "è‡ªå®šä¹‰"

        if start_dt and end_dt:
            days = (end_dt - start_dt).days + 1
            prev_end = start_dt - timedelta(days=1)
            prev_start = prev_end - timedelta(days=days - 1)
            prev_end_excl = prev_end + timedelta(days=1)
            prev_label = f"{_date_str(prev_start)} ~ {_date_str(prev_end)}"
        else:
            prev_start = None
            prev_end_excl = None
            prev_label = ""

        if yoy and start_dt and end_dt:
            yoy_start = _tag_report_shift_year_safe(start_dt, -1)
            yoy_end = _tag_report_shift_year_safe(end_dt, -1)
            yoy_end_excl = yoy_end + timedelta(days=1)
            yoy_label = f"{_date_str(yoy_start)} ~ {_date_str(yoy_end)}"
        else:
            yoy_start = None
            yoy_end_excl = None
            yoy_label = ""

    display_start = _date_str(cur_start)
    display_end = _date_str(cur_end_excl - timedelta(days=1)) if cur_end_excl else ""

    return {
        "mode": mode,
        "cur_start": cur_start,
        "cur_end_excl": cur_end_excl,
        "prev_start": prev_start,
        "prev_end_excl": prev_end_excl,
        "yoy_start": yoy_start,
        "yoy_end_excl": yoy_end_excl,
        "cur_label": cur_label,
        "prev_label": prev_label,
        "yoy_label": yoy_label,
        "display_start": display_start,
        "display_end": display_end,
    }


@app.get("/reports/tags", response_class=HTMLResponse)
def tags_report_page(
    request: Request,
    user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
    time_mode: str | None = None,
    start: str | None = None,
    end: str | None = None,
    platform: str | None = None,
    agent_user_id: str | None = None,
    yoy: int | None = None,
):
    """æ ‡ç­¾æŠ¥è¡¨ï¼ˆæ•°æ®é€è§†è¡¨ï¼‰ã€‚"""

    period = _build_tag_report_period(time_mode, start, end, bool(yoy))

    platform_f = (platform or "").strip()
    agent_user_id_int = None
    try:
        agent_user_id_int = int(agent_user_id) if agent_user_id else None
    except Exception:
        agent_user_id_int = None

    if user.role == Role.agent:
        agent_user_id_int = user.id

    raw_platforms = session.exec(select(Conversation.platform)).all()
    platform_options = sorted(
        {
            (p[0] if isinstance(p, (list, tuple)) else p)
            for p in raw_platforms
            if (p[0] if isinstance(p, (list, tuple)) else p)
        }
    )
    users = session.exec(select(User).order_by(User.role.asc(), User.name.asc())).all()
    agent_options = [
        {"id": u.id, "label": f"{u.name}ï¼ˆ{u.role}ï¼‰"}
        for u in users
        if u.role in (Role.agent, Role.supervisor) and getattr(u, "is_active", True)
    ]

    cats = session.exec(
        select(TagCategory)
        .where(TagCategory.is_active == True)
        .order_by(TagCategory.sort_order.asc(), TagCategory.id.asc())
    ).all()
    tags = session.exec(
        select(TagDefinition)
        .where(TagDefinition.is_active == True)
        .order_by(TagDefinition.category_id.asc(), TagDefinition.sort_order.asc(), TagDefinition.id.asc())
    ).all()
    cat_by_id = {int(c.id): c for c in cats if c.id}

    tag_meta = [
        {
            "id": int(t.id),
            "name": t.name,
            "category_id": int(t.category_id or 0),
            "category_name": (cat_by_id.get(int(t.category_id or 0)).name if cat_by_id.get(int(t.category_id or 0)) else "æœªåˆ†ç±»"),
            "standard": (t.standard or "") or (t.description or "") or "æœªé…ç½®",
            "sort_order": int(t.sort_order or 0),
        }
        for t in tags
        if t.id
    ]

    category_meta = [
        {
            "id": int(c.id),
            "name": c.name,
            "sort_order": int(c.sort_order or 0),
        }
        for c in cats
        if c.id
    ]

    user_meta = [
        {
            "id": int(u.id),
            "name": u.name,
            "label": f"{u.name}ï¼ˆ{u.role}ï¼‰",
        }
        for u in users
        if u.id
    ]

    time_modes = [
        ("all", "æ±‡æ€»"),
        ("last_year", "å¹´"),
        ("last_quarter", "å­£åº¦"),
        ("last_month", "æœˆ"),
        ("last_week", "å‘¨"),
        ("yesterday", "æ—¥"),
        ("custom", "è‡ªå®šä¹‰"),
    ]

    return _template(
        request,
        "tags_report.html",
        user=user,
        wide_layout=True,
        flash="",
        time_modes=time_modes,
        filters={
            "time_mode": period["mode"],
            "start": period["display_start"] or (start or ""),
            "end": period["display_end"] or (end or ""),
            "platform": platform_f,
            "agent_user_id": agent_user_id_int or "",
            "yoy": bool(yoy),
        },
        period={
            "current_label": period["cur_label"],
            "prev_label": period["prev_label"],
            "yoy_label": period["yoy_label"],
        },
        platform_options=platform_options,
        agent_options=agent_options,
        tag_meta=tag_meta,
        category_meta=category_meta,
        user_meta=user_meta,
    )


@app.get("/api/reports/tags/pivot-data")
def tags_report_pivot_data(
    request: Request,
    user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
    time_mode: str | None = None,
    start: str | None = None,
    end: str | None = None,
    platform: str | None = None,
    agent_user_id: str | None = None,
    reception_scenario: str | None = None,
    satisfaction_change: str | None = None,
    yoy: int | None = None,
):
    """æ ‡ç­¾æŠ¥è¡¨æ•°æ®é€è§†è¡¨æ•°æ®æºã€‚"""
    from sqlalchemy import func

    period = _build_tag_report_period(time_mode, start, end, bool(yoy))

    platform_f = (platform or "").strip()
    scenario_f = (reception_scenario or "").strip()
    satisfaction_f = (satisfaction_change or "").strip()
    agent_user_id_int = None
    try:
        agent_user_id_int = int(agent_user_id) if agent_user_id else None
    except Exception:
        agent_user_id_int = None

    if user.role == Role.agent:
        agent_user_id_int = user.id

    latest_subq = (
        select(func.max(ConversationAnalysis.id).label("analysis_id"))
        .group_by(ConversationAnalysis.conversation_id)
    ).subquery()

    binding_user_id = func.coalesce(AgentBinding.user_id, Conversation.agent_user_id)
    conv_date = func.coalesce(Conversation.ended_at, Conversation.started_at, Conversation.uploaded_at)

    def _apply_filters(stmt, start_dt, end_excl):
        if start_dt:
            stmt = stmt.where(conv_date >= start_dt)
        if end_excl:
            stmt = stmt.where(conv_date < end_excl)
        if platform_f:
            stmt = stmt.where(Conversation.platform == platform_f)
        if agent_user_id_int:
            stmt = stmt.where(binding_user_id == agent_user_id_int)
        if scenario_f:
            stmt = stmt.where(ConversationAnalysis.reception_scenario == scenario_f)
        if satisfaction_f:
            stmt = stmt.where(ConversationAnalysis.satisfaction_change == satisfaction_f)
        return stmt

    tag_base_stmt = (
        select(
            TagDefinition.id.label("tag_id"),
            TagDefinition.category_id.label("category_id"),
            Conversation.platform.label("platform"),
            binding_user_id.label("agent_user_id"),
            ConversationAnalysis.reception_scenario.label("scenario"),
            ConversationAnalysis.satisfaction_change.label("satisfaction"),
            func.count(func.distinct(Conversation.id)).label("cnt"),
        )
        .select_from(ConversationTagHit)
        .join(ConversationAnalysis, ConversationAnalysis.id == ConversationTagHit.analysis_id)
        .join(Conversation, Conversation.id == ConversationAnalysis.conversation_id)
        .join(latest_subq, latest_subq.c.analysis_id == ConversationAnalysis.id)
        .join(TagDefinition, TagDefinition.id == ConversationTagHit.tag_id)
        .outerjoin(AgentBinding, and_(AgentBinding.platform == Conversation.platform, AgentBinding.agent_account == Conversation.agent_account))
        .where(TagDefinition.is_active == True)
    )

    tag_group_cols = (
        TagDefinition.id,
        TagDefinition.category_id,
        Conversation.platform,
        binding_user_id,
        ConversationAnalysis.reception_scenario,
        ConversationAnalysis.satisfaction_change,
    )

    def _tag_rows(start_dt, end_excl):
        stmt = _apply_filters(tag_base_stmt, start_dt, end_excl).group_by(*tag_group_cols)
        return session.exec(stmt).all()

    cur_rows = _tag_rows(period["cur_start"], period["cur_end_excl"])
    prev_rows = _tag_rows(period["prev_start"], period["prev_end_excl"]) if period["prev_start"] else []
    yoy_rows = _tag_rows(period["yoy_start"], period["yoy_end_excl"]) if period["yoy_start"] else []

    def _tag_index(rows):
        idx: dict[tuple, int] = {}
        for tag_id, cat_id, plat, uid, scenario, satisfaction, cnt in rows:
            key = (
                int(tag_id or 0),
                int(cat_id or 0),
                str(plat or ""),
                int(uid or 0),
                str(scenario or ""),
                str(satisfaction or ""),
            )
            idx[key] = int(cnt or 0)
        return idx

    cur_idx = _tag_index(cur_rows)
    prev_idx = _tag_index(prev_rows)
    yoy_idx = _tag_index(yoy_rows)

    all_tag_keys = set(cur_idx.keys()) | set(prev_idx.keys()) | set(yoy_idx.keys())
    tag_hits = [
        {
            "tag_id": k[0],
            "category_id": k[1],
            "platform": k[2],
            "agent_user_id": k[3],
            "reception_scenario": k[4],
            "satisfaction_change": k[5],
            "current": cur_idx.get(k, 0),
            "prev": prev_idx.get(k, 0),
            "yoy": yoy_idx.get(k, 0),
        }
        for k in all_tag_keys
    ]

    total_base_stmt = (
        select(
            Conversation.platform.label("platform"),
            binding_user_id.label("agent_user_id"),
            ConversationAnalysis.reception_scenario.label("scenario"),
            ConversationAnalysis.satisfaction_change.label("satisfaction"),
            func.count(func.distinct(Conversation.id)).label("cnt"),
        )
        .select_from(Conversation)
        .join(ConversationAnalysis, ConversationAnalysis.conversation_id == Conversation.id)
        .join(latest_subq, latest_subq.c.analysis_id == ConversationAnalysis.id)
        .outerjoin(AgentBinding, and_(AgentBinding.platform == Conversation.platform, AgentBinding.agent_account == Conversation.agent_account))
    )

    total_group_cols = (
        Conversation.platform,
        binding_user_id,
        ConversationAnalysis.reception_scenario,
        ConversationAnalysis.satisfaction_change,
    )

    def _total_rows(start_dt, end_excl):
        stmt = _apply_filters(total_base_stmt, start_dt, end_excl).group_by(*total_group_cols)
        return session.exec(stmt).all()

    total_cur_rows = _total_rows(period["cur_start"], period["cur_end_excl"])
    total_prev_rows = _total_rows(period["prev_start"], period["prev_end_excl"]) if period["prev_start"] else []
    total_yoy_rows = _total_rows(period["yoy_start"], period["yoy_end_excl"]) if period["yoy_start"] else []

    def _total_index(rows):
        idx: dict[tuple, int] = {}
        for plat, uid, scenario, satisfaction, cnt in rows:
            key = (str(plat or ""), int(uid or 0), str(scenario or ""), str(satisfaction or ""))
            idx[key] = int(cnt or 0)
        return idx

    total_cur_idx = _total_index(total_cur_rows)
    total_prev_idx = _total_index(total_prev_rows)
    total_yoy_idx = _total_index(total_yoy_rows)

    all_total_keys = set(total_cur_idx.keys()) | set(total_prev_idx.keys()) | set(total_yoy_idx.keys())
    totals = [
        {
            "platform": k[0],
            "agent_user_id": k[1],
            "reception_scenario": k[2],
            "satisfaction_change": k[3],
            "current": total_cur_idx.get(k, 0),
            "prev": total_prev_idx.get(k, 0),
            "yoy": total_yoy_idx.get(k, 0),
        }
        for k in all_total_keys
    ]

    # å‘½ä¸­ä»»æ„ã€Œæœ‰æ•ˆæ ‡ç­¾ã€çš„å»é‡å¯¹è¯æ•°ï¼ˆç”¨äºâ€œè´¨æ£€å¯¹è¯æ•°â€ä¸â€œå¯¹è¯æ•°å æ¯”â€ï¼‰
    tagged_total_base_stmt = (
        select(
            Conversation.platform.label("platform"),
            binding_user_id.label("agent_user_id"),
            ConversationAnalysis.reception_scenario.label("scenario"),
            ConversationAnalysis.satisfaction_change.label("satisfaction"),
            func.count(func.distinct(Conversation.id)).label("cnt"),
        )
        .select_from(ConversationTagHit)
        .join(ConversationAnalysis, ConversationAnalysis.id == ConversationTagHit.analysis_id)
        .join(Conversation, Conversation.id == ConversationAnalysis.conversation_id)
        .join(latest_subq, latest_subq.c.analysis_id == ConversationAnalysis.id)
        .join(TagDefinition, TagDefinition.id == ConversationTagHit.tag_id)
        .outerjoin(AgentBinding, and_(AgentBinding.platform == Conversation.platform, AgentBinding.agent_account == Conversation.agent_account))
        .where(TagDefinition.is_active == True)
    )

    tagged_total_group_cols = (
        Conversation.platform,
        binding_user_id,
        ConversationAnalysis.reception_scenario,
        ConversationAnalysis.satisfaction_change,
    )

    def _tagged_total_rows(start_dt, end_excl):
        stmt = _apply_filters(tagged_total_base_stmt, start_dt, end_excl).group_by(*tagged_total_group_cols)
        return session.exec(stmt).all()

    tagged_total_cur_rows = _tagged_total_rows(period["cur_start"], period["cur_end_excl"])
    tagged_total_prev_rows = _tagged_total_rows(period["prev_start"], period["prev_end_excl"]) if period["prev_start"] else []
    tagged_total_yoy_rows = _tagged_total_rows(period["yoy_start"], period["yoy_end_excl"]) if period["yoy_start"] else []

    def _tagged_total_index(rows):
        idx: dict[tuple, int] = {}
        for plat, uid, scenario, satisfaction, cnt in rows:
            key = (str(plat or ""), int(uid or 0), str(scenario or ""), str(satisfaction or ""))
            idx[key] = int(cnt or 0)
        return idx

    tagged_total_cur_idx = _tagged_total_index(tagged_total_cur_rows)
    tagged_total_prev_idx = _tagged_total_index(tagged_total_prev_rows)
    tagged_total_yoy_idx = _tagged_total_index(tagged_total_yoy_rows)

    all_tagged_total_keys = set(tagged_total_cur_idx.keys()) | set(tagged_total_prev_idx.keys()) | set(tagged_total_yoy_idx.keys())
    tagged_totals = [
        {
            "platform": k[0],
            "agent_user_id": k[1],
            "reception_scenario": k[2],
            "satisfaction_change": k[3],
            "current": tagged_total_cur_idx.get(k, 0),
            "prev": tagged_total_prev_idx.get(k, 0),
            "yoy": tagged_total_yoy_idx.get(k, 0),
        }
        for k in all_tagged_total_keys
    ]

    tagged_category_base_stmt = (
        select(
            TagDefinition.category_id.label("category_id"),
            Conversation.platform.label("platform"),
            binding_user_id.label("agent_user_id"),
            ConversationAnalysis.reception_scenario.label("scenario"),
            ConversationAnalysis.satisfaction_change.label("satisfaction"),
            func.count(func.distinct(Conversation.id)).label("cnt"),
        )
        .select_from(ConversationTagHit)
        .join(ConversationAnalysis, ConversationAnalysis.id == ConversationTagHit.analysis_id)
        .join(Conversation, Conversation.id == ConversationAnalysis.conversation_id)
        .join(latest_subq, latest_subq.c.analysis_id == ConversationAnalysis.id)
        .join(TagDefinition, TagDefinition.id == ConversationTagHit.tag_id)
        .outerjoin(AgentBinding, and_(AgentBinding.platform == Conversation.platform, AgentBinding.agent_account == Conversation.agent_account))
        .where(TagDefinition.is_active == True)
    )

    tagged_category_group_cols = (
        TagDefinition.category_id,
        Conversation.platform,
        binding_user_id,
        ConversationAnalysis.reception_scenario,
        ConversationAnalysis.satisfaction_change,
    )

    def _tagged_category_rows(start_dt, end_excl):
        stmt = _apply_filters(tagged_category_base_stmt, start_dt, end_excl).group_by(*tagged_category_group_cols)
        return session.exec(stmt).all()

    tagged_category_cur_rows = _tagged_category_rows(period["cur_start"], period["cur_end_excl"])
    tagged_category_prev_rows = _tagged_category_rows(period["prev_start"], period["prev_end_excl"]) if period["prev_start"] else []
    tagged_category_yoy_rows = _tagged_category_rows(period["yoy_start"], period["yoy_end_excl"]) if period["yoy_start"] else []

    def _tagged_category_index(rows):
        idx: dict[tuple, int] = {}
        for cat_id, plat, uid, scenario, satisfaction, cnt in rows:
            key = (int(cat_id or 0), str(plat or ""), int(uid or 0), str(scenario or ""), str(satisfaction or ""))
            idx[key] = int(cnt or 0)
        return idx

    tagged_category_cur_idx = _tagged_category_index(tagged_category_cur_rows)
    tagged_category_prev_idx = _tagged_category_index(tagged_category_prev_rows)
    tagged_category_yoy_idx = _tagged_category_index(tagged_category_yoy_rows)

    all_tagged_category_keys = (
        set(tagged_category_cur_idx.keys()) | set(tagged_category_prev_idx.keys()) | set(tagged_category_yoy_idx.keys())
    )
    tagged_category_totals = [
        {
            "category_id": k[0],
            "platform": k[1],
            "agent_user_id": k[2],
            "reception_scenario": k[3],
            "satisfaction_change": k[4],
            "current": tagged_category_cur_idx.get(k, 0),
            "prev": tagged_category_prev_idx.get(k, 0),
            "yoy": tagged_category_yoy_idx.get(k, 0),
        }
        for k in all_tagged_category_keys
    ]

    def _date_str(d: datetime | None) -> str:
        return d.date().isoformat() if d else ""

    cur_end = period["cur_end_excl"] - timedelta(days=1) if period["cur_end_excl"] else None
    prev_end = period["prev_end_excl"] - timedelta(days=1) if period["prev_end_excl"] else None
    yoy_end = period["yoy_end_excl"] - timedelta(days=1) if period["yoy_end_excl"] else None

    return {
        "mode": period["mode"],
        "period": {
            "current_label": period["cur_label"],
            "prev_label": period["prev_label"],
            "yoy_label": period["yoy_label"],
        },
        "range": {
            "current_start": _date_str(period["cur_start"]),
            "current_end": _date_str(cur_end),
            "prev_start": _date_str(period["prev_start"]),
            "prev_end": _date_str(prev_end),
            "yoy_start": _date_str(period["yoy_start"]),
            "yoy_end": _date_str(yoy_end),
        },
        "tag_hits": tag_hits,
        "totals": totals,
        "tagged_totals": tagged_totals,
        "tagged_category_totals": tagged_category_totals,
    }


@app.get("/api/reports/tags/conversations")
def get_tag_conversations(
    request: Request,
    user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
    tag_id: int | None = None,
    category_id: int | None = None,
    start: str | None = None,
    end: str | None = None,
    platform: str | None = None,
    agent_user_id: str | None = None,
    reception_scenario: str | None = None,
    satisfaction_change: str | None = None,
    page: int | None = 1,
    per_page: int | None = 20,
):
    """æ ‡ç­¾æŠ¥è¡¨ï¼šæ‰“å¼€â€œå¯¹åº”å¯¹è¯åˆ—è¡¨â€ï¼ˆåªçœ‹æ¯ä¸ªå¯¹è¯çš„æœ€æ–°è´¨æ£€ï¼‰ã€‚

    - tag_id: åªçœ‹å‘½ä¸­è¯¥æ ‡ç­¾çš„å¯¹è¯
    - category_id: åªçœ‹å‘½ä¸­è¯¥æ ‡ç­¾åˆ†ç±»ä¸‹ä»»ä¸€æ ‡ç­¾çš„å¯¹è¯
    - ä¸¤è€…éƒ½ä¸ä¼ ï¼šåªçœ‹å½“å‰ç­›é€‰æ¡ä»¶ä¸‹çš„â€œå·²è´¨æ£€å¯¹è¯â€ï¼ˆæœ€æ–°è´¨æ£€ï¼‰
    """
    from sqlalchemy import func

    def _parse_date(d: str | None):
        if not d:
            return None
        d = d.strip()
        if not d:
            return None
        try:
            y, m, day = [int(x) for x in d.split("-")]
            return datetime(y, m, day)
        except Exception:
            return None

    start_dt = _parse_date(start)
    end_dt = _parse_date(end)
    if end_dt:
        end_dt = end_dt + timedelta(days=1)  # Make it exclusive

    platform_f = (platform or "").strip()
    scenario_f = (reception_scenario or "").strip()
    satisfaction_f = (satisfaction_change or "").strip()
    tag_id_i = int(tag_id or 0) or None
    category_id_i = int(category_id or 0) or None
    agent_user_id_int = None
    try:
        agent_user_id_int = int(agent_user_id) if agent_user_id else None
    except Exception:
        agent_user_id_int = None

    # Agent role: only see self
    if user.role == Role.agent:
        agent_user_id_int = user.id

    page_i = 1
    per_page_i = 20
    try:
        page_i = int(page or 1)
    except Exception:
        page_i = 1
    try:
        per_page_i = int(per_page or 20)
    except Exception:
        per_page_i = 20
    if page_i < 1:
        page_i = 1
    if per_page_i < 1:
        per_page_i = 20
    if per_page_i > 50:
        per_page_i = 50

    # Get latest analysis per conversation
    latest_subq = (
        select(
            ConversationAnalysis.conversation_id.label("cid"),
            func.max(ConversationAnalysis.id).label("analysis_id"),
        )
        .group_by(ConversationAnalysis.conversation_id)
    ).subquery()

    # Real-time mapping via AgentBinding if exists
    binding_user_id = func.coalesce(AgentBinding.user_id, Conversation.agent_user_id)
    event_ts_expr = func.coalesce(Conversation.ended_at, Conversation.started_at, Conversation.uploaded_at)
    event_ts = event_ts_expr.label("event_ts")

    msg_count_subq = (
        select(Message.conversation_id.label("cid"), func.count(Message.id).label("message_count"))
        .group_by(Message.conversation_id)
    ).subquery()

    where_terms = []
    if start_dt:
        where_terms.append(event_ts_expr >= start_dt)
    if end_dt:
        where_terms.append(event_ts_expr < end_dt)
    if platform_f:
        where_terms.append(Conversation.platform == platform_f)
    if agent_user_id_int:
        where_terms.append(binding_user_id == agent_user_id_int)
    if scenario_f:
        where_terms.append(ConversationAnalysis.reception_scenario == scenario_f)
    if satisfaction_f:
        where_terms.append(ConversationAnalysis.satisfaction_change == satisfaction_f)

    if tag_id_i:
        where_terms.append(
            exists(
                select(1)
                .select_from(ConversationTagHit)
                .where(
                    ConversationTagHit.analysis_id == latest_subq.c.analysis_id,
                    ConversationTagHit.tag_id == tag_id_i,
                )
            )
        )
    elif category_id_i:
        where_terms.append(
            exists(
                select(1)
                .select_from(ConversationTagHit)
                .join(TagDefinition, TagDefinition.id == ConversationTagHit.tag_id)
                .where(
                    ConversationTagHit.analysis_id == latest_subq.c.analysis_id,
                    TagDefinition.category_id == category_id_i,
                )
            )
        )

    base_stmt = (
        select(Conversation.id.label("cid"))
        .select_from(Conversation)
        .join(latest_subq, latest_subq.c.cid == Conversation.id)
        .join(ConversationAnalysis, ConversationAnalysis.id == latest_subq.c.analysis_id)
        .outerjoin(AgentBinding, and_(AgentBinding.platform == Conversation.platform, AgentBinding.agent_account == Conversation.agent_account))
    )
    if where_terms:
        base_stmt = base_stmt.where(*where_terms)

    total = session.exec(select(func.count()).select_from(base_stmt.subquery())).one()

    stmt = (
        select(
            Conversation.id,
            Conversation.platform,
            Conversation.buyer_id,
            Conversation.agent_account,
            User.name.label("agent_name"),
            event_ts,
            Conversation.started_at,
            Conversation.ended_at,
            Conversation.uploaded_at,
            msg_count_subq.c.message_count,
            ConversationAnalysis.reception_scenario,
            ConversationAnalysis.satisfaction_change,
        )
        .select_from(Conversation)
        .join(latest_subq, latest_subq.c.cid == Conversation.id)
        .join(ConversationAnalysis, ConversationAnalysis.id == latest_subq.c.analysis_id)
        .outerjoin(AgentBinding, and_(AgentBinding.platform == Conversation.platform, AgentBinding.agent_account == Conversation.agent_account))
        .outerjoin(User, User.id == binding_user_id)
        .outerjoin(msg_count_subq, msg_count_subq.c.cid == Conversation.id)
    )

    if where_terms:
        stmt = stmt.where(*where_terms)

    stmt = stmt.order_by(event_ts.desc(), Conversation.id.desc())
    stmt = stmt.offset((page_i - 1) * per_page_i).limit(per_page_i)

    def _dt_str(dt: datetime | None) -> str:
        if not dt:
            return ""
        try:
            return dt.strftime("%Y-%m-%d %H:%M")
        except Exception:
            return str(dt)

    items = []
    for (
        cid,
        plat,
        buyer_id_v,
        agent_account_v,
        agent_name_v,
        event_ts_v,
        started_at_v,
        ended_at_v,
        uploaded_at_v,
        message_count_v,
        scenario_v,
        satisfaction_v,
    ) in session.exec(stmt).all():
        items.append(
            {
                "cid": cid,
                "platform": plat or "",
                "buyer_id": buyer_id_v or "",
                "agent_account": agent_account_v or "",
                "agent_name": agent_name_v or "",
                "event_time": _dt_str(event_ts_v),
                "started_at": _dt_str(started_at_v),
                "ended_at": _dt_str(ended_at_v),
                "uploaded_at": _dt_str(uploaded_at_v),
                "message_count": int(message_count_v or 0),
                "reception_scenario": scenario_v or "",
                "satisfaction_change": satisfaction_v or "",
            }
        )

    return {
        "total": int(total or 0),
        "page": page_i,
        "per_page": per_page_i,
        "items": items,
        "cids": [x["cid"] for x in items],
    }


# =========================
# Agent Analysis Report: å®¢æœæ—¥æœŸã€åœºæ™¯ã€æ»¡æ„åº¦äº¤å‰åˆ†æ
# =========================


@app.get("/reports/agent-analysis", response_class=HTMLResponse)
def agent_analysis_report_page(
    request: Request,
    user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
    group_by: str | None = None,  # day/week/month
    start: str | None = None,
    end: str | None = None,
    platform: str | None = None,
    agent_user_id: str | None = None,
):
    """å®¢æœæ—¥æœŸã€åœºæ™¯ã€æ»¡æ„åº¦äº¤å‰åˆ†ææŠ¥è¡¨ã€‚
    
    å¤šç»´åº¦å±•ç¤ºï¼šå®¢æœ -> æ—¥æœŸ -> åœºæ™¯ -> æ»¡æ„åº¦ -> å¯¹è¯æ•°
    """
    from sqlalchemy import func, case
    from datetime import datetime, timedelta
    
    def _parse_date(d: str | None):
        if not d:
            return None
        d = d.strip()
        if not d:
            return None
        try:
            y, m, day = [int(x) for x in d.split("-")]
            return datetime(y, m, day)
        except Exception:
            return None
    
    # å‚æ•°è§£æ
    group_by_mode = (group_by or "day").strip() or "day"
    start_dt = _parse_date(start)
    end_dt = _parse_date(end)
    now = datetime.utcnow()
    today = datetime(now.year, now.month, now.day)
    
    # é»˜è®¤æœ€è¿‘7å¤©
    if not start_dt and not end_dt:
        end_dt = today - timedelta(days=1)
        start_dt = end_dt - timedelta(days=6)
    if start_dt and not end_dt:
        end_dt = start_dt
    if end_dt and not start_dt:
        start_dt = end_dt
    if start_dt and end_dt and start_dt > end_dt:
        start_dt, end_dt = end_dt, start_dt
    
    # è½¬æ¢ä¸ºåŒ…å«ç»“æŸæ—¥æœŸçš„èŒƒå›´
    end_excl = (end_dt + timedelta(days=1)) if end_dt else None
    
    platform_f = (platform or "").strip()
    agent_user_id_int = None
    try:
        agent_user_id_int = int(agent_user_id) if agent_user_id else None
    except Exception:
        agent_user_id_int = None
    
    # Agent role: only see self
    if user.role == Role.agent:
        agent_user_id_int = user.id
    
    # è·å–ç­›é€‰é€‰é¡¹
    raw_platforms = session.exec(select(Conversation.platform)).all()
    platform_options = sorted(
        {
            (p[0] if isinstance(p, (list, tuple)) else p)
            for p in raw_platforms
            if (p[0] if isinstance(p, (list, tuple)) else p)
        }
    )
    users = session.exec(select(User).order_by(User.role.asc(), User.name.asc())).all()
    agent_options = [
        {"id": u.id, "label": f"{u.name}ï¼ˆ{u.role}ï¼‰"}
        for u in users
        if u.role in (Role.agent, Role.supervisor) and getattr(u, "is_active", True)
    ]
    
    # è·å–æœ€æ–°åˆ†æ
    latest_subq = (
        select(func.max(ConversationAnalysis.id).label("analysis_id"))
        .group_by(ConversationAnalysis.conversation_id)
    ).subquery()
    
    # å®æ—¶ç»‘å®šå®¢æœ
    binding_user_id = func.coalesce(AgentBinding.user_id, Conversation.agent_user_id)
    
    # æ—¥æœŸåˆ†ç»„è¡¨è¾¾å¼
    conv_date = func.coalesce(Conversation.ended_at, Conversation.started_at, Conversation.uploaded_at)
    if group_by_mode == "week":
        # PostgreSQL: date_trunc('week', timestamp)
        date_group = func.date_trunc('week', conv_date)
    elif group_by_mode == "month":
        date_group = func.date_trunc('month', conv_date)
    else:  # day
        date_group = func.date_trunc('day', conv_date)
    
    # æ„å»ºæŸ¥è¯¢ï¼šå®¢æœã€æ—¥æœŸã€åœºæ™¯ã€æ»¡æ„åº¦ -> å¯¹è¯æ•°
    stmt = (
        select(
            binding_user_id.label("agent_user_id"),
            date_group.label("period"),
            ConversationAnalysis.reception_scenario.label("scenario"),
            ConversationAnalysis.satisfaction_change.label("satisfaction"),
            func.count(func.distinct(Conversation.id)).label("cnt"),
        )
        .select_from(Conversation)
        .join(ConversationAnalysis, ConversationAnalysis.conversation_id == Conversation.id)
        .join(latest_subq, latest_subq.c.analysis_id == ConversationAnalysis.id)
        .outerjoin(AgentBinding, and_(
            AgentBinding.platform == Conversation.platform,
            AgentBinding.agent_account == Conversation.agent_account
        ))
    )
    
    # åº”ç”¨ç­›é€‰
    if start_dt:
        stmt = stmt.where(conv_date >= start_dt)
    if end_excl:
        stmt = stmt.where(conv_date < end_excl)
    if platform_f:
        stmt = stmt.where(Conversation.platform == platform_f)
    if agent_user_id_int:
        stmt = stmt.where(binding_user_id == agent_user_id_int)
    
    # åˆ†ç»„ç»Ÿè®¡
    stmt = stmt.group_by(binding_user_id, date_group, ConversationAnalysis.reception_scenario, ConversationAnalysis.satisfaction_change)
    stmt = stmt.order_by(binding_user_id.asc(), date_group.asc(), ConversationAnalysis.reception_scenario.asc(), ConversationAnalysis.satisfaction_change.asc())
    
    rows = session.exec(stmt).all()
    
    # æ„å»ºç”¨æˆ·æ˜ å°„
    user_by_id = {u.id: u for u in users if u.id}
    
    # æ„å»ºå¤šçº§æ•°æ®ç»“æ„
    data_tree = {}
    for agent_id, period, scenario, satisfaction, cnt in rows:
        agent_id = int(agent_id or 0)
        if agent_id not in data_tree:
            data_tree[agent_id] = {}
        
        period_key = period.date().isoformat() if period else "æœªçŸ¥"
        if period_key not in data_tree[agent_id]:
            data_tree[agent_id][period_key] = {}
        
        scenario = str(scenario or "æœªçŸ¥")
        if scenario not in data_tree[agent_id][period_key]:
            data_tree[agent_id][period_key][scenario] = {}
        
        satisfaction = str(satisfaction or "æœªçŸ¥")
        data_tree[agent_id][period_key][scenario][satisfaction] = int(cnt or 0)
    
    # ç”Ÿæˆè¡¨æ ¼è¡Œï¼ˆå¤šçº§å±•å¼€ï¼‰
    table_rows = []
    
    # é¢„å®šä¹‰åœºæ™¯å’Œæ»¡æ„åº¦é¡ºåº
    scenario_order = ["å”®å‰", "å”®å", "å”®å‰å’Œå”®å", "å…¶ä»–æ¥å¾…åœºæ™¯", "æ— æ³•åˆ¤å®š", "æœªçŸ¥"]
    satisfaction_order = ["å¤§å¹…å‡å°‘", "å°å¹…å‡å°‘", "æ— æ˜¾è‘—å˜åŒ–", "å°å¹…å¢åŠ ", "å¤§å¹…å¢åŠ ", "æ— æ³•åˆ¤å®š", "æœªçŸ¥"]
    
    for agent_id in sorted(data_tree.keys()):
        agent_user = user_by_id.get(agent_id)
        agent_name = agent_user.name if agent_user else f"æœªç»‘å®šå®¢æœ({agent_id})"
        
        # Level 0: å®¢æœæ±‡æ€»
        agent_total = sum(
            sum(
                sum(scenario_dict.values())
                for scenario_dict in period_dict.values()
            )
            for period_dict in data_tree[agent_id].values()
        )
        
        table_rows.append({
            "level": 0,
            "group_id": f"agent-{agent_id}",
            "parent_id": None,
            "is_header": True,
            "label": agent_name,
            "count": agent_total,
            "details": "",
        })
        
        # Level 1: æ—¥æœŸ
        for period_key in sorted(data_tree[agent_id].keys()):
            period_dict = data_tree[agent_id][period_key]
            period_total = sum(
                sum(scenario_dict.values())
                for scenario_dict in period_dict.values()
            )
            
            table_rows.append({
                "level": 1,
                "group_id": f"agent-{agent_id}-period-{period_key}",
                "parent_id": f"agent-{agent_id}",
                "is_header": True,
                "label": period_key,
                "count": period_total,
                "details": "",
            })
            
            # Level 2: åœºæ™¯
            for scenario in scenario_order:
                if scenario not in period_dict:
                    continue
                
                scenario_dict = period_dict[scenario]
                scenario_total = sum(scenario_dict.values())
                
                table_rows.append({
                    "level": 2,
                    "group_id": f"agent-{agent_id}-period-{period_key}-scenario-{scenario}",
                    "parent_id": f"agent-{agent_id}-period-{period_key}",
                    "is_header": True,
                    "label": scenario,
                    "count": scenario_total,
                    "details": "",
                })
                
                # Level 3: æ»¡æ„åº¦ï¼ˆæ˜ç»†ï¼‰
                for satisfaction in satisfaction_order:
                    if satisfaction not in scenario_dict:
                        continue
                    
                    cnt = scenario_dict[satisfaction]
                    
                    table_rows.append({
                        "level": 3,
                        "group_id": f"agent-{agent_id}-period-{period_key}-scenario-{scenario}",
                        "parent_id": f"agent-{agent_id}-period-{period_key}-scenario-{scenario}",
                        "is_header": False,
                        "label": satisfaction,
                        "count": cnt,
                        "details": f"å®¢æœ={agent_name}, æ—¥æœŸ={period_key}, åœºæ™¯={scenario}, æ»¡æ„åº¦={satisfaction}",
                        "agent_id": agent_id,
                        "period": period_key,
                        "scenario": scenario,
                        "satisfaction": satisfaction,
                    })
    
    return templates.TemplateResponse(
        "agent_analysis.html",
        {
            "request": request,
            "user": user,
            "table_rows": table_rows,
            "filters": {
                "group_by": group_by_mode,
                "start": start_dt.date().isoformat() if start_dt else "",
                "end": end_dt.date().isoformat() if end_dt else "",
                "platform": platform_f,
                "agent_user_id": str(agent_user_id_int) if agent_user_id_int else "",
            },
            "platform_options": platform_options,
            "agent_options": agent_options,
            "group_by_options": [
                ("day", "æŒ‰å¤©"),
                ("week", "æŒ‰å‘¨"),
                ("month", "æŒ‰æœˆ"),
            ],
        },
    )


@app.get("/api/reports/agent-analysis/conversations")
def get_agent_analysis_conversations(
    request: Request,
    user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
    agent_id: int | None = None,
    period: str | None = None,
    scenario: str | None = None,
    satisfaction: str | None = None,
    platform: str | None = None,
):
    """è·å–æŒ‡å®šæ¡ä»¶çš„å¯¹è¯åˆ—è¡¨ï¼ˆCIDï¼‰"""
    from sqlalchemy import func
    
    if not all([period, scenario, satisfaction]):
        return {"error": "ç¼ºå°‘å¿…éœ€å‚æ•°", "cids": []}
    
    def _parse_date(d: str | None):
        if not d:
            return None
        d = d.strip()
        if not d:
            return None
        try:
            y, m, day = [int(x) for x in d.split("-")]
            return datetime(y, m, day)
        except Exception:
            return None
    
    period_dt = _parse_date(period)
    if not period_dt:
        return {"error": "æ—¥æœŸæ ¼å¼é”™è¯¯", "cids": []}
    
    # Agent role: only see self
    if user.role == Role.agent:
        agent_id = user.id
    
    # è·å–æœ€æ–°åˆ†æ
    latest_subq = (
        select(func.max(ConversationAnalysis.id).label("analysis_id"))
        .group_by(ConversationAnalysis.conversation_id)
    ).subquery()
    
    # å®æ—¶ç»‘å®šå®¢æœ
    binding_user_id = func.coalesce(AgentBinding.user_id, Conversation.agent_user_id)
    
    conv_date = func.coalesce(Conversation.ended_at, Conversation.started_at, Conversation.uploaded_at)
    
    # æŒ‰å¤©æŸ¥è¯¢
    period_start = period_dt
    period_end = period_dt + timedelta(days=1)
    
    platform_f = (platform or "").strip()
    
    stmt = (
        select(Conversation.id)
        .select_from(Conversation)
        .join(ConversationAnalysis, ConversationAnalysis.conversation_id == Conversation.id)
        .join(latest_subq, latest_subq.c.analysis_id == ConversationAnalysis.id)
        .outerjoin(AgentBinding, and_(
            AgentBinding.platform == Conversation.platform,
            AgentBinding.agent_account == Conversation.agent_account
        ))
        .where(conv_date >= period_start)
        .where(conv_date < period_end)
    )
    
    # å¤„ç†åœºæ™¯å’Œæ»¡æ„åº¦ç­›é€‰ï¼ˆåŒ…æ‹¬ç©ºå€¼/"æœªçŸ¥"ï¼‰
    if scenario == "æœªçŸ¥":
        stmt = stmt.where((ConversationAnalysis.reception_scenario == "") | (ConversationAnalysis.reception_scenario == None))
    else:
        stmt = stmt.where(ConversationAnalysis.reception_scenario == scenario)
    
    if satisfaction == "æœªçŸ¥":
        stmt = stmt.where((ConversationAnalysis.satisfaction_change == "") | (ConversationAnalysis.satisfaction_change == None))
    else:
        stmt = stmt.where(ConversationAnalysis.satisfaction_change == satisfaction)
    
    if agent_id:
        stmt = stmt.where(binding_user_id == agent_id)
    if platform_f:
        stmt = stmt.where(Conversation.platform == platform_f)
    
    stmt = stmt.order_by(Conversation.id.desc())
    cids = [row for row in session.exec(stmt).all()]
    
    return {"cids": cids}


# =========================
# Training workflow: focus tag + AI reflection + AI simulation
# =========================


def _pick_default_focus_tag(a: ConversationAnalysis | None) -> str:
    """ç”±äºæ ‡ç­¾å­—æ®µå·²åºŸå¼ƒï¼Œè¿”å›ç©ºå­—ç¬¦ä¸²è®©ç”¨æˆ·æ‰‹åŠ¨å¡«å†™ç„¦ç‚¹æ ‡ç­¾"""
    return ""


@app.post("/tasks/{task_id}/focus")
def set_task_focus_tag(
    request: Request,
    task_id: int,
    focus_negative_tag: str = Form(""),
    user: User = Depends(require_role("admin", "supervisor")),
    session: Session = Depends(get_session),
):
    t = session.get(TrainingTask, task_id)
    if not t:
        return _template(request, "error.html", user=user, message="æ‰¾ä¸åˆ°ä»»åŠ¡")
    t.focus_negative_tag = (focus_negative_tag or "").strip()
    session.add(t)
    session.commit()
    return _redirect(f"/tasks/{t.id}")


@app.post("/tasks/{task_id}/reflection/question")
async def generate_reflection_question(
    request: Request,
    task_id: int,
    user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
):
    t = session.get(TrainingTask, task_id)
    if not t:
        return _template(request, "error.html", user=user, message="æ‰¾ä¸åˆ°ä»»åŠ¡")
    if user.role == Role.agent and t.assigned_to_user_id != user.id:
        return _template(request, "error.html", user=user, message="æ— æƒé™")

    conv = session.get(Conversation, t.conversation_id)
    a = session.exec(
        select(ConversationAnalysis)
        .where(ConversationAnalysis.conversation_id == conv.id)
        .order_by(ConversationAnalysis.id.desc())
    ).first()

    if not t.focus_negative_tag:
        t.focus_negative_tag = _pick_default_focus_tag(a)

    system = {
        "role": "system",
        "content": (
            "ä½ æ˜¯Marllenå“ç‰Œçš„å®¢æœåŸ¹è®­æ•™ç»ƒã€‚ä½ çš„ç›®æ ‡æ˜¯å¸®åŠ©å®¢æœä»ä¸€æ¬¡è´Ÿé¢æ¥å¾…ä¸­å¤ç›˜æˆé•¿ã€‚\n"
            "è¯·åªè¾“å‡ºä¸€å¥æ¸…æ™°çš„â€˜è‡ªæˆ‘æ€è€ƒé¢˜ç›®â€™ï¼Œç”¨äºè®©å®¢æœè§£é‡Šï¼šä¸ºä»€ä¹ˆè¿™æ¬¡ä¼šäº§ç”Ÿè¯¥è´Ÿé¢æ ‡ç­¾ï¼Œä»¥åŠä¸‹æ¬¡åº”è¯¥æ€ä¹ˆåšã€‚\n"
            "é¢˜ç›®è¦è®©è¯­è¨€å’Œé€»è¾‘ä¸€èˆ¬çš„äººä¹Ÿèƒ½å›ç­”ï¼ˆä¸ç»•ã€ä¸è¦å¤šé—®é¢˜ï¼‰ã€‚"
        ),
    }

    context = {
        "role": "user",
        "content": (
            f"è´Ÿé¢æ ‡ç­¾ï¼ˆæœ¬æ¬¡è®­ç»ƒèšç„¦ï¼‰ï¼š{t.focus_negative_tag or 'æœªæŒ‡å®š'}\n\n"
            f"AIè¯Šæ–­æ‘˜è¦ï¼š{(a.day_summary if a else '') or 'æš‚æ— '}\n\n"
            f"è¯„ä»·è§£æï¼š{(a.tag_parsing if a else '') or 'æš‚æ— '}\n\n"
            f"ä¸»ç®¡/ç®¡ç†å‘˜ç•™è¨€ï¼š{t.notes or 'æš‚æ— '}"
        ),
    }

    try:
        q = (await chat_completion([system, context], temperature=0.2)).strip()
    except AIError as e:
        return _template(request, "error.html", user=user, message=str(e))

    t.reflection_question = q
    session.add(t)
    session.commit()
    return _redirect(f"/tasks/{t.id}#reflection")


@app.post("/tasks/{task_id}/reflection/submit")
async def submit_reflection(
    request: Request,
    task_id: int,
    answer: str = Form(...),
    user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
):
    t = session.get(TrainingTask, task_id)
    if not t:
        return _template(request, "error.html", user=user, message="æ‰¾ä¸åˆ°ä»»åŠ¡")
    if user.role == Role.agent and t.assigned_to_user_id != user.id:
        return _template(request, "error.html", user=user, message="æ— æƒé™")

    question = (t.reflection_question or "").strip()
    if not question:
        return _template(request, "error.html", user=user, message="è¯·å…ˆç”Ÿæˆè‡ªæˆ‘æ€è€ƒé¢˜ç›®")

    system = {
        "role": "system",
        "content": (
            "ä½ æ˜¯Marllenå“ç‰Œçš„å®¢æœåŸ¹è®­æ•™ç»ƒã€‚\n"
            "è¯·å®¡æ ¸å®¢æœçš„è‡ªæˆ‘æ€è€ƒå›ç­”æ˜¯å¦è¿‡å…³ï¼šæ˜¯å¦èƒ½è¯´æ˜åŸå› ã€å¤ç›˜é”™è¯¯ç‚¹ã€ç»™å‡ºå¯æ‰§è¡Œçš„æ”¹è¿›åšæ³•ä¸è¯æœ¯æ€è·¯ã€‚\n"
            "åªè¾“å‡ºJSONï¼Œæ ¼å¼ï¼š{\"passed\":true/false,\"score\":0-100,\"feedback\":\"...\"}ã€‚"
        ),
    }
    content = {
        "role": "user",
        "content": f"é¢˜ç›®ï¼š{question}\n\nå®¢æœå›ç­”ï¼š{answer.strip()}",
    }

    raw = ""
    try:
        raw = (await chat_completion([system, content], temperature=0.2)).strip()
    except AIError as e:
        return _template(request, "error.html", user=user, message=str(e))

    import json
    passed = None
    score = None
    feedback = ""

    try:
        # best-effort parse: find first {...}
        jtxt = raw
        if "{" in raw and "}" in raw:
            jtxt = raw[raw.find("{") : raw.rfind("}") + 1]
        obj = json.loads(jtxt)
        passed = bool(obj.get("passed"))
        score = int(obj.get("score")) if obj.get("score") is not None else None
        feedback = str(obj.get("feedback") or "")
    except Exception:
        # fallback: treat whole text as feedback
        feedback = raw

    r = TrainingReflection(
        task_id=t.id,
        user_id=user.id,
        question=question,
        answer=answer.strip(),
        ai_passed=passed,
        ai_score=score,
        ai_feedback=feedback,
    )
    session.add(r)

    # advance status a bit
    if t.status in [TaskStatus.open]:
        t.status = TaskStatus.practicing
        session.add(t)

    session.commit()
    return _redirect(f"/tasks/{t.id}#reflection")


@app.post("/tasks/{task_id}/simulate/start")
def start_simulation(
    request: Request,
    task_id: int,
    user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
):
    t = session.get(TrainingTask, task_id)
    if not t:
        return _template(request, "error.html", user=user, message="æ‰¾ä¸åˆ°ä»»åŠ¡")
    if user.role == Role.agent and t.assigned_to_user_id != user.id:
        return _template(request, "error.html", user=user, message="æ— æƒé™")

    focus = (t.focus_negative_tag or "").strip()
    if not focus:
        # attempt pick from latest analysis
        a = session.exec(
            select(ConversationAnalysis)
            .where(ConversationAnalysis.conversation_id == t.conversation_id)
            .order_by(ConversationAnalysis.id.desc())
        ).first()
        focus = _pick_default_focus_tag(a)
        t.focus_negative_tag = focus
        session.add(t)

    # one simulation per (task, user) for MVP
    sim = session.exec(
        select(TrainingSimulation).where(TrainingSimulation.task_id == t.id, TrainingSimulation.user_id == user.id)
    ).first()
    if not sim:
        sim = TrainingSimulation(task_id=t.id, user_id=user.id, focus_tag=focus)
        session.add(sim)
        session.commit()

        # seed system message
        session.add(
            TrainingSimulationMessage(
                simulation_id=sim.id,
                role="system",
                content=f"æœ¬æ¬¡æ¨¡æ‹Ÿè®­ç»ƒèšç„¦ï¼š{focus or 'æœªæŒ‡å®š'}ã€‚ä½ å¯ä»¥åƒçœŸå®æ¥å¾…ä¸€æ ·å›å¤é¡¾å®¢ã€‚",
            )
        )
        session.commit()

    return _redirect(f"/tasks/{t.id}#simulate")


@app.post("/tasks/{task_id}/simulate/send")
async def simulate_send(
    request: Request,
    task_id: int,
    message: str = Form(...),
    user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
):
    t = session.get(TrainingTask, task_id)
    if not t:
        return _template(request, "error.html", user=user, message="æ‰¾ä¸åˆ°ä»»åŠ¡")
    if user.role == Role.agent and t.assigned_to_user_id != user.id:
        return _template(request, "error.html", user=user, message="æ— æƒé™")

    sim = session.exec(
        select(TrainingSimulation).where(TrainingSimulation.task_id == t.id, TrainingSimulation.user_id == user.id)
    ).first()
    if not sim:
        # auto-start
        sim = TrainingSimulation(task_id=t.id, user_id=user.id, focus_tag=t.focus_negative_tag)
        session.add(sim)
        session.commit()

    # store user message
    session.add(TrainingSimulationMessage(simulation_id=sim.id, role="user", content=message.strip()))
    session.commit()

    # build recent context
    msgs = session.exec(
        select(TrainingSimulationMessage)
        .where(TrainingSimulationMessage.simulation_id == sim.id)
        .order_by(TrainingSimulationMessage.created_at.asc(), TrainingSimulationMessage.id.asc())
        .limit(40)
    ).all()

    # include AI diagnosis context (latest)
    a = session.exec(
        select(ConversationAnalysis)
        .where(ConversationAnalysis.conversation_id == t.conversation_id)
        .order_by(ConversationAnalysis.id.desc())
    ).first()

    system = {
        "role": "system",
        "content": (
            "ä½ åœ¨åšMarllenå®¢æœæ¨¡æ‹Ÿè®­ç»ƒã€‚ä½ éœ€è¦åŒæ—¶æ‰®æ¼”â€˜é¡¾å®¢â€™ä¸â€˜æ•™ç»ƒâ€™ï¼š\n"
            "- å…ˆä»¥é¡¾å®¢èº«ä»½ç»§ç»­å¯¹è¯ï¼ˆè‡ªç„¶çœŸå®ï¼Œå¯èƒ½è¿½é—®ã€è´¨ç–‘ã€è¡¨è¾¾æƒ…ç»ªï¼‰ã€‚\n"
            "- ç„¶åå¦èµ·ä¸€è¡Œï¼Œä»¥ã€æ•™ç»ƒç‚¹è¯„ã€‘å¼€å¤´ï¼Œç»™å®¢æœä¸€å¥éå¸¸å…·ä½“å¯æ‰§è¡Œçš„æ”¹è¿›å»ºè®®ï¼ˆå«å¯å¤åˆ¶çš„è¯æœ¯ï¼‰ã€‚\n"
            "å½“å‰è®­ç»ƒèšç„¦è´Ÿé¢æ ‡ç­¾ï¼š" + (sim.focus_tag or t.focus_negative_tag or "æœªæŒ‡å®š") + "\n"
            "AIè¯Šæ–­æ‘˜è¦ï¼š" + ((a.day_summary if a else "") or "æš‚æ— ")
        ),
    }

    chat_msgs = [system]
    for m in msgs:
        if m.role == "system":
            continue
        role = "assistant" if m.role == "assistant" else "user"
        chat_msgs.append({"role": role, "content": m.content})

    try:
        reply = (await chat_completion(chat_msgs, temperature=0.35)).strip()
    except AIError as e:
        return _template(request, "error.html", user=user, message=str(e))

    session.add(TrainingSimulationMessage(simulation_id=sim.id, role="assistant", content=reply))

    if t.status in [TaskStatus.open]:
        t.status = TaskStatus.practicing
        session.add(t)

    session.commit()
    return _redirect(f"/tasks/{t.id}#simulate")
